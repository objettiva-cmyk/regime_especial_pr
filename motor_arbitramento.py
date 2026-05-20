# -*- coding: utf-8 -*-
"""
Arbitramento de Custo - RIR/2018 Art. 308
Critério: 70% do MAIOR preço de venda por Produto/NCM/Competência.

Versao 1.3.8 - 2026-05-18  (Layout: capitalização de colunas, INTERVENCAO_ANALISTA centralizada, ALTERA_CALCULO centralizada, CAPA quebra de texto + justificado, remoção DESCRICAO_PENDENCIA, comentários de cabeçalho, TIPO_PRODUTO comentário INV)

Alterações incorporadas nesta versão:

* **F-01 - Remoção de ARQUIVO_AJUSTE e LINK_ABRIR_AJUSTE de 03_INVENTARIO_VALORIZADO:**
  Para todos os casos que geram INTERVENCAO_ANALISTA=SIM na aba 03, já existe linha
  correspondente em 04_PENDENCIAS com informação mais completa. As colunas
  INTERVENCAO_ANALISTA e ACAO_ANALISTA foram mantidas como indicadores informativos.
  O link de ação fica centralizado na CAPA e na aba 04.

* **F-02 (GAP-01) - PRECO_DO_PROPRIO_PRODUTO_EM_MES_ANTERIOR gera pendência em 04_PEND:**
  Anteriormente gerava INTERVENCAO=SIM em 03_INV mas não criava linha em 04_PENDENCIAS
  nem no TEMPLATE_AJUSTES. Agora gera PRECO_USADO_DE_MES_ANTERIOR_REQUER_REVISAO
  com NIVEL=REVISAR e TIPO_AJUSTE_SUGERIDO=APROVAR_FALLBACK.

* **F-03 (GAP-02) - ALERTA_NCM_DIVERGENTE_INVENTARIO gera pendência em 04_PEND:**
  Divergência de NCM gerava INTERVENCAO=SIM em 03_INV sem linha em 04_PENDENCIAS.
  Agora gera ALERTA_NCM_DIVERGENTE_REQUER_VERIFICACAO com NIVEL=CRITICO e
  TIPO_AJUSTE_SUGERIDO=CORRIGIR_CADASTRO_PRODUTO.

* **F-04 - VALOR_ATUAL preenchido no TEMPLATE_AJUSTES:**
  Antes sempre vazio. Agora preenchido com MAIOR_PRECO_UNITARIO (pendências de
  ARBITRAMENTO) ou CUSTO_ARBITRADO_70 (pendências de INVENTARIO).

* **F-05 - Mensagem de orientação nos dropdowns do template:**
  Adicionadas mensagens de input_title/input_message nas validações de TIPO_AJUSTE
  e STATUS_AJUSTE do TEMPLATE_AJUSTES.

* **F-06 - Passo-a-passo expandido na CAPA:**
  Instrução numerada em 3 passos na CENTRAL DO ANALISTA, com alerta explícito
  de que as abas de output são somente leitura.

* Versão anterior: 1.3.5 (cobertura de pendências e usabilidade do analista)

* **F-07 - Fluxo operacional do analista corrigido:**
  A CAPA passa a orientar abertura do arquivo operacional input/ajustes/AJUSTES_ANALISTA_RIR70.xlsx.
  O template em output fica somente como cópia auditável, eliminando a armadilha de Ctrl+S no arquivo errado.

* **F-08 - Nível ATENCAO consolidado em REVISAR:**
  Pendências de preço usado de mês anterior passam a ser contadas em TOTAL_REVISAR/TOTAL_ACIONAVEL.

* **F-09 - CFOP de transferência para destino externo:**
  CFOP_TRANSFERENCIA_DESTINO_EXTERNO passa a sugerir REVISAR_CLASSIFICACAO_OPERACAO, não APROVAR_ALERTA_PRECO.

* **F-10 - Critério de custeio RIR70 explícito:**
  A matriz pode trazer CRITERIO_CUSTEIO_RIR70. PA e MR vinculado ao grupo permanecem em RIR70_70_MAIOR_PRECO; MR de terceiros deve usar CUSTO_AQUISICAO e fica bloqueado/revisão até existir rotina própria de custo de aquisição.

* **F-11 - Template de ajustes mais claro:**
  Incluídos ALTERA_CALCULO, VALOR_REFERENCIA_INFORMADO, CAMINHO_EVIDENCIA, JUSTIFICATIVA_COMPARABILIDADE e RESPONSAVEL_VALIDACAO_FISCAL.
"""
from __future__ import annotations

from pathlib import Path
from datetime import datetime, date, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP, getcontext
from collections import defaultdict, Counter
import calendar, hashlib, gc, json, logging, math, os, re, socket, sys
import unicodedata, xml.etree.ElementTree as ET, io, zipfile

try:
    import pandas as pd
    import xlsxwriter
except Exception as exc:
    print("ERRO: execute o .bat para instalar pandas, xlsxwriter e dependências.")
    sys.exit(1)

try:
    import python_calamine
    HAS_CALAMINE = True
except Exception:
    HAS_CALAMINE = False

getcontext().prec = 28
BASE_DIR    = Path(__file__).resolve().parent
INPUT_DIR   = BASE_DIR / "input"
OUTPUT_DIR  = BASE_DIR / "output"
OUTPUT_DIR.mkdir(exist_ok=True)
CONFIG_PATH = BASE_DIR / "config_arbitramento_rir70.json"
RUN_TS      = datetime.now()
SCRIPT_VER  = "1.3.8"
VERSAO_CRITERIO_PRECO = "RIR70_PRECO_BRUTO_COM_ICMS_v1"
OUT_BASE    = f"Arbitramento_Custo_RIR70_COMPETENCIA_{RUN_TS:%Y%m%d_%H%M%S}"
OUTPUT_PATH = OUTPUT_DIR / f"{OUT_BASE}.xlsx"
LOG_PATH    = OUTPUT_DIR / f"{OUT_BASE}.log"
META_PATH   = OUTPUT_DIR / f"{OUT_BASE}.metadados.json"
_OUTPUT_INITIALIZED = False
HEADER_SCAN = 40
CENT2 = Decimal("0.01"); CENT3 = Decimal("0.001"); CENT4 = Decimal("0.0001"); ZERO = Decimal("0")

class ConfigError(ValueError):
    """Erro de configuração que deve interromper o motor antes do processamento."""
    pass

# Escopo fechado da aba 03_INVENTARIO_PA solicitado para o inventário de PA.
# A aba 03 só deve exibir/calcular produtos com estes NCMs.
NCM_ABA03_PERMITIDOS = [
    "21069010", "21069030", "22021000", "22029900", "22030000",
    "22060090", "22084000", "22085000", "22086000", "22087000", "22089000",
]

# Lista cirúrgica definida para classificação operacional TIPO_ITEM_RIR70.
# A NCM 22030000 foi incluída expressamente na regra RIR70 v6.
NCM_TIPO_ITEM_RIR70 = [
    "21069010", "21069030", "22021000", "22029900", "22030000",
    "22060090", "22084000", "22085000", "22086000", "22087000", "22089000",
]

# ─── Configuração padrão ──────────────────────────────────────────────────────
DEFAULT_CONFIG = {
    "periodo_base_inicio": "2025-01-01",
    "periodo_base_fim":    "2025-12-31",
    # Quando Sim, o motor pergunta o período na abertura do BAT/terminal.
    # Em execução não interativa, mantém o período do config para não quebrar automações.
    "perguntar_periodo_ao_iniciar": "Sim",
    # Valor fiscal de referência: por determinação operacional deste projeto,
    # usa-se o valor bruto do produto e deduz-se APENAS o IPI.
    # ICMS NÃO é deduzido, em aderência ao RIR/2018 art. 308, par.1º.
    "tratamento_ipi_valor_produto": "vprod_ja_sem_ipi",
    "deduzir_ipi_do_valor_produto_bruto": False,
    "deduzir_icms_do_valor_produto": False,
    "excluir_icms_st_do_preco_base": True,
    "excluir_fcp_st_do_preco_base": True,
    # Datas mínimas de movimentação por CNPJ emitente: registros anteriores à data são ignorados.
    # Exemplo: {"03408722000763": "01/11/2026"}
    "data_inicial_por_cnpj": {},
    "bloquear_calculo_sem_coluna_ipi": False,
    "competencia_prioritaria": "data_emissao",
    "permitir_fallback_competencia_data_emissao": True,
    "excluir_intercompany_do_calculo": True,
    "cnpjs_grupo": ["03408722000178","03408722000330","03408722000410","03408722000763"],
    "CFOP_VENDA_CALCULO": ["5101","5102","5401","5403","5405","6101","6102","6401","6403","6404"],
    "CFOP_REVISAR": ["5110","6110","5116","5117","6116","6117","5404","6405"],
    "CFOP_NAO_CALCULA": ["5151","6151","5152","6152","5901","6901","5910","6910","5921","6921","5124","6124","5915","6915","5556","6556","5922","6922","5949","6949","7949","1201","2201","1411","2411"],
    "status_documento_bloqueantes": [
        "CANCELADA","CANCELADO","CANC","INUTILIZADA","INUTILIZADO",
        "EXTEMPORANEA","EXTEMPORANEO","DENEGADA","DENEGADO","NAO_CONFIRMADA"
    ],
    "fator_unidade_maximo": 500,
    "limiar_outlier_maior_vs_medio": 2.5,
    "excluir_outlier_do_calculo": False,  # Mantido apenas por compatibilidade; cálculo legal nunca exclui outlier automaticamente.
    "media_unitaria_ponderada_por_qtde": True,
    "media_oficial_alerta": "ponderada_por_qtd_calculo",
    "calcular_media_ponderada": True,
    "limiar_alerta_desvio_percentual": 50,
    "limiar_critico_desvio_percentual": 100,
    "limiar_revisao_obrigatoria_percentual": 200,
    "formato_data": "DD/MM/AAAA",
    "gerar_aba_inventario_pa": True,
    "almoxarifados_inventario_pa": ["01","50"],
    "filtrar_almoxarifado_inventario_pa": True,
    "unidades_embalagem_fator_1": ["CX","CXS","CAIXA","FD","FARDO","PCT","PACOTE","PACK"],
    "unidades_base_fator_1": ["UN","UND","UNID","UNIDADE","KG","G","L","LT","ML","M","MT"],
    "usar_subpastas_input": True,
    "pasta_movimentacao": "input/movimento_item",
    "pasta_documentos":   "input/documentos",
    "pasta_inventario":   "input/inventario",
    "pasta_auxiliares":   "input/auxiliares",
    "pasta_xml":          "input/xml",
    "usar_xml_como_movimentacao": True,
    "usar_zip_xml": True,
    "arquivos_movimentacao_excluir_contem": [],
    "arquivos_movimentacao_priorizar_contem": ["DETALHADA","DETALHADO"],
    "cnpj_inventario_padrao": "03408722000410",
    "cnpj_inventario_por_token_arquivo": {},
    "processar_apenas_cnpj_inventario_padrao": False,
    "considerar_todos_emitentes_movimentacao": True,
    "segregar_calculo_por_cnpj_emitente": True,
    "usar_documentos_consolidados_para_enriquecimento": True,
    "permitir_documentos_em_input_raiz": True,
    "permitir_documentos_em_movimento_item": False,
    "bloquear_situacao_documento_numerica_2": True,
    "permitir_movimentacao_em_input_raiz": True,
    "permitir_inventario_em_input_raiz": True,
    "arquivos_inventario_prefixos": ["INVENTARIO","ESTOQUE","SALDO"],
    "arquivos_movimentacao_prefixos": ["BASE","MOVIMENTACAO","MOVIMENTAÇÃO","SAIDAS","SAÍDAS","NFE","NF"],
    "normalizar_codigo_item_para_cruzamento": True,
    "validar_dimensao_qcom_qtrib": True,
    "permitir_fator_manual": True,
    "arquivo_fator_manual": "fatores_manuais.xlsx",
    "gravar_inventario_sem_base_arbitrada": True,
    "permitir_fallback_base_inventario_competencia_anterior": True,
    "priorizar_base_competencia_anterior_mesmo_cnpj": True,
    "converter_inventario_para_unidade_tributavel": True,
    "permitir_fallback_base_inventario_entre_cnpjs": True,
    "destacar_fallback_base_inventario_entre_cnpjs": True,
    "permitir_fallback_base_produto_similar": True,
    "arquivo_mapa_similaridade_produto": "MAPA_SIMILARIDADE_PRODUTO.xlsx",
    "criterio_similaridade_produto": "cadastro_homologado",
    "fallback_limite_inferior_competencia": "01/2025",
    "usar_leitor_xlsx_streaming": True,
    "mapeamento_colunas": {},
    "score_minimo_colunas_fiscais": 5,
    # ----- Parâmetros de melhoria -----
    # Ativa processamento paralelo de arquivos XML. Quando True, o motor usa
    # ThreadPoolExecutor com o número de workers definido em `max_workers_xml`
    # para ler arquivos XML em paralelo, reduzindo o tempo de importação.
    "usar_leitura_xml_paralela": False,
    # Define o número máximo de threads de leitura de XML quando o paralelismo
    # estiver ativado. Ajuste conforme a capacidade de CPU e I/O do servidor.
    "max_workers_xml": 4,
    # Indica se os arquivos XML estão em rede (UNC). Ainda não utilizado,
    # incluído para futura implementação de tempo de espera e reconexão.
    "xml_em_rede": False,
    # RESERVADO - detecção por desvio padrão não implementada nesta versão.
    # Quando implementado, gerará alertas adicionais com base em desvios estatísticos.
    "usar_deteccao_preco_fora_padrao": False,  # reservado - não altera o cálculo atual
    "desvios_padrao_alerta": 3,                # reservado - usado apenas quando usar_deteccao_preco_fora_padrao=True
    # Controla abertura automática do arquivo Excel após a execução do motor.
    "abrir_excel_ao_final": False,
    # Controla se o terminal de comando deve ser fechado automaticamente ao
    # final da execução. Útil para ambientes Windows.
    "fechar_cmd_ao_final": False,
    # NCMs que entram no cálculo de arbitramento (02_ARBITRAMENTO e 03_INVENTARIO_PA).
    # Lista vazia = todos os NCMs. Preencher com os NCMs das saídas sujeitas ao RIR/2018 Art. 308.
    "ncms_arbitramento": [
        "21069010",  # Concentrados/preparações alimentícias quando classificados como produto acabado
        "21069030",  # Suplementos líquidos/produtos acabados - incluído para não suprimir vendas da empresa 03.408.722/0007-63
        "22021000",  # Águas + sucos industrializados
        "22029900",  # Outras bebidas não alcoólicas
        "22030000",  # Cervejas de malte
        "22060090",  # Outras bebidas fermentadas (coquetéis, drinks)
        "22084000",  # Rum e aguardente de cana
        "22085000",  # Gin e genebra
        "22086000",  # Vodca
        "22087000",  # Licores e cremes
        "22089000"   # Outras bebidas destiladas/espirituosas
    ],
    # Escopo específico da aba 03_INVENTARIO_PA.
    # Este filtro é aplicado mesmo quando o inventário contém outros NCMs.
    "ncms_inventario_pa": NCM_ABA03_PERMITIDOS.copy(),
    # Nao = output executivo enxuto para analista. Sim = inclui colunas tecnicas de rastreabilidade.
    "exibir_colunas_diagnostico": "Nao",
    # Central do analista / correções manuais auditáveis
    "pasta_ajustes_analista": "input/ajustes",
    "arquivo_ajustes_analista": "AJUSTES_ANALISTA_RIR70.xlsx",
    "gerar_template_ajustes_analista": True,
    "criar_arquivo_ajustes_analista_se_ausente": True,
    "gerar_links_analista": True,
    "MODO_REPROCESSAMENTO": "COMPLETO",
    "preservar_ajustes_analista_existentes": True,
    "MAX_MESES_RETROATIVOS_PRECO": 12,
    "max_meses_retroativos_preco": 12,
    "usar_precos_referencia": True,
    "arquivo_precos_referencia": "precos_referencia.xlsx",
    "validar_erros_excel_output": "Sim",
    "max_celulas_validacao_output": 0
    ,
    # Impede o cálculo quando a matriz de cadastro não possui valor em
    # CATEGORIA_ITEM. A ausência de CATEGORIA_ITEM compromete a
    # classificação correta de CFOPs de devolução/transferência e deve
    # ser tratada como pendência. Ajuste para False em testes.
    "validar_categoria_item": True,
    # Critério explícito de custeio. Ausência na matriz não bloqueia: PA e MR vinculado ao grupo
    # continuam em RIR70_70_MAIOR_PRECO; MR sem vínculo é bloqueado.
    "validar_criterio_custeio_rir70": True,
    "bloquear_custo_aquisicao_sem_rotina": True,
    "alertar_cnpj_raiz_grupo_nao_cadastrado": True,
    "gerar_diagnostico_matriz_rir70": True
}

# ─── Candidatos de colunas ────────────────────────────────────────────────────
COLUMN_CANDIDATES = {
    "chave":               ["CHAVEACESSO","CHAVE ACESSO","CHAVE ACESSO NFE","CHAVE NF-E","CHAVE NFE","CHAVENFE","CHAVE"],
    "competencia":         ["PERIODOREFERENCIA","PERÍODO REFERÊNCIA","PERIODO REFERENCIA","COMPETENCIA","COMP","MESANO","MES ANO","REFERENCIA"],
    "data_emissao":        ["DATAEMISSAONFE","DATA EMISSAO NFE","DATAEMISSAO","DATA EMISSAO","DT EMISSAO","DTEMISSAO","EMISSAO","DATA","DHEMI","DH EMISSAO","DATA AUTORIZACAO","DATA NF"],
    "numero_nf":           ["NUMERODOCUMENTO","NUMERO DOCUMENTO","NUMERONF","NUMERO NF","NUMERO NFE","NNF","NFS","NF","DOCUMENTO","NUM_DOC"],
    "serie":               ["SERIEDOCUMENTO","SERIE DOCUMENTO","SERIE","SERIE NF","SERIE NFE","SER","SERIE_DOC"],
    "gtin":                ["GTINPRODUTO","GTIN PRODUTO","GTIN","CEAN","EAN","COD BARRA","CODIGO BARRAS","BARRA"],
    "ncm":                 ["NCMPRODUTO","NCM PRODUTO","NCM","CODIGO NCM","NBM","CLAS FISC","CLASSIFICACAO FISCAL"],
    "cfop":                ["CFOPPRODUTO","CFOP PRODUTO","CFOP"],
    "codigo_item":         ["PRODUTO","CODIGO ITEM","CODITEM","COD ITEM","COD_ITEM","CPROD","CODIGO PRODUTO","COD_PRODUTO","IDITEM","ID ITEM"],
    "item_nf":             ["ITEM","NUMITEM","NUM ITEM","ITEM NFE","ITEM NF","NITEM"],
    "descricao":           ["DESCRICAOPRODUTO","DESCRICAO PRODUTO","DESCRICAO COMPLETA","DESCRICAOCOMPLETA","DESC PRODUTO","DESC_TECNICA","DESCRICAO","XPROD"],
    "unid_comercial":      ["UNIDADECOMERCIAL","UNIDADE COMERCIAL","UNID. COMERCIAL","UNID COMERCIAL","UCOM","UNIDADE","UN","UMP"],
    "qtd_comercial":       ["QUANTIDADEUNIDADECOMERCIAL","QUANTIDADE UNIDADE COMERCIAL","QTD COMERCIAL","QUANTIDADE COMERCIAL","QCOM","QTDE","QUANTIDADE","QTD"],
    "valor_unitario":      ["VALORUNITARIOCOMERCIAL","VALOR UNITARIO COMERCIAL","VALOR UNITARIO","VLR UNITARIO","VUNCOM","PRECO UNIT","PRECO UNITARIO"],
    "valor_unitario_tributavel": ["VALORUNITARIOTRIBUTAVEL","VALOR UNITARIO TRIBUTAVEL","VALOR UNIT. TRIBUTAVEL","VALORUNITTRIBUTAVEL","VLR UNITARIO TRIBUTAVEL","VLR UNIT TRIB","VUNTRIB","VALOR UN TRIB","PRECO UNIT TRIB","PRECO UNITARIO TRIBUTAVEL"],
    "valor_total_produto": ["VALORTOTALPRODUTO","VALOR TOTAL PRODUTO","VPROD","VALOR PRODUTO","VLR MERCADORIA","VALOR MERCADORIA","VALOR TOTAL ITEM NFE","VALORTOTALITEMNFE"],
    "valor_ipi":           ["VALORIPI","VALOR IPI","VIPI","IPI","VALORIPIDEVOLUCAO","VALOR IPI DEVOLUCAO"],
    "situacao_documento":  ["SITUACAODOCUMENTO","SITUACAO DOCUMENTO","SITUACAO NFE","SITUACAO NF","STATUS DOCUMENTO","STATUS NFE","STATUS NF","CANCELAMENTO","STATUS CANCELAMENTO","CANCELADA","SITUACAO","STATUS"],
    "cnpj_emitente":       [
        "CNPJDOEMITENTE", "CNPJ DO EMITENTE", "CnpjDoEmitente",
        "CNPJOUCPFDOEMITENTE", "CNPJ OU CPF DO EMITENTE", "CnpjOuCpfDoEmitente",
        "CNPJEMITENTE", "CNPJ EMITENTE", "CNPJCPFEMITENTE", "CNPJ CPF EMITENTE",
        "EMITENTE CNPJ", "CNPJ EMIT", "CNPJ_EMITENTE"
    ],
    "cnpj_destinatario":   [
        "CNPJDODESTINATARIO", "CNPJ DO DESTINATARIO", "CnpjDoDestinatario",
        "CNPJOUCPFDODESTINATARIO", "CNPJ OU CPF DO DESTINATARIO", "CnpjOuCpfDoDestinatario",
        "CNPJDESTINATARIO", "CNPJ DESTINATARIO", "CNPJCPFDESTINATARIO", "CNPJ CPF DESTINATARIO",
        "DESTINATARIO CNPJ", "CNPJ DEST", "CNPJ_DESTINATARIO"
    ],
    "nome_emitente":       ["RAZAOSOCIALEMITENTE","RAZAO SOCIAL EMITENTE","NOMEEMITENTE","NOME EMITENTE","EMITENTE","XNOME EMIT","XNOME EMITENTE","RAZAO SOCIAL DO EMITENTE","EMPRESA EMITENTE","NOMEFORNECEDOR","NOME FORNECEDOR","XNOMEEMIT","XNOME","NOME_EMIT"],
    "nome_destinatario":   ["RAZAOSOCIALDESTINATARIO","RAZAO SOCIAL DESTINATARIO","NOMEDESTINATARIO","NOME DESTINATARIO","DESTINATARIO","XNOME DEST","XNOME DESTINATARIO","RAZAO SOCIAL DO DESTINATARIO","NOMECLIENTE","NOME CLIENTE","XNOMEDEST"],
    "uf_destinatario":     ["UFDESTINATARIO","UF DESTINATARIO","UF_DESTINATARIO","UF DEST","DEST UF","UFDEST","UF DO DESTINATARIO","UF","UFDESTINO","UF DESTINO"],
    "cest":                ["CESTPRODUTO","CEST PRODUTO","CODIGO CEST","COD CEST","CEST"],
    "qtd_tributavel":      ["QUANTIDADEUNIDADETRIBUTAVEL","QUANTIDADE UNIDADE TRIBUTAVEL","QTD TRIBUTAVEL","QUANTIDADE TRIBUTAVEL","QTRIB","QTDE TRIB","QTDE TRIBUTAVEL"],
}

INVENTORY_QTY_TOKENS = frozenset({
    "QTDE","QTD","QUANT","QUANTIDADE","QATU",
    "CFQTDE","CFQUANT","CFQATU","QTDEATUAL","QUANTATUAL","QUANTIDADEATUAL",
    "SALDOQTDE","SALDOQUANTIDADE","QTDEFISICA","QTDPHYSICAL",
})
INVENTORY_FINANCIAL_BLOCKED = frozenset({
    "VLR","VALOR","PRECO","PREC","CUSTO","CMV","FINANC","VUNIT",
    "VEND","MEDIO","TOTAL","ESTOQUE","MONTANTE","IMPORTANCIA",
})
INVENTORY_CANDIDATES = {
    "codigo_item":    ["COD_ITEM","CODIGO_ITEM","CODITEM","COD ITEM","CODIGO ITEM","CPROD","PRODUTO","COD_PRODUTO","CODIGO PRODUTO","REFERENCIA","SKU"],
    "qtde_inventario":["CF_QTDE","CF_QUANT","CF_QATU","CFQTDE","CFQUANT","CFQATU","QTDE ATUAL","QTDEATUAL","QUANTIDADE ATUAL","QUANTIDADEATUAL","SALDO QTDE","SALDOQTDE","QTDE FISICA","QTD FISICA","QUANTIDADE FISICA","QTDE","QTD","QUANTIDADE"],
    "descricao":      ["DESCRICAO","DESCRICAO PRODUTO","DESC PRODUTO","DESC","NOME PRODUTO","XPROD","DESCRICAO_ITEM"],
    "unid_comercial": ["UNIDADE","UNID","UN","UND","UNIDADE COMERCIAL","UNID COMERCIAL","UCOM","UNIDADE_MEDIDA"],
    "ncm":            ["NCM","CODIGO NCM","NCM PRODUTO","NCMPRODUTO","COD NCM","CLASSIF FISCAL","NBM"],
    "cod_almox":      ["COD_ALMOX","COD ALMOX","CODIGO ALMOXARIFADO","ALMOXARIFADO","ALMOX","DEPOSITO","LOCAL","ARMAZEM","CENTRO_CUSTO","CENTRO CUSTO"],
}

logger = logging.getLogger("arbitramento_rir70")
MAPPING_AUDIT_ROWS: list = []

# ─── Utilitários ──────────────────────────────────────────────────────────────
def clean_str(v) -> str:
    if v is None: return ""
    try:
        if isinstance(v, float) and math.isnan(v): return ""
    except Exception: pass
    t = str(v).strip()
    return "" if t.lower() in {"nan","nat","none"} else t

def norm_text(v) -> str:
    t = unicodedata.normalize("NFKD", clean_str(v)).encode("ascii","ignore").decode("ascii")
    return re.sub(r"[^A-Z0-9]","",t.upper())

def only_digits(v) -> str: return re.sub(r"\D","",clean_str(v))

def normalize_code_text(v) -> str:
    t = clean_str(v)
    return t[:-2] if re.fullmatch(r"\d+\.0",t) else t

def normalize_chave(v) -> str:
    d = only_digits(v); return d if len(d)==44 else clean_str(v)

def normalize_cnpj(v) -> str:
    d = only_digits(v); return d[:14] if len(d)>=14 else d

def normalize_ncm(v) -> str:
    d = only_digits(v); return d[:8] if len(d)>=8 else d

def normalize_cfop(v) -> str:
    d = only_digits(v); return d[:4] if len(d)>=4 else d

def normalize_cest(v) -> str:
    d = only_digits(v); return d[:7].zfill(7) if d else ""

def normalize_uf(v) -> str:
    t = re.sub(r"[^A-Za-z]","",clean_str(v)).upper(); return t[:2] if len(t)>=2 else t

def to_decimal(v):
    if v is None: return None
    if isinstance(v, Decimal): return v
    if isinstance(v,(int,float)) and not isinstance(v,bool):
        if isinstance(v,float) and math.isnan(v): return None
        return Decimal(str(v))
    t = clean_str(v).replace("R$","").replace("%","").replace("\u00a0","").strip()
    t = re.sub(r"[^0-9,\.\-]","",t)
    if not t or t in {"-",",","."}: return None
    if "," in t and "." in t:
        t = t.replace(".","").replace(",",".") if t.rfind(",")>t.rfind(".") else t.replace(",","")
    elif "," in t:
        t = t.replace(".","").replace(",",".")
    try: return Decimal(t)
    except InvalidOperation: return None

def q4(v): return v.quantize(CENT4, rounding=ROUND_HALF_UP) if isinstance(v,Decimal) else None
def q3(v): return v.quantize(CENT3, rounding=ROUND_HALF_UP) if isinstance(v,Decimal) else None
def q2(v): return v.quantize(CENT2, rounding=ROUND_HALF_UP) if isinstance(v,Decimal) else None

def _json_default(obj):
    """Conversor seguro para logs/metadados JSON.

    O config usa frozenset/set pre-computados para performance; json.dumps puro
    quebra nesses objetos. Esta função mantém o log auditável sem remover as
    otimizações internas do motor.
    """
    if isinstance(obj, (set, frozenset)):
        return sorted(obj)
    if isinstance(obj, Decimal):
        return str(obj)
    if isinstance(obj, (datetime, date)):
        return obj.isoformat()
    if isinstance(obj, Path):
        return str(obj)
    return str(obj)

def safe_json_dumps(obj, **kwargs):
    kwargs.setdefault("ensure_ascii", False)
    kwargs.setdefault("default", _json_default)
    return json.dumps(obj, **kwargs)

def parse_date(v, preferred=None):
    if isinstance(v, datetime): return v.replace(tzinfo=None)
    if isinstance(v, date):     return datetime(v.year,v.month,v.day)
    t = clean_str(v)
    if not t: return None
    if re.fullmatch(r"\d+\.0",t): t = t[:-2]
    fmap = {"DD/MM/AAAA":"%d/%m/%Y","DD-MM-AAAA":"%d-%m-%Y","AAAA-MM-DD":"%Y-%m-%d","MM/DD/AAAA":"%m/%d/%Y"}
    pref = fmap.get(clean_str(preferred).upper())
    fmts = [pref] if pref else []
    fmts += [f for f in ["%d/%m/%Y","%Y-%m-%d","%d-%m-%Y","%Y%m%d","%d/%m/%y"] if f not in fmts]
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}.*",t):
        try: return datetime.fromisoformat(t[:19]).replace(tzinfo=None)
        except Exception:
            pass
    for f in fmts:
        try: return datetime.strptime(t[:10], f)
        except Exception:
            pass
    num = to_decimal(v)
    if num is not None and Decimal("1")<=num<=Decimal("90000"):
        return datetime(1899,12,30)+timedelta(days=int(num))
    return None

def parse_config_date(v):
    d = parse_date(v); return datetime(d.year,d.month,d.day) if d else None

def format_date_abnt(v) -> str:
    d = parse_date(v)
    return f"{d.day:02d}/{d.month:02d}/{d.year:04d}" if d else clean_str(v)

def _parse_competencia_str(v):
    """Interpreta uma string de competência retornando "MM/AAAA" ou "N/D"."""
    t = clean_str(v)
    d = only_digits(t)
    # Formato MMyyyy ou mmyyyy
    if len(d) == 6:
        # mmYYYY
        if 1 <= int(d[:2]) <= 12:
            return f"{d[:2]}/{d[2:6]}"
        # YYYYmm
        if 1 <= int(d[4:6]) <= 12:
            return f"{d[4:6]}/{d[:4]}"
    # Formato DDMMYYYY ou YYYYMMDD
    if len(d) == 8:
        # ddMMYYYY
        if 1 <= int(d[2:4]) <= 12 and d[4:8].startswith(("19", "20")):
            return f"{d[2:4]}/{d[4:8]}"
        # YYYYMMDD
        if d[:4].startswith(("19", "20")) and 1 <= int(d[4:6]) <= 12:
            return f"{d[4:6]}/{d[:4]}"
    m = re.search(r"(0?[1-9]|1[0-2])\D+(20\d{2}|19\d{2})", t)
    if m:
        return f"{int(m.group(1)):02d}/{m.group(2)}"
    return "N/D"

def parse_competencia(v, data_emissao, config=None):
    """
    Converte um valor de competência (mês/ano) para o formato "MM/AAAA".

    Este utilitário aceita opcionalmente um objeto de configuração para
    obedecer às chaves ``competencia_prioritaria`` e
    ``permitir_fallback_competencia_data_emissao``.  Quando
    ``competencia_prioritaria`` for ``"competencia"`` (ignorando
    maiúsculas), o campo de entrada ``v`` terá precedência; quando for
    ``"data_emissao"`` ou ausente, a data de emissão será usada caso
    exista.  Se a prioridade for "competencia" e o valor informado
    estiver ausente ou não puder ser convertido, o fallback para a data
    de emissão somente ocorrerá se
    ``permitir_fallback_competencia_data_emissao`` for verdadeiro.  Sem
    fallback, será retornado "N/D".
    """
    def _from_emissao(de):
        if isinstance(de, datetime):
            return f"{de.month:02d}/{de.year}"
        return None
    # Se config não fornecido, mantém comportamento anterior: data_emissao
    # prevalece sobre valor informado, porque a estrutura de NF-e sempre
    # traz a data de emissão e muitas planilhas deixam a competência vazia.
    if config:
        pri = norm_text(config.get("competencia_prioritaria", "data_emissao"))
        fallback = bool(config.get("permitir_fallback_competencia_data_emissao", True))
        # Prioriza a competência informada
        if pri == "competencia":
            val = _parse_competencia_str(v)
            if val != "N/D":
                return val
            # valor ausente ou inválido
            if fallback:
                from_emissao = _from_emissao(data_emissao)
                return from_emissao or "N/D"
            return "N/D"
        # Caso contrário, prioriza data de emissão
        from_emissao = _from_emissao(data_emissao)
        if from_emissao:
            return from_emissao
        # se não houver data de emissão, tenta parsear valor informado
        return _parse_competencia_str(v)
    # Sem configuração: comportamento legada: data de emissão prevalece
    if isinstance(data_emissao, datetime):
        return f"{data_emissao.month:02d}/{data_emissao.year}"
    return _parse_competencia_str(v)

def competencia_sort_key(v) -> int:
    m = re.match(r"^(\d{2})/(\d{4})$",clean_str(v))
    return int(m.group(2))*100+int(m.group(1)) if m else 999999

def competencia_month_index(v):
    m = re.match(r"^(\d{2})/(\d{4})$", clean_str(v))
    if not m:
        return None
    mes = int(m.group(1)); ano = int(m.group(2))
    if not 1 <= mes <= 12:
        return None
    return ano * 12 + mes

def meses_entre_competencias(origem, destino):
    i0 = competencia_month_index(origem)
    i1 = competencia_month_index(destino)
    if i0 is None or i1 is None:
        return None
    return i1 - i0

def _fallback_mes_dentro_limite(origem, destino, config):
    diff = meses_entre_competencias(origem, destino)
    if diff is None or diff < 0:
        return False
    if diff == 0:
        return True
    try:
        limite = int(config.get("MAX_MESES_RETROATIVOS_PRECO") or config.get("max_meses_retroativos_preco") or 12)
    except Exception:
        limite = 12
    return diff <= max(0, limite)

def make_unique_headers(headers):
    seen = defaultdict(int); result = []
    for h in headers:
        b = clean_str(h) or "COLUNA_SEM_NOME"
        seen[b] += 1
        result.append(b if seen[b]==1 else f"{b}_{seen[b]}")
    return result

def normalize_item_join_key(v) -> str:
    raw = normalize_code_text(v); txt = clean_str(raw).strip()
    if not txt: return ""
    tn = re.sub(r"\s+", "", txt.upper())
    # Senior/Focco/SAP: códigos podem vir com sufixos operacionais usados
    # para diferenciar origem/processo, mas que não devem quebrar o cruzamento
    # entre movimentação e inventário. Ex.: 0103558-F, 0103558FC,
    # 0103558-PA, 0103558-IND -> 103558.
    m_suffix = re.fullmatch(r"0*(\d+)(?:[-_]?)(F|FC|PA|IND)$", tn)
    if m_suffix:
        try:
            return str(int(m_suffix.group(1)))
        except Exception:
            return m_suffix.group(1).lstrip("0") or "0"
    if re.fullmatch(r"\d+", tn):
        try: return str(int(tn))
        except Exception: return tn.lstrip("0") or "0"
    return norm_text(txt)

def codigo_produto_key(v) -> str:
    """Chave operacional do RIR70: preserva sufixos como -F, pois PA e MR têm regra fiscal distinta."""
    t = normalize_code_text(v).strip().upper()
    return re.sub(r"\s+", "", t)

def normalize_almox(v) -> str:
    t = clean_str(v)
    if not t: return ""
    num = to_decimal(t)
    if num is not None:
        try:
            if num == num.to_integral_value():
                return str(int(num)).zfill(2)
        except Exception:
            pass
    d = only_digits(t)
    if not d: return ""
    try:
        return str(int(d)).zfill(2)
    except Exception:
        return d.zfill(2)

def _looks_like_uf(t) -> bool:
    return bool(re.fullmatch(r"[A-Z]{2}",clean_str(t).strip().upper()))

def is_financial_inventory_header(h:str)->bool:
    h = norm_text(h)
    if any(q in h for q in INVENTORY_QTY_TOKENS): return False
    return any(tag in h for tag in INVENTORY_FINANCIAL_BLOCKED)

def is_physical_quantity_header(h:str)->bool:
    h = norm_text(h)
    return bool(h) and any(tok in h for tok in INVENTORY_QTY_TOKENS) and not is_financial_inventory_header(h)

def validate_cnpj(cnpj:str)->bool:
    d = only_digits(cnpj)
    if len(d)!=14 or len(set(d))==1: return False
    w1=[5,4,3,2,9,8,7,6,5,4,3,2]
    r1=sum(int(x)*w for x,w in zip(d[:12],w1))%11; d1=0 if r1<2 else 11-r1
    w2=[6,5,4,3,2,9,8,7,6,5,4,3,2]
    r2=sum(int(x)*w for x,w in zip(d[:13],w2))%11; d2=0 if r2<2 else 11-r2
    return int(d[12])==d1 and int(d[13])==d2

def validate_nfe_key(chave) -> str:
    d = only_digits(chave)
    if len(d)!=44: return "NÃO VALIDADO ERRO - FORMATO"
    body=d[:43]; dv=int(d[43])
    ws=[]; w=2
    for _ in range(43): ws.append(w); w=2 if w==9 else w+1
    tot=sum(int(n)*ww for n,ww in zip(reversed(body),ws))
    calc=11-tot%11; calc=0 if calc>=10 else calc
    return "VALIDADO OK" if calc==dv else "NÃO VALIDADO ERRO - DV INCORRETO"

def normalize_document_status(v)->str:
    raw=clean_str(v); norm=norm_text(raw)
    if not norm: return ""
    if norm in {"1","10","AUTORIZADO","AUTORIZADA","AUTORIZADONFE","USOAUTORIZADO"}: return "AUTORIZADA"
    if norm in {"2","20","CANCELADO","CANCELADA","CANC","CANCELAMENTOHOMOLOGADO"}: return "CANCELADA"
    if "CANCEL" in norm or norm.startswith("CANC"): return "CANCELADA"
    if "DENEG" in norm: return "DENEGADA"
    if "INUTIL" in norm: return "INUTILIZADA"
    if "EXTEMPOR" in norm: return "EXTEMPORANEA"
    return raw.upper()

# ─── Mapeamento de colunas ────────────────────────────────────────────────────
def pick_column(headers, logical_name, config_mapping=None):
    config_mapping = config_mapping or {}
    if logical_name in config_mapping:
        wanted = norm_text(config_mapping[logical_name])
        for idx,h in enumerate(headers):
            if norm_text(h)==wanted: return idx
    norm_map = {norm_text(h):idx for idx,h in enumerate(headers)}
    for c in COLUMN_CANDIDATES[logical_name]:
        k = norm_text(c)
        if k in norm_map: return norm_map[k]
    forbidden = {"NF","UF","UN","PR","S","N","NFE","IPI"}
    exact_only = {"ncm","cfop","codigo_item","chave","situacao_documento","cnpj_emitente","cnpj_destinatario"}
    if logical_name in exact_only: return None
    party_tokens = {
        "nome_emitente":    {"EMIT","FORNEC","RAZAO","NOMEEMIT","EMPRESA"},
        "nome_destinatario":{"DEST","CLIENTE","RAZAO","NOMEDEST","COMPRADOR"},
    }
    if logical_name in party_tokens:
        req = party_tokens[logical_name]
        for c in COLUMN_CANDIDATES[logical_name]:
            k = norm_text(c)
            if len(k)<4: continue
            if k in norm_map: return norm_map[k]
            for normalized,idx in norm_map.items():
                if len(normalized)<4: continue
                if not any(t in normalized for t in req): continue
                if k in normalized or normalized in k: return idx
        return None
    for c in COLUMN_CANDIDATES[logical_name]:
        k = norm_text(c)
        if len(k)<4 or k in forbidden: continue
        for normalized,idx in norm_map.items():
            if len(normalized)<4 or normalized in forbidden: continue
            if k in normalized or normalized in k: return idx
    return None

def pick_inventory_column(headers, logical_name):
    nmap = {norm_text(h):idx for idx,h in enumerate(headers) if norm_text(h)}
    def allowed(nh): return is_physical_quantity_header(nh) if logical_name=="qtde_inventario" else True
    for c in INVENTORY_CANDIDATES[logical_name]:
        k=norm_text(c)
        if logical_name=="qtde_inventario" and not is_physical_quantity_header(k): continue
        if k in nmap and allowed(k): return nmap[k]
    for c in INVENTORY_CANDIDATES[logical_name]:
        k=norm_text(c)
        if len(k)<4: continue
        if logical_name=="qtde_inventario" and not is_physical_quantity_header(k): continue
        for nh,idx in nmap.items():
            if len(nh)<4 or not allowed(nh): continue
            if k in nh or nh in k: return idx
    return None

def classify_file_headers(headers):
    """Classifica a aba como ITEM_FISCAL, INVENTARIO, DOCUMENTO_CONSOLIDADO ou IGNORADO.

    A classificação se baseia na presença de colunas mínimas. Em vez de considerar
    apenas alguns nomes específicos, utiliza os candidatos definidos em
    ``COLUMN_CANDIDATES`` para aumentar a robustez. Campos essenciais para
    ITEM_FISCAL são: chave de acesso, NCM, CFOP, quantidade, valor total
    do produto e código do item.  Se apenas código e quantidade de estoque
    estiverem presentes, classifica como INVENTARIO.
    """
    n = {norm_text(h) for h in headers if clean_str(h)}
    def _any(cands):
        return any(norm_text(c) in n for c in cands)
    # Campos essenciais baseados em COLUMN_CANDIDATES
    chave_cands = COLUMN_CANDIDATES.get("chave", [])
    ncm_cands = COLUMN_CANDIDATES.get("ncm", [])
    cfop_cands = COLUMN_CANDIDATES.get("cfop", [])
    qtd_cands = COLUMN_CANDIDATES.get("qtd_comercial", []) + COLUMN_CANDIDATES.get("qtd_tributavel", [])
    vprod_cands = COLUMN_CANDIDATES.get("valor_total_produto", [])
    produto_cands = COLUMN_CANDIDATES.get("codigo_item", [])
    has_chave = _any(chave_cands)
    has_ncm = _any(ncm_cands)
    has_cfop = _any(cfop_cands)
    has_qtd = _any(qtd_cands)
    has_vprod = _any(vprod_cands)
    has_produto = _any(produto_cands)
    has_item = has_chave and has_ncm and has_cfop and has_qtd and has_vprod and has_produto
    # Documento consolidado: chave de acesso e informações de nota sem itens
    has_doc = _any(["CHAVEACESSO"]) and (
        _any(["NOMEEMITENTE","RAZAOSOCIALEMITENTE","NOMEDESTINATARIO"]) or
        _any(["VALORTOTALNOTA"]) or _any(["SITUACAO"]))
    # Inventário: código de produto e quantidade em estoque (planilhas de inventário)
    inv_codigo_cands = ["CODITEM","CODIGOPRODUTO"]
    inv_qtd_cands = ["CFQTDE","CFQUANT","CFQATU","QTDEATUAL","QUANTIDADEATUAL","SALDOQTDE","QTDE","QUANTIDADE","QTD"]
    has_inv = _any(inv_codigo_cands) and _any(inv_qtd_cands)
    if has_inv and not has_item:
        return "INVENTARIO"
    if has_item:
        return "ITEM_FISCAL"
    if has_doc:
        return "DOCUMENTO_CONSOLIDADO"
    return "IGNORADO"

def data_quality_score(preview_df, header_idx, headers, mapping):
    score=Decimal("0")
    sample=preview_df.iloc[header_idx+1:min(header_idx+11,len(preview_df))]
    if sample.empty: return 0
    def col_vals(name):
        idx=mapping.get(name)
        if idx is None or idx>=sample.shape[1]: return []
        return [clean_str(v) for v in sample.iloc[:,idx].tolist() if clean_str(v)]
    chaves=col_vals("chave")
    if chaves: score+=Decimal(str(sum(1 for v in chaves if len(only_digits(v))==44)/len(chaves)))*20
    datas=col_vals("data_emissao")
    if datas: score+=Decimal(str(sum(1 for v in datas if parse_date(v) is not None)/len(datas)))*15
    cfops=col_vals("cfop")
    if cfops: score+=Decimal(str(sum(1 for v in cfops if re.fullmatch(r"\d{4}",only_digits(v)[:4] or ""))/len(cfops)))*10
    vals=col_vals("valor_total_produto")
    if vals: score+=Decimal(str(sum(1 for v in vals if to_decimal(v) is not None)/len(vals)))*10
    return float(score)

def find_header_row(preview_df, config_mapping=None):
    logical_req=["chave","data_emissao","ncm","cfop","codigo_item","descricao","unid_comercial","valor_total_produto"]
    best={"idx":0,"score":-1,"mapped":0,"confidence":0}
    for idx in range(min(len(preview_df),HEADER_SCAN)):
        headers=make_unique_headers(preview_df.iloc[idx].tolist())
        mapping={k:pick_column(headers,k,config_mapping) for k in COLUMN_CANDIDATES}
        mapped=sum(1 for k in logical_req if mapping.get(k) is not None)
        quality=data_quality_score(preview_df,idx,headers,mapping)
        score=mapped*12+quality
        if score>best["score"]:
            best={"idx":int(idx),"score":float(score),"mapped":mapped,"confidence":min(1,score/150)}
    return best

def find_inventory_header(preview_df):
    best={"idx":0,"score":-1}
    for idx in range(min(len(preview_df),HEADER_SCAN)):
        headers=make_unique_headers(preview_df.iloc[idx].tolist())
        score=sum(1 for k in ["codigo_item","qtde_inventario"] if pick_inventory_column(headers,k) is not None)*10
        for b in ["descricao","unid_comercial","ncm","cod_almox"]:
            if pick_inventory_column(headers,b) is not None: score+=1
        if score>best["score"]: best={"idx":idx,"score":score}
    return best

# ─── Leitor XLSX streaming ────────────────────────────────────────────────────
# Global cache for sheet metadata when streaming XLSX files.  In a parallel
# environment this dictionary may be accessed by multiple threads when
# processing multiple XML or XLSX files concurrently.  To avoid race
# conditions when reading and writing to this cache, we protect it with
# a reentrant lock.  See `_xlsx_sheet_paths` and `_xlsx_shared_strings_cached`.
_XLSX_CACHE: dict = {}

# A global lock used to protect access to the `_XLSX_CACHE` dictionary.  Without
# synchronisation, concurrent writes can lead to lost updates or partially
# initialised entries when XML parsing happens in parallel.  The overhead of
# acquiring this lock is negligible compared to the cost of reading and
# parsing XLSX files.
try:
    import threading as _threading
except Exception:
    _threading = None
_XLSX_CACHE_LOCK = _threading.RLock() if _threading else None

def _xlsx_cache_key(p):
    p = Path(p)
    try:
        st = p.stat()
        return (str(p.resolve()), st.st_mtime_ns, st.st_size)
    except Exception:
        return (str(p), 0, 0)

def _xlsx_shared_strings(zf):
    try:
        root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
    except Exception:
        return []
    ns={"a":"http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    return ["".join(t.text or "" for t in si.findall(".//a:t",ns)) for si in root.findall("a:si",ns)]

def _xlsx_col_index(ref:str)->int:
    m=re.match(r"([A-Z]+)",clean_str(ref).upper())
    if not m: return 0
    n=0
    for ch in m.group(1): n=n*26+(ord(ch)-64)
    return n-1

def _xlsx_sheet_paths(file_path):
    key=_xlsx_cache_key(file_path)
    # First attempt to read from cache under lock to avoid race conditions.
    if _XLSX_CACHE_LOCK:
        with _XLSX_CACHE_LOCK:
            cached=_XLSX_CACHE.get(key, {})
            if "sheet_paths" in cached:
                return cached["sheet_paths"]
    else:
        cached=_XLSX_CACHE.get(key, {})
        if "sheet_paths" in cached:
            return cached["sheet_paths"]
    # Not cached: parse workbook.xml to extract sheet names.
    with zipfile.ZipFile(file_path) as zf:
        wb_root=ET.fromstring(zf.read("xl/workbook.xml"))
        rel_root=ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rels={r.attrib.get("Id"):r.attrib.get("Target") for r in rel_root}
        ns={"a":"http://schemas.openxmlformats.org/spreadsheetml/2006/main",
            "r":"http://schemas.openxmlformats.org/officeDocument/2006/relationships"}
        result=[]
        for sheet in wb_root.find("a:sheets",ns):
            name=sheet.attrib.get("name","")
            rid=sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            target=rels.get(rid,"")
            if not target:
                continue
            path=target.lstrip("/")
            if not path.startswith("xl/"): path="xl/"+path
            if path not in zf.namelist():
                path="xl/worksheets/"+Path(path).name
            result.append((name,path))
    # Update the cache under the lock to ensure consistency.
    if _XLSX_CACHE_LOCK:
        with _XLSX_CACHE_LOCK:
            _XLSX_CACHE.setdefault(key, {})["sheet_paths"] = result
    else:
        _XLSX_CACHE.setdefault(key, {})["sheet_paths"] = result
    return result

def _xlsx_shared_strings_cached(file_path):
    key=_xlsx_cache_key(file_path)
    if _XLSX_CACHE_LOCK:
        with _XLSX_CACHE_LOCK:
            cached=_XLSX_CACHE.get(key, {})
            if "shared" in cached:
                return cached["shared"]
    else:
        cached=_XLSX_CACHE.get(key, {})
        if "shared" in cached:
            return cached["shared"]
    with zipfile.ZipFile(file_path) as zf:
        shared=_xlsx_shared_strings(zf)
    if _XLSX_CACHE_LOCK:
        with _XLSX_CACHE_LOCK:
            _XLSX_CACHE.setdefault(key, {})["shared"] = shared
    else:
        _XLSX_CACHE.setdefault(key, {})["shared"] = shared
    return shared

def xlsx_sheet_names(file_path): return [n for n,_ in _xlsx_sheet_paths(file_path)]

def iter_xlsx_rows_stream(file_path, sheet_name=None, max_rows=None, max_cols=None):
    selected=[(n,p) for n,p in _xlsx_sheet_paths(file_path) if sheet_name is None or n==sheet_name]
    if not selected: return
    _,sheet_path=selected[0]
    shared=_xlsx_shared_strings_cached(file_path)
    with zipfile.ZipFile(file_path) as zf:
        with zf.open(sheet_path) as fh:
            yielded=0
            for event,elem in ET.iterparse(fh,events=("end",)):
                if not elem.tag.endswith("}row"):
                    elem.clear()  # libera nós não-row imediatamente para não acumular DOM
                    continue
                row_values=[]
                for c in elem:
                    if not c.tag.endswith("}c"): continue
                    cidx=_xlsx_col_index(c.attrib.get("r","A1"))
                    if max_cols is not None and cidx>=max_cols: continue
                    t=c.attrib.get("t",""); v=None
                    if t=="inlineStr":
                        v="".join(x.text or "" for x in c.findall(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t"))
                    else:
                        vn=c.find("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v")
                        if vn is not None: v=vn.text
                    if v is None: value=""
                    elif t=="s":
                        try:
                            si = int(v)
                            value = shared[si] if 0 <= si < len(shared) else ""
                        except Exception:
                            value = ""
                    elif t=="b": value="1" if clean_str(v)=="1" else "0"
                    else: value=v
                    while len(row_values)<=cidx: row_values.append("")
                    row_values[cidx]=value
                yielded+=1; yield row_values; elem.clear()
                if max_rows is not None and yielded>=max_rows: break

# ─── Utilitários de arquivo ────────────────────────────────────────────────────
def excel_engine(p):
    s=p.suffix.lower()
    if HAS_CALAMINE and s in {".xlsx",".xlsm",".xls",".xlsb"}: return "calamine"
    return "xlrd" if s==".xls" else "openpyxl"

def _excel_files_in_dir(d):
    p=Path(d) if Path(d).is_absolute() else BASE_DIR/d
    if not p.exists() or not p.is_dir(): return []
    return sorted([f for f in p.iterdir() if f.suffix.lower() in {".xlsx",".xlsm",".xls"} and not f.name.startswith("~$")])

def _is_manual_factor_file(p,config=None):
    cfg=config or DEFAULT_CONFIG
    return p.name.lower()==clean_str(cfg.get("arquivo_fator_manual","fatores_manuais.xlsx")).lower()

def _filename_contains_any(p,fragments):
    name=norm_text(p.stem)
    return any(norm_text(f) and norm_text(f) in name for f in fragments if norm_text(f))

def _filename_matches_prefix(p,prefixes):
    name=norm_text(p.stem)
    return any(name.startswith(norm_text(pf)) or norm_text(pf) in name for pf in prefixes if norm_text(pf))

def _is_excluded_movement_file(p,config):
    return _filename_contains_any(p,config.get("arquivos_movimentacao_excluir_contem",[]))


def ensure_runtime_folders(config):
    INPUT_DIR.mkdir(exist_ok=True); OUTPUT_DIR.mkdir(exist_ok=True)
    folder_keys = ["pasta_movimentacao", "pasta_documentos", "pasta_inventario", "pasta_auxiliares", "pasta_xml", "pasta_ajustes_analista", "pasta_evidencias"]
    if config.get("usar_subpastas_input", True):
        for key in folder_keys:
            folder = Path(config.get(key, ""))
            if clean_str(folder):
                if not folder.is_absolute():
                    folder = BASE_DIR / folder
                folder.mkdir(parents=True, exist_ok=True)


def validar_governanca_operacional(config):
    """Valida coerencia minima entre script, config e comando congelado."""
    esperado = clean_str(config.get("versao_script_esperada"))
    if esperado and esperado != SCRIPT_VER:
        raise ConfigError(f"versao_script_esperada={esperado} diverge do SCRIPT_VER={SCRIPT_VER}.")
    regra = clean_str(config.get("regra_motor_vfinal"))
    if regra and SCRIPT_VER not in regra:
        raise ConfigError(f"regra_motor_vfinal nao referencia a versao {SCRIPT_VER}: {regra}")
    config["_governanca_operacional"] = "OK - script, config e regra_motor_vfinal coerentes"
    return True

def discover_files():
    INPUT_DIR.mkdir(exist_ok=True)
    files=[p for p in INPUT_DIR.iterdir() if p.suffix.lower() in {".xlsx",".xlsm",".xls"} and not p.name.startswith("~$")]
    return sorted([p for p in files if p.name.lower()!=DEFAULT_CONFIG["arquivo_fator_manual"].lower()])


def discover_inventory_files(config):
    files=[]
    if config.get("usar_subpastas_input",True):
        files.extend(_excel_files_in_dir(config.get("pasta_inventario","input/inventario")))
    if config.get("permitir_inventario_em_input_raiz",True):
        inv_pref=config.get("arquivos_inventario_prefixos",["INVENTARIO","ESTOQUE","SALDO"])
        for p in discover_files():
            if not _is_manual_factor_file(p,config) and _filename_matches_prefix(p,inv_pref): files.append(p)
    seen=set(); out=[]
    for p in files:
        rp=str(p.resolve())
        if rp not in seen: seen.add(rp); out.append(p)
    return sorted(out)

def discover_xml_files(config):
    if not config.get("usar_xml_como_movimentacao",True): return []
    folders=[BASE_DIR/config.get("pasta_xml","input/xml"), BASE_DIR/"input"/"XML"]
    files=[]
    for folder in folders:
        if folder.exists() and folder.is_dir():
            files.extend([p for p in folder.rglob("*.xml") if not p.name.startswith("~$")])
            if config.get("usar_zip_xml",True):
                files.extend([p for p in folder.rglob("*.zip") if not p.name.startswith("~$")])
    seen=set(); out=[]
    for p in sorted(files):
        rp=str(p.resolve())
        if rp not in seen: seen.add(rp); out.append(p)
    return out


# ─── Fatores manuais ──────────────────────────────────────────────────────────
def load_manual_factors(config):
    manual={}
    if not config.get("permitir_fator_manual",True): return manual
    manual_name=config.get("arquivo_fator_manual","fatores_manuais.xlsx")
    aux_folder=Path(config.get("pasta_auxiliares","input/auxiliares"))
    if not aux_folder.is_absolute(): aux_folder=BASE_DIR/aux_folder
    for path in [INPUT_DIR/manual_name, aux_folder/manual_name, BASE_DIR/manual_name,
                 (INPUT_DIR/manual_name).with_suffix(".csv"), (aux_folder/manual_name).with_suffix(".csv")]:
        if not path.exists(): continue
        try:
            df=pd.read_csv(path,sep=None,engine="python",dtype=object,keep_default_na=False) if path.suffix.lower()==".csv" else pd.read_excel(path,dtype=object,keep_default_na=False)
            cols=list(df.columns)
            c_cod=pick_column(cols,"codigo_item"); c_ncm=pick_column(cols,"ncm")
            c_factor=next((i for i,h in enumerate(cols) if norm_text(h) in {"FATORMANUAL","FATOR","FATORUNIDADE"}),None)
            if c_factor is None: continue
            for _,row in df.iterrows():
                fator=to_decimal(row.iloc[c_factor])
                if fator is None or fator<=0: continue
                cod=normalize_code_text(row.iloc[c_cod]) if c_cod is not None else ""
                ncm=normalize_ncm(row.iloc[c_ncm]) if c_ncm is not None else ""
                manual[(cod,ncm,"")]=int(fator); manual[(cod,ncm,norm_text(clean_str(row.iloc[c_ncm] if c_ncm else ""))[:80])]=int(fator)
        except Exception as exc: logger.exception("Fatores manuais: %s",exc)
    return manual

def find_manual_factor(cod,ncm,descricao,manual_factors):
    desc_key=norm_text(descricao)[:80]
    for key in [(cod,ncm,desc_key),(cod,ncm,""),(cod,"",""),("",ncm,desc_key)]:
        if key in manual_factors: return manual_factors[key]
    return None


def status_dimensional(qtd_com,qtd_trib,fator,fallback=False):
    if fator is None: return "REVISAR ATENCAO FATOR AUSENTE"
    if qtd_com is None or qtd_com<=ZERO: return "REVISAR ATENCAO QTDE COM ZERADA"
    if qtd_trib is None or qtd_trib<=ZERO: return "REVISAR ATENCAO QTDE TRIB ZERADA"
    if fallback: return "REVISAR ATENCAO QTDE TRIB AUSENTE - USADO QTDE COM"
    if qtd_com!=qtd_trib: return "REVISAR ATENCAO QTDE COM <> QTDE TRIB"
    return "OK OK"

def derive_unid_tributavel(unidade, fator=None, descricao=""):
    """Retorna a unidade usada na apuração tributável.

    Quando a unidade comercial/descrição indicar caixa, fardo, pacote ou pack
    com fator maior que 1, a quantidade é convertida para unidade tributável UN.
    Para KG/L/ML etc., preserva a própria unidade de medida.
    """
    unit = norm_text(unidade)
    desc = norm_text(descricao)
    try:
        fator_int = int(fator) if fator is not None else 1
    except Exception:
        fator_int = 1
    embalagem = {"CX", "CXS", "CAIXA", "FD", "FARDO", "PCT", "PACOTE", "PACK"}
    base_units = {"UN", "UND", "UNID", "UNIDADE", "KG", "G", "L", "LT", "ML", "M", "MT"}
    if fator_int > 1 or unit in embalagem or any(e in desc for e in embalagem):
        return "UN"
    if unit in base_units:
        return unit
    return clean_str(unidade).upper() or "UN"


def _hdr_idx(headers,aliases):
    nmap={norm_text(h):idx for idx,h in enumerate(headers)}
    for a in aliases:
        if norm_text(a) in nmap: return nmap[norm_text(a)]
    return None

def _val_at(row,idx):
    if idx is None: return ""
    try:
        return row[idx]
    except Exception:
        return ""

def row_value(row,mapping,name):
    idx=mapping.get(name)
    if idx is None or idx>=len(row): return ""
    return row.iloc[idx] if hasattr(row,"iloc") else row[idx]

def _legacy_build_record_rir70(row,mapping,config,manual_factors,ipi_column_present,row_origin=""):
    data_emissao=parse_date(row_value(row,mapping,"data_emissao"),config.get("formato_data"))
    # A competência deve respeitar as configurações de prioridade e fallback definidas no config.
    competencia=parse_competencia(row_value(row,mapping,"competencia"), data_emissao, config)
    situacao_documento_raw=clean_str(row_value(row,mapping,"situacao_documento"))
    situacao_documento=normalize_document_status(situacao_documento_raw)
    descricao=clean_str(row_value(row,mapping,"descricao"))
    unidade=clean_str(row_value(row,mapping,"unid_comercial"))
    cod_item=normalize_code_text(row_value(row,mapping,"codigo_item"))
    ncm=normalize_ncm(row_value(row,mapping,"ncm"))
    cfop=normalize_cfop(row_value(row,mapping,"cfop"))
    cnpj_emit=normalize_cnpj(row_value(row,mapping,"cnpj_emitente"))
    cnpj_dest=normalize_cnpj(row_value(row,mapping,"cnpj_destinatario"))
    nome_emit=clean_str(row_value(row,mapping,"nome_emitente"))
    nome_dest=clean_str(row_value(row,mapping,"nome_destinatario"))
    uf_dest=normalize_uf(row_value(row,mapping,"uf_destinatario"))
    cest=normalize_cest(row_value(row,mapping,"cest"))
    manual_factor=find_manual_factor(cod_item,ncm,descricao,manual_factors)
    fator,status_unidade,obs_unidade=extract_factor(descricao,unidade,config,manual_factor)
    qtd=to_decimal(row_value(row,mapping,"qtd_comercial"))
    qtd_trib_input=to_decimal(row_value(row,mapping,"qtd_tributavel"))
    qtd_trib=qtd_trib_input
    qtd_trib_fallback=False

    # Conversão fiscal de embalagem para unidade tributável.
    # Regra de auditoria:
    #   - Se o input/XML já trouxer qTrib válida e diferente da qCom, ela prevalece.
    #   - Se a unidade/descrição indicar caixa/fardo/pacote/pack com fator > 1
    #     e qTrib vier ausente ou igual à qCom, converte: QTDE TRIB = QTDE COM × FATOR.
    #   - Essa quantidade convertida é a única usada na aba 02_ARBITRAMENTO e no CMV.
    if qtd is not None:
        fator_dec = Decimal(fator) if fator not in (None, 0) else Decimal("1")
        if fator_dec > 1 and (qtd_trib is None or qtd_trib == qtd):
            qtd_trib = qtd * fator_dec
            qtd_trib_fallback = True
            status_unidade = "CONVERTIDO OK CX/PACK -> UN"
            obs_unidade = f"Quantidade tributável calculada: {qtd} × {fator_dec} = {qtd_trib}."
        elif qtd_trib is None:
            qtd_trib = qtd
            qtd_trib_fallback = True
            obs_unidade = "Quantidade tributável ausente; adotada quantidade comercial."
        elif fator_dec > 1 and qtd_trib == qtd * fator_dec:
            status_unidade = "VALIDADO OK QTDE TRIB INFORMADA"
            obs_unidade = "Quantidade tributável informada já corresponde à quantidade comercial convertida pelo fator."

    unid_trib = derive_unid_tributavel(unidade, fator, descricao)
    valor_total=to_decimal(row_value(row,mapping,"valor_total_produto"))
    valor_ipi=to_decimal(row_value(row,mapping,"valor_ipi")) if ipi_column_present else None
    valor_unit_inf=to_decimal(row_value(row,mapping,"valor_unitario"))
    valor_unit_trib_inf=to_decimal(row_value(row,mapping,"valor_unitario_tributavel"))
    if not ipi_column_present:
        status_ipi="IPI NÃO LOCALIZADO - SEM DEDUÇÃO DO VALOR BRUTO ATENCAO"
        ipi_calc=ZERO
        ipi_bloqueado=bool(config.get("bloquear_calculo_sem_coluna_ipi",False))
    else:
        status_ipi="LOCALIZADO OK" if valor_ipi is not None else "LOCALIZADO VAZIO - ZERO OK"
        ipi_calc=valor_ipi if valor_ipi is not None else ZERO
        ipi_bloqueado=False
    valor_base=None
    if valor_total is not None:
        tratamento=clean_str(config.get("tratamento_ipi_valor_produto","")).lower() or "valor_item_inclui_ipi"
        deduzir_ipi = bool(config.get("deduzir_ipi_do_valor_produto_bruto", tratamento=="valor_item_inclui_ipi"))
        # Regra deliberada do projeto: ICMS NÃO é deduzido do preço de venda.
        # O RIR/2018 art. 308, par.1º, determina o preço de venda sem exclusão de ICMS.
        deduzir_icms = bool(config.get("deduzir_icms_do_valor_produto", False))
        if deduzir_icms:
            logger.warning("Config deduzir_icms_do_valor_produto=True ignorada: o motor não deduz ICMS do preço-base.")
        valor_base = valor_total - (ipi_calc if deduzir_ipi and isinstance(ipi_calc, Decimal) else ZERO)
        if valor_base < ZERO:
            valor_base = ZERO

    # Valores expostos na aba 01_MOVIMENTACAO para auditoria NF-e:
    # - Valor Comercial = valor total do produto informado na unidade comercial (vProd).
    # - Valor Trib. = valor reconciliado pela unidade tributável. Se existir vUnTrib
    #   no XML/input, calcula qTrib × vUnTrib; caso contrário, usa o mesmo valor-base
    #   do item para evitar criar diferença artificial.
    valor_comercial_total = valor_total
    if valor_unit_trib_inf is not None and qtd_trib not in (None, ZERO):
        valor_tributavel_total = valor_unit_trib_inf * qtd_trib
    else:
        valor_tributavel_total = valor_base

    usou_vu = not (valor_base is not None and qtd not in (None,ZERO))
    if usou_vu:
        if bool(config.get("deduzir_ipi_do_valor_produto_bruto", clean_str(config.get("tratamento_ipi_valor_produto","")).lower()=="valor_item_inclui_ipi")) and ipi_calc and valor_unit_inf:
            qref=qtd_trib or qtd
            if qref and qref>ZERO: valor_unit_inf=valor_unit_inf-(ipi_calc/qref)
        unit_comercial=valor_unit_inf
    else:
        unit_comercial=valor_base/qtd

    # Preço unitário fiscal: sempre por unidade tributável.
    # Quando há valor total e QTDE TRIB convertida, divide diretamente pelo denominador fiscal.
    # Isso evita subavaliação em CX6/CX12/CX24 e evita dupla conversão quando qTrib já veio pronta.
    if valor_base is not None and qtd_trib not in (None, ZERO):
        unit_sem_ipi = q4(valor_base / qtd_trib)
    elif unit_comercial is not None and fator not in (None, 0):
        unit_sem_ipi = q4(unit_comercial / Decimal(fator))
    else:
        unit_sem_ipi = None

    valor_unit_comercial_aud = valor_unit_inf
    if valor_unit_comercial_aud is None and valor_comercial_total is not None and qtd not in (None, ZERO):
        valor_unit_comercial_aud = valor_comercial_total / qtd
    valor_unit_tributavel_aud = valor_unit_trib_inf if valor_unit_trib_inf is not None else unit_sem_ipi

    chave=normalize_chave(row_value(row,mapping,"chave"))
    valid_chave=validate_nfe_key(chave)
    tipo_operacao,participa_cfop,motivo_cfop=classify_cfop(cfop,config)
    cnpjs_grupo={normalize_cnpj(c) for c in config.get("cnpjs_grupo",[]) if normalize_cnpj(c)}
    intercompany=bool(cnpjs_grupo and cnpj_emit in cnpjs_grupo and cnpj_dest in cnpjs_grupo)
    participa=participa_cfop
    if config.get("bloquear_chave_nfe_invalida", False) and not clean_str(valid_chave).startswith("VALIDADO"):
        participa=False; tipo_operacao="CHAVE NF-E INVÁLIDA"; motivo_cfop=valid_chave
    if valor_total is not None and valor_total<=ZERO: participa=False; tipo_operacao="VALOR INVÁLIDO"; motivo_cfop="Valor total do item ≤ zero."
    bloq_norm=norm_text(situacao_documento)
    bloqueantes=[norm_text(s) for s in config.get("status_documento_bloqueantes",[]) if norm_text(s)]
    situacao_bloqueada=bool(bloq_norm and any(s in bloq_norm for s in bloqueantes))
    if situacao_bloqueada: participa=False; tipo_operacao="DOCUMENTO BLOQUEADO"; motivo_cfop=f"Situação bloqueante: {situacao_documento}"
    if config.get("excluir_intercompany_do_calculo",True) and intercompany: participa=False; tipo_operacao="INTERCOMPANY"; motivo_cfop="Emitente e destinatário no grupo"
    if ipi_bloqueado: participa=False; tipo_operacao="IPI NÃO MAPEADO"; motivo_cfop="Coluna IPI ausente - bloqueado por config"
    if unit_sem_ipi is None or unit_sem_ipi<=0: participa=False
    cnpj_key=cnpj_emit if config.get("segregar_calculo_por_cnpj_emitente",True) else ""
    item_key=normalize_item_join_key(cod_item) or norm_text(descricao)[:80]
    return {
        "Chave NF-e":            chave,
        "Validação Chave":       valid_chave,
        "Status IPI":            status_ipi,
        "Competência":           competencia,
        "Data Emissão":          data_emissao,
        "Número NF":             normalize_code_text(row_value(row,mapping,"numero_nf")),
        "Série":                 normalize_code_text(row_value(row,mapping,"serie")),
        "CNPJ Emitente":         cnpj_emit,
        "Nome Emitente":         "" if _looks_like_uf(nome_emit) else nome_emit,
        "CNPJ Destinatário":     cnpj_dest,
        "Nome Destinatário":     "" if _looks_like_uf(nome_dest) else nome_dest,
        "UF Dest":               uf_dest,
        "NCM":                   ncm,
        "CFOP":                  cfop,
        "Situação Documento":    situacao_documento,
        "Tipo Operação":         tipo_operacao,
        "Código Item":           cod_item,
        "Descrição":             descricao,
        "Unid. Comercial":       unidade,
        "Fator Unidade":         int(fator) if fator is not None else None,
        "Unid. Trib.":           unid_trib,
        "Unid.":                 unid_trib,
        "Status Unidade":        status_unidade,
        "QTDE COM":              q4(qtd) if qtd is not None else None,
        "QTDE":                  q4(qtd_trib) if qtd_trib is not None else None,
        "QTDE TRIB":             q4(qtd_trib) if qtd_trib is not None else None,
        "QTDE TRIB Informada":   q4(qtd_trib_input) if qtd_trib_input is not None else None,
        "Vlr Unit. Comercial":   q4(valor_unit_comercial_aud) if isinstance(valor_unit_comercial_aud, Decimal) else None,
        "Valor Comercial":       q2(valor_comercial_total) if isinstance(valor_comercial_total, Decimal) else None,
        "Vlr Unit. Tributável":       q4(valor_unit_tributavel_aud) if isinstance(valor_unit_tributavel_aud, Decimal) else None,
        "Valor Tributável":           q2(valor_tributavel_total) if isinstance(valor_tributavel_total, Decimal) else None,
        "Vlr Unitário Base RIR70":  unit_sem_ipi,
        "Participa Cálculo":     "SIM" if participa else "NÃO",
        "Motivo Exclusão":       "" if participa else (motivo_cfop or "Não participante"),
        "_participa":            participa,
        "_motivo":               motivo_cfop,
        "_monthly_key":          (cnpj_key, item_key, ncm, competencia),
        "_group_key":            (cnpj_key, item_key, ncm),
        "_cnpj_key":             cnpj_emit,
        "_sort_comp":            competencia_sort_key(competencia),
        "_sort_data":            data_emissao or datetime.max,
        "_sort_desc":            descricao.upper(),
        "_qtd":                  qtd,
        "_qtd_trib":             qtd_trib,
        "_valor_total":          valor_total,
        "_ipi":                  ipi_calc,
        "_intercompany":         intercompany,
        "_situacao_bloqueada":   situacao_bloqueada,
        "_row_origin":           row_origin,
        "_item_nf":              normalize_code_text(row_value(row,mapping,"item_nf")),
    }


def _record_has_value(v):
    if v is None: return False
    if isinstance(v,(Decimal,datetime)): return True
    return clean_str(v) not in {"","N/D","ND","NONE","NAN"}

def merge_record_missing_fields(existing,incoming):
    if not existing or not incoming: return 0
    fields=["Nome Emitente","Nome Destinatário","UF Dest","NCM","CFOP","CNPJ Emitente","CNPJ Destinatário","Número NF","Série","Data Emissão","Situação Documento","QTDE TRIB","CEST"]
    changed=0
    for f in fields:
        if not _record_has_value(existing.get(f)) and _record_has_value(incoming.get(f)):
            existing[f]=incoming.get(f); changed+=1
    return changed

def rebuild_record_group_keys(r, config):
    """Recalcula as chaves internas quando CNPJ Emitente, Código Item ou NCM são corrigidos."""
    cnpj_key = normalize_cnpj(r.get("CNPJ Emitente", "")) if config.get("segregar_calculo_por_cnpj_emitente", True) else ""
    item_key = normalize_item_join_key(r.get("Código Item", "")) or norm_text(r.get("Descrição", ""))[:80]
    ncm = normalize_ncm(r.get("NCM", ""))
    comp = clean_str(r.get("Competência", ""))
    r["_cnpj_key"] = cnpj_key
    r["_monthly_key"] = (cnpj_key, item_key, ncm, comp)
    r["_group_key"] = (cnpj_key, item_key, ncm)
    return r

def _exact_header_idx(headers, aliases):
    """Localiza coluna somente por equivalência normalizada exata, sem fuzzy match."""
    nmap = {norm_text(h): idx for idx, h in enumerate(headers) if norm_text(h)}
    for alias in aliases:
        key = norm_text(alias)
        if key in nmap:
            return nmap[key]
    return None

def _cnpj_emitente_aliases():
    return [
        "CnpjDoEmitente", "CNPJ DO EMITENTE", "CNPJDOEMITENTE",
        "CnpjOuCpfDoEmitente", "CNPJ OU CPF DO EMITENTE", "CNPJOUCPFDOEMITENTE",
        "CNPJ Emitente", "CNPJEMITENTE", "CNPJ_EMITENTE", "CNPJ CPF EMITENTE"
    ]

def _cnpj_destinatario_aliases():
    return [
        "CnpjDoDestinatario", "CNPJ DO DESTINATARIO", "CNPJDODESTINATARIO",
        "CnpjOuCpfDoDestinatario", "CNPJ OU CPF DO DESTINATARIO", "CNPJOUCPFDODESTINATARIO",
        "CNPJ Destinatario", "CNPJDESTINATARIO", "CNPJ_DESTINATARIO", "CNPJ CPF DESTINATARIO"
    ]


# ─── Documentos consolidados ──────────────────────────────────────────────────
def load_document_master(config,logs):
    master={}; detail_rows=[]
    if not config.get("usar_documentos_consolidados_para_enriquecimento",True): return master,detail_rows
    files=discover_document_files(config)
    if not files:
        logs.append(("-","DOCUMENTOS","DOCUMENTOS_NAO_LOCALIZADOS","Nenhum arquivo documental consolidado localizado."))
        return master,detail_rows
    for file_path in files:
        try:
            sheet_names=xlsx_sheet_names(file_path) if file_path.suffix.lower() in {".xlsx",".xlsm"} else []
            if not sheet_names:
                with pd.ExcelFile(file_path, engine=excel_engine(file_path)) as xl:
                    sheet_names = list(xl.sheet_names)
            for sh in sheet_names:
                try:
                    rows=list(iter_xlsx_rows_stream(file_path,sh)) if file_path.suffix.lower() in {".xlsx",".xlsm"} else []
                    if not rows: continue
                    headers=make_unique_headers(rows[0])
                    if classify_file_headers(headers)!="DOCUMENTO_CONSOLIDADO": continue
                    idx={"chave":_hdr_idx(headers,["ChaveAcesso","Chave Acesso"]),"situacao":_hdr_idx(headers,["Situacao","SituacaoDocumento","Status"]),"nome_emit":_hdr_idx(headers,["NomeEmitente","RazaoSocialEmitente"]),"nome_dest":_hdr_idx(headers,["NomeDestinatario","RazaoSocialDestinatario"]),"cnpj_emit":_hdr_idx(headers,["CnpjDoEmitente","CnpjEmitente"]),"cnpj_dest":_hdr_idx(headers,["CnpjDoDestinatario","CnpjDestinatario"])}
                    imp=0
                    for row in rows[1:]:
                        chave=normalize_chave(_val_at(row,idx["chave"]))
                        if not re.fullmatch(r"\d{44}",clean_str(chave)): continue
                        master[chave]={"ChaveAcesso":chave,"Status Normalizado":normalize_document_status(_val_at(row,idx["situacao"])),"NomeEmitente":clean_str(_val_at(row,idx["nome_emit"])),"NomeDestinatario":clean_str(_val_at(row,idx["nome_dest"])),"CnpjEmitente":normalize_cnpj(_val_at(row,idx["cnpj_emit"])),"CnpjDestinatario":normalize_cnpj(_val_at(row,idx["cnpj_dest"]))}
                        imp+=1
                    logs.append((file_path.name,sh,"DOCUMENTO_CONSOLIDADO",f"Importados={imp}"))
                except Exception as e: logs.append((file_path.name,sh,"DOCUMENTOS_ERRO",str(e)))
        except Exception as e: logs.append((file_path.name,"-","DOCUMENTOS_ERRO",str(e)))
    return master,list(master.values())

def apply_document_master(records,document_master,logs,config=None):
    if config is None: config = DEFAULT_CONFIG
    if not document_master:
        logs.append(("-","DOCUMENTOS","ENRIQUECIMENTO_NAO_APLICADO","Nenhum documento consolidado disponível.")); return
    enriched=blocked=0
    for r in records:
        doc=document_master.get(clean_str(r.get("Chave NF-e")))
        if not doc: continue
        enriched+=1
        # Enriquecimento por Chave NF-e a partir dos documentos consolidados.
        # Importante para a aba 01_MOVIMENTACAO: quando o movimento não traz
        # CNPJ Emitente/Destinatário, usa CnpjDoEmitente/CnpjDoDestinatario
        # capturados dos arquivos da pasta input/documentos ou da pasta input,
        # conforme configuração de descoberta do pacote.
        if doc.get("CnpjEmitente") and not clean_str(r.get("CNPJ Emitente")):
            r["CNPJ Emitente"] = normalize_cnpj(doc["CnpjEmitente"])
            r["_cnpj_key"] = r["CNPJ Emitente"]
            cnpj_key = r["CNPJ Emitente"] if config.get("segregar_calculo_por_cnpj_emitente", True) else ""
            item_key = normalize_item_join_key(r.get("Código Item", "")) or norm_text(r.get("Descrição", ""))[:80]
            r["_monthly_key"] = (cnpj_key, item_key, normalize_ncm(r.get("NCM", "")), r.get("Competência", ""))
            r["_group_key"] = (cnpj_key, item_key, normalize_ncm(r.get("NCM", "")))
        if doc.get("CnpjDestinatario") and not clean_str(r.get("CNPJ Destinatário")):
            r["CNPJ Destinatário"] = normalize_cnpj(doc["CnpjDestinatario"])
        if doc.get("NomeEmitente") and not clean_str(r.get("Nome Emitente")): r["Nome Emitente"]=doc["NomeEmitente"]
        if doc.get("NomeDestinatario") and not clean_str(r.get("Nome Destinatário")): r["Nome Destinatário"]=doc["NomeDestinatario"]
        sn=clean_str(doc.get("Status Normalizado"))
        if sn:
            r["Situação Documento"]=sn
            if sn in {"CANCELADA","DENEGADA","INUTILIZADA","EXTEMPORANEA"}:
                if r.get("_participa"): blocked+=1
                r["_participa"]=False; r["Participa Cálculo"]="NÃO"; r["Motivo Exclusão"]=f"Situação bloqueante pelo consolidado: {sn}"
    logs.append(("-","DOCUMENTOS","ENRIQUECIMENTO_APLICADO",f"Enriquecidos={enriched}; bloqueados={blocked}"))

def apply_party_directory(records,document_master,logs):
    party={}

    def _put(cnpj, nome):
        cnpj=normalize_cnpj(cnpj); nome=clean_str(nome)
        if not cnpj or not nome or _looks_like_uf(nome):
            return
        # Mantém a razão social mais completa quando houver duplicidade por CNPJ.
        if cnpj not in party or len(nome) > len(party[cnpj]):
            party[cnpj]=nome

    for doc in (document_master or {}).values():
        _put(doc.get("CnpjEmitente",""), doc.get("NomeEmitente",""))
        _put(doc.get("CnpjDestinatario",""), doc.get("NomeDestinatario",""))
    for r in records:
        _put(r.get("CNPJ Emitente",""), r.get("Nome Emitente",""))
        # Em operações entre empresas do grupo, o mesmo CNPJ pode aparecer como destinatário
        # em uma NF e como emitente em outra. Usar esse diretório reduz Nome Emitente vazio.
        _put(r.get("CNPJ Destinatário",""), r.get("Nome Destinatário",""))

    filled=0
    for r in records:
        cnpj=normalize_cnpj(r.get("CNPJ Emitente",""))
        if cnpj and (not clean_str(r.get("Nome Emitente","")) or _looks_like_uf(r.get("Nome Emitente",""))) and party.get(cnpj):
            r["Nome Emitente"]=party[cnpj]; filled+=1
    logs.append(("-","DOCUMENTOS","ENRIQUECIMENTO_CNPJ",f"Nome Emitente preenchido/normalizado por CNPJ={filled}; CNPJs conhecidos={len(party)}"))

# ─── XML ──────────────────────────────────────────────────────────────────────
def _xml_find_text(parent,path,default=""):
    if parent is None: return default
    v=parent.findtext(path); return clean_str(v) if v is not None else default

def _xml_attr_chave(root):
    inf=root.find(".//{*}infNFe")
    if inf is None: return ""
    v=clean_str(inf.attrib.get("Id",""))
    return v[3:] if v.startswith("NFe") else v

def _xml_ipi_value(det):
    for path in [".//{*}IPI/{*}IPITrib/{*}vIPI",".//{*}IPI/{*}IPINT/{*}vIPI",".//{*}vIPI"]:
        v=det.findtext(path)
        if clean_str(v): return clean_str(v)
    return "0"

def _xml_status(root):
    cstat=clean_str(root.findtext(".//{*}protNFe/{*}infProt/{*}cStat") or root.findtext(".//{*}cStat") or "")
    xmotivo=clean_str(root.findtext(".//{*}protNFe/{*}infProt/{*}xMotivo") or root.findtext(".//{*}xMotivo") or "")
    if cstat=="100": return "AUTORIZADA"
    if cstat in {"101","135","155"} or "cancel" in xmotivo.lower(): return "CANCELADA"
    if cstat in {"110","111"} or "inutiliz" in xmotivo.lower(): return "INUTILIZADA"
    if cstat in {"301","302"} or "denegad" in xmotivo.lower(): return "DENEGADA"
    return "NAO_CONFIRMADA"

_XML_LOGICAL = ["chave","competencia","data_emissao","numero_nf","serie","gtin","ncm","cfop","codigo_item","item_nf","descricao","unid_comercial","qtd_comercial","valor_unitario","valor_unitario_tributavel","valor_total_produto","valor_ipi","situacao_documento","cnpj_emitente","cnpj_destinatario","nome_emitente","nome_destinatario","uf_destinatario","cest","qtd_tributavel"]

def _xml_rows_from_bytes(xml_bytes,source="XML"):
    try: root=ET.fromstring(xml_bytes)
    except Exception as e: return [],f"Falha XML {source}: {e}"
    chave=_xml_attr_chave(root)
    ide=root.find(".//{*}ide"); emit=root.find(".//{*}emit"); dest=root.find(".//{*}dest")
    status=_xml_status(root)
    data_emissao=_xml_find_text(ide,"./{*}dhEmi") or _xml_find_text(ide,"./{*}dEmi")
    numero_nf=_xml_find_text(ide,"./{*}nNF"); serie=_xml_find_text(ide,"./{*}serie")
    cnpj_emit=_xml_find_text(emit,"./{*}CNPJ") or _xml_find_text(emit,"./{*}CPF")
    nome_emit=_xml_find_text(emit,"./{*}xNome")
    cnpj_dest=_xml_find_text(dest,"./{*}CNPJ") or _xml_find_text(dest,"./{*}CPF")
    nome_dest=_xml_find_text(dest,"./{*}xNome")
    uf_dest=_xml_find_text(dest,"./{*}enderDest/{*}UF")
    rows=[]
    for det in root.findall(".//{*}det"):
        prod=det.find("./{*}prod")
        if prod is None: continue
        v={"chave":chave,"competencia":"","data_emissao":data_emissao,"numero_nf":numero_nf,"serie":serie,"gtin":_xml_find_text(prod,"./{*}cEAN"),"ncm":_xml_find_text(prod,"./{*}NCM"),"cfop":_xml_find_text(prod,"./{*}CFOP"),"codigo_item":_xml_find_text(prod,"./{*}cProd"),"item_nf":clean_str(det.attrib.get("nItem","")),"descricao":_xml_find_text(prod,"./{*}xProd"),"unid_comercial":_xml_find_text(prod,"./{*}uCom"),"qtd_comercial":_xml_find_text(prod,"./{*}qCom"),"valor_unitario":_xml_find_text(prod,"./{*}vUnCom"),"valor_unitario_tributavel":_xml_find_text(prod,"./{*}vUnTrib"),"valor_total_produto":_xml_find_text(prod,"./{*}vProd"),"valor_ipi":_xml_ipi_value(det),"situacao_documento":status,"cnpj_emitente":cnpj_emit,"cnpj_destinatario":cnpj_dest,"nome_emitente":nome_emit,"nome_destinatario":nome_dest,"uf_destinatario":uf_dest,"cest":_xml_find_text(prod,"./{*}CEST"),"qtd_tributavel":_xml_find_text(prod,"./{*}qTrib") or _xml_find_text(prod,"./{*}qCom")}
        rows.append([v.get(k,"") for k in _XML_LOGICAL])
    return rows,"OK"

def _xml_rows_from_zip(zip_path):
    all_rows=[]
    inner_hashes=[]
    try:
        with zipfile.ZipFile(zip_path) as zf:
            for name in sorted(zf.namelist()):
                if not name.lower().endswith(".xml"):
                    continue
                data = zf.read(name)
                inner_hashes.append(f"{name}:{_sha256_bytes(data)}")
                rows,_=_xml_rows_from_bytes(data,f"{Path(zip_path).name}|{name}")
                all_rows.extend(rows)
        if inner_hashes:
            # Rastreia o conteúdo fiscal efetivamente lido, e não o arquivo ZIP contêiner.
            RIR70_SOURCE_CONTROL_CODES[Path(zip_path).name] = _sha256_bytes("\n".join(inner_hashes).encode("utf-8"))
        else:
            RIR70_SOURCE_CONTROL_CODES[Path(zip_path).name] = sha256_file(zip_path)
    except Exception as e:
        return [],f"Falha ZIP XML: {e}"
    return all_rows,"OK"

def process_record_row(row,mapping,config,manual_factors,ipi_col,file_path,sheet_name,periodo_inicio,periodo_fim,fingerprints,records,duplicates,row_index=None):
    if all(clean_str(v)=="" for v in row): return "VAZIA"
    row_origin=f"{file_path.name}|{sheet_name}|{row_index or 'N/D'}"
    r=build_record(row,mapping,config,manual_factors,ipi_col,row_origin)
    dt=r.get("Data Emissão")
    if periodo_inicio and (dt is None or dt<periodo_inicio): return "FORA_PERIODO"
    if periodo_fim and (dt is None or dt>periodo_fim): return "FORA_PERIODO"
    data_inicial_por_cnpj = config.get("data_inicial_por_cnpj") or {}
    if data_inicial_por_cnpj:
        cnpj_emit_r = normalize_cnpj(r.get("CNPJ Emitente") or "")
        limite_str = data_inicial_por_cnpj.get(cnpj_emit_r)
        if limite_str:
            limite_dt = parse_config_date(limite_str)
            if limite_dt and (dt is None or dt < limite_dt):
                return "FORA_PERIODO"
    fp=duplicate_fingerprint(r)
    if fp in fingerprints:
        existing=fingerprints.get(fp)
        merged=merge_record_missing_fields(existing,r) if existing is not None else 0
        duplicates.append((file_path.name,sheet_name,r.get("Chave NF-e",""),r.get("Número NF",""),r.get("Código Item",""),f"DUPLICATA {'ENRIQUECIDA' if merged else 'DESCARTADA'}"))
        return "DUPLICATA"
    fingerprints[fp]=r; records.append(r); return "IMPORTADA"


def load_xml_inputs(config, manual_factors, periodo_inicio, periodo_fim, fingerprints, records, duplicates, logs):
    """
    Carrega notas fiscais eletrônicas em formato XML. Quando a opção
    `usar_leitura_xml_paralela` está ativada no config, a leitura dos
    arquivos XML e ZIP é realizada em paralelo usando `ThreadPoolExecutor`.
    A importação em paralelo melhora a performance em grandes volumes de
    documentos, mas o processamento individual dos registros ainda é feito
    sequencialmente para manter a consistência das estruturas compartilhadas
    (records, duplicates, fingerprints).
    """
    files = discover_xml_files(config)
    if not files:
        logs.append(("-", "XML", "XML_NAO_LOCALIZADO", "Nenhum XML em input/xml."))
        return
    mapping = {name: idx for idx, name in enumerate(_XML_LOGICAL)}

    def _parse_xml_file(file_path):
        """Lê um único arquivo XML ou ZIP e retorna (file_path, rows, status, sha256)."""
        try:
            if file_path.suffix.lower() == ".zip":
                rows, status = _xml_rows_from_zip(file_path)
                sha = ""
            else:
                data = Path(file_path).read_bytes()
                sha = _sha256_bytes(data)
                rows, status = _xml_rows_from_bytes(data, file_path.name)
            return file_path, rows, status, sha
        except Exception as exc:
            # Em caso de erro, retorna status de erro
            return file_path, [], f"Erro ao ler XML: {exc}", ""

    # Se o paralelismo estiver habilitado, utilizar ThreadPoolExecutor para ler
    # os arquivos em paralelo. Caso contrário, processar sequencialmente.
    parsed_results = []
    if config.get("usar_leitura_xml_paralela", False):
        max_workers = int(config.get("max_workers_xml", 4) or 4)
        # Se xml_em_rede=True, limitar paralelismo para evitar travamento em UNC/rede lenta.
        if config.get("xml_em_rede", False):
            max_workers = min(max_workers, 2)
            logger.info("XML em rede detectado: max_workers_xml limitado a %d.", max_workers)
        timeout_per_file = 60  # segundos por XML antes de desistir
        try:
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                future_to_file = {executor.submit(_parse_xml_file, f): f for f in files}
                for future in as_completed(future_to_file):
                    try:
                        file_path, rows, status, sha = future.result(timeout=timeout_per_file)
                    except Exception as exc:
                        fp = future_to_file[future]
                        logs.append((fp.name, "XML", "ERRO_XML", f"Timeout ou erro no processamento paralelo: {exc}"))
                        continue
                    parsed_results.append((file_path, rows, status, sha))
        except Exception as exc:
            logs.append(("-", "XML", "ERRO_XML", f"Falha no processamento paralelo de XML: {exc}. Revertendo para processamento sequencial."))
            parsed_results.clear()
    if not parsed_results:
        for file_path in files:
            file_path, rows, status, sha = _parse_xml_file(file_path)
            parsed_results.append((file_path, rows, status, sha))
    # Ordenação determinística por nome de arquivo: garante que, em caso de colisão de
    # deduplicação, a linha "sobrevivente" seja sempre a mesma entre execuções.
    parsed_results.sort(key=lambda x: x[0].name)

    # Processa os resultados de leitura (rows, status, sha) sequencialmente para
    # atualizar as estruturas compartilhadas de forma segura.
    for file_path, rows, status, sha in parsed_results:
        if sha:
            RIR70_SOURCE_CONTROL_CODES[file_path.name] = sha
        if status != "OK":
            logs.append((file_path.name, "XML", "ERRO_XML", status))
            continue
        before = len(records)
        out = dup = empty = 0
        for i, row in enumerate(rows, 1):
            s = process_record_row(
                row, mapping, config, manual_factors, True, file_path, "XML_NFE",
                periodo_inicio, periodo_fim, fingerprints, records, duplicates, row_index=i
            )
            if s == "FORA_PERIODO":
                out += 1
            elif s == "DUPLICATA":
                dup += 1
            elif s == "VAZIA":
                empty += 1
        logs.append(
            (
                file_path.name,
                "XML",
                "XML_PROCESSADO",
                f"Importados={len(records) - before}; fora_periodo={out}; duplicadas={dup}; codigo_controle=arquivo_lido",
            )
        )


# ─── CÁLCULO CENTRAL: 70% do Maior Preço ─────────────────────────────────────

# ─── Montagem das linhas de output ────────────────────────────────────────────
# Colunas da aba 01_MOVIMENTACAO - linha a linha, todas as NFs; Nome Emitente removido da saída por solicitação do usuário
MOV_COLS = [
    "Competência", "Data Emissão", "CNPJ Emitente",
    "Nº NF", "CFOP", "NCM", "Código Item", "Descrição",
    "Unid. Comercial", "Fator Unidade", "Unid. Trib.",
    "QTDE COM", "QTDE TRIB",
    "Vlr Unit. Comercial", "Valor Comercial",
    "Vlr Unit. Tributável", "Valor Tributável",
    "Vlr Unitário Base RIR70",
    "Participa Cálculo", "Motivo Exclusão",
]

# Colunas da aba 02_ARBITRAMENTO - 1 linha por produto/NCM/competência
ARB_COLS = [
    "Competência", "CNPJ Emitente", "Código Item", "NCM", "Descrição",
    "Unid. Comercial", "Fator Unidade", "Unid. Trib.", "Unid.",
    "Qtde Movimentada",
    "Vlr Médio",          # média das vendas participantes
    "Maior Vlr Unitário", # base do cálculo
    "Ratio Maior/Médio",  # alertar outliers: 1.0 = única venda; >2 = alerta
    "Base Arbitrada (70%)", # 70% do Maior - RESULTADO PRINCIPAL
    "CMV Competência",      # Base × Qtde tributável
    "Nº NF Referência",   # NF com o maior preço
    "Data NF Referência",
]

# Colunas da aba 03_INVENTARIO_PA
INV_COLS = [
    "CNPJ Inventário", "Competência", "Código Item", "NCM", "Descrição",
    "Unid.", "QTDE Inventário", "Base Arbitrada (70%)", "Valor Inventário PA", "Status",
]

# Colunas da aba 04_LOG_EXCLUSOES
LOG_COLS = [
    "Competência", "CNPJ Emitente", "Nº NF", "Código Item", "NCM",
    "CFOP", "Tipo Operação", "Motivo Exclusão", "Arquivo/Origem",
]

# Aba 02A - validação da origem e conciliação entre movimento e arbitramento.
ORIGEM_COLS = [
    "CNPJ Emitente", "Competência", "NCM", "Linhas Origem", "Linhas Participantes",
    "Itens Arbitrados", "Qtde Trib Mov.", "Qtde Trib Arbitrada", "Dif. Qtde",
    "CMV Arbitrado", "Status Auditoria",
]

# Aba 03A - conciliação específica do inventário PA e origem da base arbitrada.
ORIGEM_INV_COLS = [
    "CNPJ Inventário", "Competência", "Itens Inventário", "Itens Calculados",
    "Itens sem Base", "Itens com Base Competência Anterior", "Itens com Base de Produto Similar", "Itens com Base de Outro CNPJ",
    "Qtde Inv. Original", "Qtde Inv. Trib.", "Dif. Conversão",
    "Valor Inventário PA", "Status Auditoria",
]







def _counter_main(counter_obj):
    if not counter_obj:
        return ""
    try:
        return counter_obj.most_common(1)[0][0]
    except Exception:
        return ""

def _inventario_factor_from_unit_or_desc(unidade, descricao, config):
    fator, status, obs = extract_factor(descricao, unidade, config or DEFAULT_CONFIG)
    if fator is None:
        # Para inventário, não converter quando não houver fator seguro. Mantém quantidade física e registra fator 1.
        return Decimal("1"), status, obs
    try:
        return Decimal(str(fator)), status, obs
    except Exception:
        return Decimal("1"), "OK", "Fator de inventário não numérico; adotado 1."




_SIMILARITY_CACHE = {}
_PRECO_REF_CACHE:  dict = {}  # cache dos preços de referência manuais (arquivo -> dict)

def _similarity_file_candidates(config):
    """Locais aceitos para o mapa de similaridade homologado."""
    fname = clean_str(config.get("arquivo_mapa_similaridade_produto", "mapa_similaridade_produto.xlsx")) or "mapa_similaridade_produto.xlsx"
    aux = Path(config.get("pasta_auxiliares", "input/auxiliares"))
    if not aux.is_absolute():
        aux = BASE_DIR / aux
    return [
        aux / fname,
        INPUT_DIR / "auxiliares" / fname,
        INPUT_DIR / fname,
        BASE_DIR / fname,
    ]

def load_similarity_map(config):
    """Carrega mapa homologado de similaridade produto a produto.

    Regras:
      - usa somente linhas com PERMITE_FALLBACK_SIMILAR = S e STATUS_HOMOLOGACAO = APROVADO;
      - chaveia pelo código normalizado, aceitando CODIGO e CODIGO_BASE;
      - o grupo deve vir da coluna GRUPO_SIMILARIDADE_HOMOLOGADO.
    """
    candidates = _similarity_file_candidates(config)
    path = next((p for p in candidates if p.exists()), None)
    if path is None:
        return {}
    try:
        st = path.stat()
        cache_key = (str(path.resolve()), st.st_mtime_ns, st.st_size)
    except Exception:
        cache_key = (str(path), 0, 0)
    if cache_key in _SIMILARITY_CACHE:
        return _SIMILARITY_CACHE[cache_key]

    try:
        df = pd.read_excel(path, sheet_name="MAPA_SIMILARIDADE", dtype=object, keep_default_na=False)
    except Exception:
        try:
            df = pd.read_excel(path, dtype=object, keep_default_na=False)
        except Exception as exc:
            logger.warning("Mapa de similaridade não pôde ser lido: %s", exc)
            return {}

    cols = {norm_text(c): c for c in df.columns}
    def _get(row, name):
        col = cols.get(norm_text(name))
        return clean_str(row.get(col, "")) if col else ""

    sim = {}
    for _, row in df.iterrows():
        grupo = _get(row, "GRUPO_SIMILARIDADE_HOMOLOGADO")
        status = norm_text(_get(row, "STATUS_HOMOLOGACAO"))
        permite = _yes_no_status(_get(row, "PERMITE_FALLBACK_SIMILAR"), "")
        status_cad = norm_text(_get(row, "STATUS_CADASTRO"))
        part_map = _yes_no_status(_get(row, "PART_ARBITRAMENTO") or _get(row, "PARTICIPA_ARBITR"), "")
        if not grupo or status != "APROVADO" or permite != "Sim" or status_cad != "ATIVO" or part_map != "Sim":
            continue
        meta = {
            "grupo": grupo,
            "tipo": _get(row, "TIPO_PRODUTO").upper(),
            "ncm": normalize_ncm(_get(row, "NCM")),
            "cest": normalize_cest(_get(row, "CEST")),
            "embalagem": norm_text(_get(row, "EMBALAGEM")),
            "volume_ml": only_digits(_get(row, "VOLUME_ML")),
            "marca": _get(row, "MARCA_INTERNA"),
            "descricao": _get(row, "DESCRICAO_TECNICA"),
        }
        for key_col in ("CODIGO", "CODIGO_BASE"):
            cod = normalize_item_join_key(_get(row, key_col))
            if cod:
                sim[cod] = meta
    _SIMILARITY_CACHE[cache_key] = sim
    return sim

def _precos_ref_file_candidates(config):
    fname = clean_str(config.get("arquivo_precos_referencia", "precos_referencia.xlsx")) or "precos_referencia.xlsx"
    aux_dir = Path(clean_str(config.get("pasta_auxiliares", "input/auxiliares")))
    if not aux_dir.is_absolute():
        aux_dir = BASE_DIR / aux_dir
    # Nome oficial + nomes recebidos por download/ajuste, sem depender de renomeação manual.
    candidates = [aux_dir / fname, INPUT_DIR / fname, BASE_DIR / fname]
    for folder in [aux_dir, INPUT_DIR, BASE_DIR]:
        if folder.exists():
            candidates.extend(sorted(folder.glob("precos_referencia*.xlsx")))
            candidates.extend(sorted(folder.glob("PRECOS_REFERENCIA*.xlsx")))
    seen = set(); out = []
    for p in candidates:
        rp = str(p.resolve()) if p.exists() else str(p)
        if rp not in seen:
            seen.add(rp); out.append(p)
    return out

def _find_header_row_for_precos_ref(preview_df):
    obrig_cod = {"CODIGO", "CODIGOPRODUTO", "CODIGOITEM", "CODPRODUTO", "CODITEM"}
    obrig_preco = {"MAIORPRECOVENDA", "PRECOREFERENCIA", "PRECOVENDA", "PRECOUNITARIO", "PRECOUNIT", "PRECO"}
    for idx in range(min(len(preview_df), 20)):
        vals = {norm_text(v) for v in preview_df.iloc[idx].tolist() if clean_str(v)}
        if vals & obrig_cod and vals & obrig_preco:
            return idx
    return 0

def _read_precos_ref_excel(path):
    """
    Lê precos_referencia.xlsx com detecção automática do cabeçalho real.

    Utiliza um contexto para `pd.ExcelFile` para garantir que o arquivo
    subjacente seja fechado imediatamente após o uso, evitando alocar
    descritores de arquivo indefinidamente.  Também valida a presença
    opcional das colunas RESPONSAVEL e DATA_DECISAO, registrando um aviso
    quando estas não forem localizadas, uma vez que tais informações
    auxiliam a rastreabilidade e auditoria dos preços de referência.
    """
    errors = []
    with pd.ExcelFile(path, engine=excel_engine(path)) as xls:
        for sheet in xls.sheet_names:
            try:
                preview = pd.read_excel(
                    path,
                    sheet_name=sheet,
                    header=None,
                    dtype=object,
                    keep_default_na=False,
                    nrows=25,
                    engine=excel_engine(path),
                )
                header_idx = _find_header_row_for_precos_ref(preview)
                df = pd.read_excel(
                    path,
                    sheet_name=sheet,
                    header=header_idx,
                    dtype=object,
                    keep_default_na=False,
                    engine=excel_engine(path),
                )
                cols = {norm_text(c) for c in df.columns}
                has_codigo = {"CODIGO", "CODIGOPRODUTO", "CODIGOITEM", "CODPRODUTO", "CODITEM"} & cols
                has_preco = {"MAIORPRECOVENDA", "PRECOREFERENCIA", "PRECOVENDA", "PRECOUNITARIO", "PRECOUNIT", "PRECO"} & cols
                if has_codigo and has_preco:
                    # Validação opcional de rastreabilidade do preço de referência
                    missing_traces = []
                    if not ({"RESPONSAVEL", "RESPONSÁVEL", "RESPONSAVELPRECO"} & cols):
                        missing_traces.append("RESPONSAVEL")
                    if not ({"DATADECISAO", "DATA_DECISAO", "DATADECISÃO"} & cols):
                        missing_traces.append("DATA_DECISAO")
                    if missing_traces:
                        logger.warning(
                            "precos_referencia.xlsx: colunas ausentes de rastreabilidade %s. Recomenda-se adicionar RESPONSAVEL e DATA_DECISAO para cada preço de referência.",
                            ", ".join(missing_traces),
                        )
                    return df, sheet, header_idx + 1
            except Exception as exc:
                errors.append(f"{sheet}: {exc}")
    raise ValueError(
        "Nenhuma aba válida em precos_referencia.xlsx. "
        + "; ".join(errors[:3])
    )

def load_precos_referencia(config):
    """Carrega tabela de preços de referência manuais como último fallback auditável.

    Regras de produção:
      - detecta cabeçalho real mesmo quando o modelo tem título/instrução nas linhas 1 e 2;
      - ignora STATUS = EXEMPLO, PENDENTE, BLOQUEADO, INATIVO ou equivalente;
      - exige CODIGO, MAIOR_PRECO_VENDA e FONTE_PRECO;
      - aceita EMPRESA/CNPJ_EMPRESA opcional para restringir a referência por CNPJ;
      - registra SHA-256 em RIR70_SOURCE_CONTROL_CODES e metadados no config;
      - nunca substitui preço real de venda já localizado: é aplicado apenas quando todos os fallbacks falham.
    """
    if not _is_sim(config.get("usar_precos_referencia", True)):
        config["_precos_referencia_info"] = {"arquivo": "", "linhas_lidas": 0, "linhas_usadas": 0, "situacao": "DESABILITADO"}
        return {}

    path = next((p for p in _precos_ref_file_candidates(config) if p.exists()), None)
    if path is None:
        config["_precos_referencia_info"] = {"arquivo": "precos_referencia.xlsx", "linhas_lidas": 0, "linhas_usadas": 0, "situacao": "NAO_LOCALIZADO"}
        return {}

    try:
        st = path.stat()
        cache_key = (str(path.resolve()), st.st_mtime_ns, st.st_size)
    except Exception:
        cache_key = (str(path), 0, 0)

    if cache_key in _PRECO_REF_CACHE:
        cached_ref, cached_info = _PRECO_REF_CACHE[cache_key]
        config["_precos_referencia_info"] = dict(cached_info)
        return cached_ref

    try:
        df, sheet_name, header_row = _read_precos_ref_excel(path)
    except Exception as exc:
        logger.warning("precos_referencia: falha na leitura (%s): %s", path.name, exc)
        config["_precos_referencia_info"] = {"arquivo": path.name, "linhas_lidas": 0, "linhas_usadas": 0, "situacao": "ERRO_LEITURA", "erro": str(exc)}
        return {}

    cols = {norm_text(c): c for c in df.columns}
    def _col(*names):
        for n in names:
            c = cols.get(norm_text(n))
            if c:
                return c
        return None

    cod_col   = _col("CODIGO","CODIGO_PRODUTO","CODIGO_ITEM","COD_PRODUTO","COD_ITEM")
    preco_col = _col("MAIOR_PRECO_VENDA","PRECO_REFERENCIA","PRECO_VENDA","PRECO_UNITARIO","PRECO_UNIT","PRECO")
    fonte_col = _col("FONTE_PRECO","FONTE","DOCUMENTO","NF_REFERENCIA","EVIDENCIA")
    status_col= _col("STATUS","SITUACAO","STATUS_PRECO","SITUACAO_PRECO")
    emp_col   = _col("EMPRESA","CNPJ_EMPRESA","CNPJ","CNPJ_EMITENTE")
    ncm_col   = _col("NCM")
    comp_col  = _col("COMPETENCIA","MES","PERIODO","MES_ANO","MESANO")
    just_col  = _col("JUSTIFICATIVA","MOTIVO","OBSERVACAO","OBS")

    if not cod_col or not preco_col or not fonte_col:
        logger.warning("precos_referencia: colunas obrigatórias não localizadas em '%s'. Necessário: CODIGO, MAIOR_PRECO_VENDA e FONTE_PRECO.", path.name)
        config["_precos_referencia_info"] = {"arquivo": path.name, "linhas_lidas": int(len(df)), "linhas_usadas": 0, "situacao": "COLUNAS_OBRIGATORIAS_AUSENTES"}
        RIR70_SOURCE_CONTROL_CODES[path.name] = sha256_file(path)
        return {}

    status_bloqueado = {"EXEMPLO", "PENDENTE", "BLOQUEADO", "BLOQUEADA", "INATIVO", "INATIVA", "NAO", "NAOAPROVADO", "NAOAPROVADA", "NÃO", "N"}
    ref: dict = {}
    linhas_lidas = int(len(df)); linhas_ok = 0; linhas_desc = 0

    for _, row in df.iterrows():
        status = norm_text(row.get(status_col, "")) if status_col else ""
        if status in status_bloqueado:
            linhas_desc += 1
            continue
        cod_raw = clean_str(row.get(cod_col, ""))
        if not cod_raw:
            linhas_desc += 1
            continue
        preco = to_decimal(row.get(preco_col, ""))
        if preco is None or preco <= ZERO:
            linhas_desc += 1
            continue
        fonte = clean_str(row.get(fonte_col, "")) if fonte_col else ""
        if not fonte:
            linhas_desc += 1
            continue
        cod_key = normalize_item_join_key(cod_raw)
        emp = normalize_cnpj(row.get(emp_col, "")) if emp_col else ""
        ncm = normalize_ncm(row.get(ncm_col, "")) if ncm_col else ""
        comp = clean_str(row.get(comp_col, "")) if comp_col else ""
        just = clean_str(row.get(just_col, "")) if just_col else ""
        if comp:
            if not re.match(r"^(\d{2})/(\d{4})$", comp):
                dt = parse_date(comp)
                comp = f"{dt.month:02d}/{dt.year}" if dt else ""
        entry = {
            "codigo": cod_raw,
            "empresa": emp,
            "ncm": ncm,
            "comp": comp,
            "preco": q6(preco),
            "base_70": q6(preco * Decimal("0.70")),
            "fonte": fonte,
            "just": just,
            "status": clean_str(row.get(status_col, "")) if status_col else "",
        }
        # Hierarquia: CNPJ específico prevalece; vazio funciona como referência geral.
        for key in [
            (emp, cod_key, ncm, comp), (emp, cod_key, ncm, ""), (emp, cod_key, "", comp), (emp, cod_key, "", ""),
            ("", cod_key, ncm, comp), ("", cod_key, ncm, ""), ("", cod_key, "", comp), ("", cod_key, "", ""),
        ]:
            if key[1]:
                ref.setdefault(key, entry)
        linhas_ok += 1

    codigo_controle = sha256_file(path)
    RIR70_SOURCE_CONTROL_CODES[path.name] = codigo_controle
    info = {
        "arquivo": path.name,
        "aba": sheet_name,
        "header_row": header_row,
        "linhas_lidas": linhas_lidas,
        "linhas_usadas": linhas_ok,
        "linhas_descartadas": linhas_desc,
        "codigo_controle": codigo_controle,
        "situacao": "OK" if linhas_ok else "SEM_LINHAS_VALIDAS",
    }
    config["_precos_referencia_info"] = info
    logger.info("precos_referencia: %d linhas válidas de %d em '%s' (aba=%s, header=%s).", linhas_ok, linhas_lidas, path.name, sheet_name, header_row)
    _PRECO_REF_CACHE[cache_key] = (ref, info)
    return ref


def _find_preco_referencia(cod, ncm, comp, ref_prices, empresa=""):
    """Lookup de preço de referência com especificidade: empresa, código, NCM e competência."""
    if not ref_prices:
        return None
    cod_key = normalize_item_join_key(cod)
    ncm_n = normalize_ncm(ncm) if ncm else ""
    emp_n = normalize_cnpj(empresa) if empresa else ""
    for key in [
        (emp_n, cod_key, ncm_n, comp), (emp_n, cod_key, ncm_n, ""), (emp_n, cod_key, "", comp), (emp_n, cod_key, "", ""),
        ("", cod_key, ncm_n, comp), ("", cod_key, ncm_n, ""), ("", cod_key, "", comp), ("", cod_key, "", ""),
    ]:
        if key[1] and key in ref_prices:
            return ref_prices[key]
    return None


def _build_similarity_indexes(arb_rows, sim_map):
    same_cnpj_current = defaultdict(list)   # (cnpj, grupo, comp) -> rows
    same_cnpj_history = defaultdict(list)   # (cnpj, grupo) -> rows
    other_current = defaultdict(list)       # (grupo, comp) -> rows
    other_history = defaultdict(list)       # grupo -> rows
    for row in arb_rows:
        cod_key = normalize_item_join_key(row.get("Código Item", ""))
        meta = sim_map.get(cod_key)
        if not meta:
            continue
        base = row.get("Base Arbitrada (70%)")
        if not isinstance(base, Decimal):
            continue
        cnpj = normalize_cnpj(row.get("CNPJ Emitente", ""))
        comp = clean_str(row.get("Competência", ""))
        grupo = meta.get("grupo", "")
        if not cnpj or not comp or not grupo:
            continue
        same_cnpj_current[(cnpj, grupo, comp)].append(row)
        same_cnpj_history[(cnpj, grupo)].append(row)
        other_current[(grupo, comp)].append(row)
        other_history[grupo].append(row)
    for rows in list(same_cnpj_current.values()) + list(same_cnpj_history.values()) + list(other_current.values()) + list(other_history.values()):
        rows.sort(key=lambda r: (competencia_sort_key(r.get("Competência", "")), clean_str(r.get("Código Item", ""))))
    return same_cnpj_current, same_cnpj_history, other_current, other_history

def _find_previous_similarity(rows_hist, target_comp, target_cod_key):
    target_key = competencia_sort_key(target_comp)
    best = None
    for cand in rows_hist:
        ck = competencia_sort_key(cand.get("Competência", ""))
        if ck < target_key and isinstance(cand.get("Base Arbitrada (70%)"), Decimal) and normalize_item_join_key(cand.get("Código Item","")) != target_cod_key:
            best = cand
        elif ck >= target_key:
            break
    return best

def _first_valid_similar(candidates, target_cod_key, cnpjs_grupo=None, target_cnpj=None):
    grupo_set = {normalize_cnpj(c) for c in (cnpjs_grupo or []) if normalize_cnpj(c)}
    for cand in sorted(candidates, key=lambda r: clean_str(r.get("Código Item", ""))):
        cand_cod = normalize_item_join_key(cand.get("Código Item", ""))
        cand_cnpj = normalize_cnpj(cand.get("CNPJ Emitente", ""))
        if cand_cod == target_cod_key:
            continue
        if target_cnpj and cand_cnpj == normalize_cnpj(target_cnpj):
            pass
        if grupo_set and cand_cnpj and cand_cnpj not in grupo_set:
            continue
        if isinstance(cand.get("Base Arbitrada (70%)"), Decimal):
            return cand
    return None


def build_inventario_origem_rows(inv_rows, logs=None):
    grouped = {}
    for r in inv_rows:
        key = (r.get("CNPJ Inventário", ""), r.get("Competência", ""))
        g = grouped.setdefault(key, {"itens": 0, "calc": 0, "sem": 0, "prev": 0, "similar": 0, "other": 0,
                                    "qtd_orig": ZERO, "qtd_trib": ZERO, "valor": ZERO})
        g["itens"] += 1
        st = clean_str(r.get("Status", ""))
        origem = clean_str(r.get("Origem Base", ""))
        if "CALCULADO" in st: g["calc"] += 1
        if "SEM BASE" in st: g["sem"] += 1
        if "COMPETÊNCIA ANTERIOR" in st or "COMPETÊNCIA ANTERIOR" in origem: g["prev"] += 1
        if "PRODUTO SIMILAR" in st or "PRODUTO SIMILAR" in origem: g["similar"] += 1
        if "OUTRO CNPJ" in st or "OUTRO CNPJ" in origem: g["other"] += 1
        q_orig = r.get("QTDE Inventário")
        q_trib = r.get("QTDE Inventário Trib.")
        val = r.get("Valor Inventário PA")
        if isinstance(q_orig, Decimal): g["qtd_orig"] += q_orig
        if isinstance(q_trib, Decimal): g["qtd_trib"] += q_trib
        if isinstance(val, Decimal): g["valor"] += val
    rows = []
    for (cnpj, comp), g in grouped.items():
        dif = g["qtd_trib"] - g["qtd_orig"]
        if g["sem"]:
            status = "REVISAR - ITENS SEM BASE ATENCAO"
        elif g["other"]:
            status = "REVISAR - BASE DE OUTRO CNPJ ATENCAO"
        elif g["similar"]:
            status = "REVISAR - BASE DE PRODUTO SIMILAR HOMOLOGADO ATENCAO"
        elif g["prev"]:
            status = "REVISAR - BASE DA COMPETÊNCIA ANTERIOR ATENCAO"
        else:
            status = "CONCILIADO OK"
        rows.append({
            "CNPJ Inventário": cnpj,
            "Competência": comp,
            "Itens Inventário": g["itens"],
            "Itens Calculados": g["calc"],
            "Itens sem Base": g["sem"],
            "Itens com Base Competência Anterior": g["prev"],
            "Itens com Base de Produto Similar": g["similar"],
            "Itens com Base de Outro CNPJ": g["other"],
            "Qtde Inv. Original": q3(g["qtd_orig"]),
            "Qtde Inv. Trib.": q3(g["qtd_trib"]),
            "Dif. Conversão": q3(dif),
            "Valor Inventário PA": q2(g["valor"]),
            "Status Auditoria": status,
        })
    rows.sort(key=lambda r: (clean_str(r.get("CNPJ Inventário", "")), competencia_sort_key(r.get("Competência", ""))))
    return rows

def build_inventario_pivot(inv_rows):
    comps = sorted({clean_str(r.get("Competência", "")) for r in inv_rows if clean_str(r.get("Competência", ""))}, key=competencia_sort_key)
    grouped = {}
    for r in inv_rows:
        key = (r.get("CNPJ Inventário", ""), r.get("Código Item", ""), r.get("NCM", ""), r.get("Descrição", ""),
               r.get("Unid. Inventário", ""), r.get("Fator Unidade", ""), r.get("Unid. Trib.", ""))
        g = grouped.setdefault(key, {"CNPJ Inventário": r.get("CNPJ Inventário", ""),
                                    "Código Item": r.get("Código Item", ""),
                                    "NCM": r.get("NCM", ""),
                                    "Descrição": r.get("Descrição", ""),
                                    "Unid. Inventário": r.get("Unid. Inventário", ""),
                                    "Fator Unid.": r.get("Fator Unidade", ""),
                                    "Unid. Trib.": r.get("Unid. Trib.", ""),
                                    "_meses": {}})
        comp = clean_str(r.get("Competência", ""))
        if comp:
            g["_meses"][comp] = {
                "qtd_orig": r.get("QTDE Inventário"),
                "qtd_trib": r.get("QTDE Inventário Trib."),
                "base": r.get("Base Arbitrada (70%)"),
                "valor": r.get("Valor Inventário PA"),
                "origem": r.get("Origem Base"),
                "cnpj_base": r.get("CNPJ Base Arbitrada"),
                "comp_base": r.get("Competência Base Arbitrada"),
                "status": r.get("Status"),
            }
    rows = list(grouped.values())
    rows.sort(key=lambda r: (clean_str(r.get("CNPJ Inventário", "")), clean_str(r.get("Código Item", "")), clean_str(r.get("Descrição", ""))))
    return rows, comps


def _write_inventario_pivot(writer, wb, periodo_txt, inv_rows, fmts):
    """Aba 03_INVENTARIO_PA em layout horizontal: produto × competência.

    A quantidade exibida é tributável, já convertida de CX/FD/PCT/PACK para UN
    quando o fator de embalagem for identificado. O status por competência mostra
    se a base veio do mesmo CNPJ, de outro CNPJ ou se não houve base arbitrada.
    """
    rows, comps = build_inventario_pivot(inv_rows)
    N = len(comps)
    FIXED = ["CNPJ Inventário", "Código Item", "NCM", "Descrição", "Unid. Inventário", "Fator Unid.", "Unid. Trib."]
    FIXED_W = [18, 16, 10, 48, 13, 11, 10]
    SUB = ["Qtde Trib.", "Base 70%", "Valor PA", "Origem/Status"]
    SUB_W = [14, 13, 15, 26]
    n_fixed = len(FIXED)
    total_ci = n_fixed + N * len(SUB)
    last_col = total_ci

    def _f(spec): return wb.add_format(spec)
    _base = {"font_name": "Arial Narrow", "font_size": 10, "valign": "vcenter"}
    _header = {"font_name": "Arial Narrow", "font_size": 11, "valign": "vcenter"}
    hdr_group = _f({**_header, "bold": True, "align": "center", "bg_color": "#0D2137", "font_color": "#FFFFFF", "border": 1, "border_color": "#1A3A5C", "text_wrap": True})
    hdr_sub = _f({**_header, "bold": True, "align": "center", "bg_color": "#1A3A5C", "font_color": "#D9EAF7", "border": 1, "border_color": "#1E5C8A", "text_wrap": True})
    hdr_total = _f({**_header, "bold": True, "align": "center", "bg_color": "#14532D", "font_color": "#FFFFFF", "border": 1, "border_color": "#0A3018", "text_wrap": True})
    title_left = _f({"font_name": "Arial Narrow", "font_size": 13, "bold": True, "align": "left", "valign": "vcenter", "bg_color": "#0D2137", "font_color": "#FFFFFF", "left": 4, "left_color": "#1A3A5C"})
    subtitle = _f({"font_name": "Arial Narrow", "font_size": 10, "italic": True, "align": "left", "valign": "vcenter", "bg_color": "#EAF2F8", "font_color": "#1A3A5C", "text_wrap": True, "border": 1, "border_color": "#1A3A5C"})
    _txt = _f({**_base}); _txt_a = _f({**_base, "bg_color": SILVER})
    _ctr = _f({**_base, "align": "center"}); _ctr_a = _f({**_base, "align": "center", "bg_color": SILVER})
    _qtd = _f({**_base, "num_format": "#,##0.0000", "align": "right"}); _qtd_a = _f({**_base, "num_format": "#,##0.0000", "align": "right", "bg_color": SILVER})
    _num = _f({**_base, "num_format": "R$ #,##0.0000", "align": "right"}); _num_a = _f({**_base, "num_format": "R$ #,##0.0000", "align": "right", "bg_color": SILVER})
    _cur = _f({**_base, "num_format": "R$ #,##0.00", "align": "right"}); _cur_a = _f({**_base, "num_format": "R$ #,##0.00", "align": "right", "bg_color": SILVER})
    _status_base = {**_base, "font_size": 8, "align": "center", "bold": True, "text_wrap": True}
    _ok = _f({**_status_base, "bg_color": GREEN_BG, "font_color": GREEN_FG})
    _warn = _f({**_status_base, "bg_color": AMBER_BG, "font_color": AMBER_FG})
    _bad = _f({**_status_base, "bg_color": RED_BG, "font_color": RED_FG})
    _tot_qty = _f({**_base, "num_format": "#,##0.0000", "align": "right", "bg_color": "#14532D", "font_color": "#FFFFFF", "bold": True})
    _tot_cur = _f({**_base, "num_format": "R$ #,##0.00", "align": "right", "bg_color": "#0D2137", "font_color": "#FFFFFF", "bold": True})

    ws = wb.add_worksheet("03_INVENTARIO_PA")
    ws.set_tab_color("#14532D")
    writer.sheets["03_INVENTARIO_PA"] = ws
    ws.hide_gridlines(2)
    ws.set_zoom(90)
    for ci, w in enumerate(FIXED_W): ws.set_column(ci, ci, w)
    for n in range(N):
        for s, w in enumerate(SUB_W): ws.set_column(n_fixed + n * len(SUB) + s, n_fixed + n * len(SUB) + s, w)
    ws.set_column(total_ci, total_ci, 16)

    ws.merge_range(0, 0, 0, last_col, f"Inventário PA - Quantidade Tributável × Base Arbitrada 70%  |  {periodo_txt}", title_left)
    ws.set_row(0, 26)
    ws.merge_range(1, 0, 1, last_col, "Leitura fiscal: quantidade de inventário convertida para unidade tributável. CX6, CX12 e CX24 são multiplicadas pelo fator da embalagem. Origem da base indica se o preço veio do mesmo CNPJ ou de outro CNPJ.", subtitle)
    ws.set_row(1, 28)

    ws.merge_range(2, 0, 2, n_fixed - 1, "Identificação do produto em estoque", hdr_group)
    for n, comp in enumerate(comps):
        c0 = n_fixed + n * len(SUB)
        ws.merge_range(2, c0, 2, c0 + len(SUB) - 1, comp, hdr_group)
    ws.write(2, total_ci, "Total", hdr_total)
    ws.set_row(2, 24)
    for ci, col in enumerate(FIXED): ws.write(3, ci, col, hdr_sub)
    for n in range(N):
        c0 = n_fixed + n * len(SUB)
        for s, sub in enumerate(SUB): ws.write(3, c0 + s, sub, hdr_sub)
    ws.write(3, total_ci, "Valor PA Total", hdr_total)
    ws.set_row(3, 32)

    col_qty_totals = [Decimal("0")] * N
    col_val_totals = [Decimal("0")] * N
    grand_total = Decimal("0")
    for ri, prod in enumerate(rows, start=4):
        alt = (ri % 2 == 0)
        ws.set_row(ri, 18)
        fixed_values = [prod.get(k, "") for k in FIXED]
        for ci, val in enumerate(fixed_values):
            fmt = _txt_a if (alt and ci == 3) else (_txt if ci == 3 else (_ctr_a if alt else _ctr))
            ws.write(ri, ci, val, fmt)
        row_total = Decimal("0")
        for n, comp in enumerate(comps):
            c0 = n_fixed + n * len(SUB)
            data = prod.get("_meses", {}).get(comp)
            if not data:
                for s in range(len(SUB)): ws.write_blank(ri, c0 + s, None, _txt_a if alt else _txt)
                continue
            qtd = data.get("qtd_trib")
            base = data.get("base")
            valor = data.get("valor")
            status = clean_str(data.get("status", ""))
            origem = clean_str(data.get("origem", ""))
            cnpj_base = clean_str(data.get("cnpj_base", ""))
            comp_base = clean_str(data.get("comp_base", ""))
            if isinstance(qtd, Decimal):
                ws.write_number(ri, c0, float(qtd), _qtd_a if alt else _qtd); col_qty_totals[n] += qtd
            else: ws.write_blank(ri, c0, None, _qtd_a if alt else _qtd)
            if isinstance(base, Decimal): ws.write_number(ri, c0 + 1, float(base), _num_a if alt else _num)
            else: ws.write_blank(ri, c0 + 1, None, _num_a if alt else _num)
            if isinstance(valor, Decimal):
                ws.write_number(ri, c0 + 2, float(valor), _cur_a if alt else _cur); row_total += valor; col_val_totals[n] += valor
            else: ws.write_blank(ri, c0 + 2, None, _cur_a if alt else _cur)
            status_txt = status
            extras = []
            if cnpj_base and "OUTRO CNPJ" in status:
                extras.append(f"Base CNPJ: {cnpj_base}")
            if comp_base and comp_base != comp:
                extras.append(f"Comp. base: {comp_base}")
            if extras:
                status_txt = f"{status}\n" + " | ".join(extras)
            fmt_status = _bad if "SEM BASE" in status_txt else (_warn if "OUTRO CNPJ" in status_txt or "FALLBACK" in origem or "ANTERIOR" in status_txt else _ok)
            ws.write(ri, c0 + 3, status_txt, fmt_status)
        grand_total += row_total
        ws.write_number(ri, total_ci, float(row_total), _cur_a if alt else _cur)

    tot_row = 4 + len(rows)
    ws.set_row(tot_row, 20)
    ws.merge_range(tot_row, 0, tot_row, n_fixed - 1, "TOTAL POR COMPETÊNCIA", hdr_total)
    for n in range(N):
        c0 = n_fixed + n * len(SUB)
        # Por decisão de auditoria visual, não totalizar Qtde Trib. na aba 03:
        # a quantidade tributável é apenas métrica de conferência linha a linha.
        ws.write_blank(tot_row, c0, None, _tot_cur)
        ws.write_blank(tot_row, c0 + 1, None, _tot_cur)
        ws.write_number(tot_row, c0 + 2, float(col_val_totals[n]), _tot_cur)
        ws.write_blank(tot_row, c0 + 3, None, _tot_cur)
    ws.write_number(tot_row, total_ci, float(grand_total), _tot_cur)
    ws.freeze_panes(4, n_fixed)
    ws.autofilter(3, 0, 3 + len(rows), last_col)





def sha256_file(path):
    h=hashlib.sha256()
    try:
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(1024 * 1024), b""):
                h.update(chunk)
    except Exception:
        return ""
    return h.hexdigest()




# =============================================================================
# PATCH MICRO-CIRURGICO RIR70 - COMANDO CONGELADO 2026-05-14
# Mantem leitores e extratores existentes; substitui regras, nomes e saida final
# para cumprir o pacote enxuto: CAPA + 01 + 02 + 03 + 04 + 05.
# =============================================================================
CENT6 = Decimal("0.000001")
RIR70_DUPLICATES = []
RIR70_SOURCE_STATS = []
RIR70_SOURCE_CONTROL_CODES = {}  # arquivo -> codigo de controle do arquivo efetivamente lido
RIR70_LAST_INVENTORY_ROWS = []
RIR70_LAST_ARB_ROWS = []

def q6(v):
    return v.quantize(CENT6, rounding=ROUND_HALF_UP) if isinstance(v, Decimal) else None

def _month_range(inicio, fim):
    if not inicio or not fim:
        return []
    y, m = inicio.year, inicio.month
    out = []
    while (y, m) <= (fim.year, fim.month):
        out.append(f"{m:02d}/{y}")
        m += 1
        if m == 13:
            m = 1; y += 1
    return out

def _competencia_to_ym(mes):
    s = clean_str(mes)
    m = re.match(r"^(\d{2})/(\d{4})$", s)
    if not m:
        return None
    return int(m.group(2)), int(m.group(1))

def _competencia_diff_meses(origem, destino):
    a = _competencia_to_ym(origem); b = _competencia_to_ym(destino)
    if not a or not b:
        return None
    return (b[0] - a[0]) * 12 + (b[1] - a[1])

def _sha256_bytes(data):
    h = hashlib.sha256()
    h.update(data or b"")
    return h.hexdigest()

def _fmt_defasagem_preco(mes_origem, mes_destino, prefixo="PRECO_DE_MES_ANTERIOR"):
    diff = _competencia_diff_meses(mes_origem, mes_destino)
    if diff is None or diff <= 1:
        return "PRECO_USADO_DE_MES_ANTERIOR"
    return f"{prefixo}_COM_DEFASAGEM_DE_{diff}_MESES"

def _empresas_grupo_cnpjs(config):
    raw = config.get("empresas_grupo") or config.get("cnpjs_grupo") or []
    out = []
    for item in raw:
        c = item.get("cnpj") if isinstance(item, dict) else item
        c = normalize_cnpj(c)
        if c and c not in out:
            out.append(c)
    for c in config.get("cnpjs_grupo", []) or []:
        c = normalize_cnpj(c)
        if c and c not in out:
            out.append(c)
    return out

def _empresas_grupo_display(config):
    raw = config.get("empresas_grupo") or []
    out = []
    if isinstance(raw, list):
        for item in raw:
            if isinstance(item, dict):
                c = normalize_cnpj(item.get("cnpj")); r = clean_str(item.get("razao_social"))
                if c:
                    out.append(f"{c} - {r}" if r else c)
            else:
                c = normalize_cnpj(item)
                if c: out.append(c)
    return out or _empresas_grupo_cnpjs(config)

def _arquivo_periodo_nome(config):
    inicio = parse_config_date(config.get("data_inicial_calculo") or config.get("periodo_base_inicio"))
    fim = parse_config_date(config.get("data_final_calculo") or config.get("periodo_base_fim"))
    if not inicio or not fim:
        return "DE_DATA_INVALIDA_ATE_DATA_INVALIDA"
    return f"DE_{inicio:%d%m%Y}_ATE_{fim:%d%m%Y}"

def _novo_output_path(config):
    global OUTPUT_PATH, OUT_BASE, LOG_PATH, META_PATH, _OUTPUT_INITIALIZED
    if _OUTPUT_INITIALIZED:
        return OUTPUT_PATH
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = f"Arbitramento_Custo_RIR70_{_arquivo_periodo_nome(config)}_PROCESSADO_{stamp}"
    path = OUTPUT_DIR / f"{base}.xlsx"
    # Nunca sobrescreve arquivo anterior.
    seq = 1
    while path.exists():
        path = OUTPUT_DIR / f"{base}_{seq:02d}.xlsx"
        seq += 1
    OUT_BASE = path.stem
    OUTPUT_PATH = path
    LOG_PATH = OUTPUT_DIR / f"{OUT_BASE}.log"
    META_PATH = OUTPUT_DIR / f"{OUT_BASE}.metadados.json"
    _OUTPUT_INITIALIZED = True
    return path

def ensure_config():
    if not CONFIG_PATH.exists():
        CONFIG_PATH.write_text(safe_json_dumps(DEFAULT_CONFIG, indent=2), encoding="utf-8")
    with CONFIG_PATH.open("r", encoding="utf-8") as f:
        loaded = json.load(f)
    chaves_obsoletas = ["cfops_validos_venda", "cfops_bloqueados_calculo"]
    encontradas = [k for k in chaves_obsoletas if isinstance(loaded, dict) and k in loaded]
    if encontradas:
        raise ConfigError(
            "Chave(s) obsoleta(s) no config: " + ", ".join(encontradas) +
            ". Use exclusivamente CFOP_VENDA_CALCULO, CFOP_REVISAR e CFOP_NAO_CALCULA."
        )
    cfg = DEFAULT_CONFIG.copy(); cfg.update(loaded or {})

    cfg.setdefault("MODO_EXECUCAO", "Producao")
    modo_exec_norm = norm_text(cfg.get("MODO_EXECUCAO"))
    if modo_exec_norm not in {"PRODUCAO", "HOMOLOGACAO", "TESTE"}:
        raise ConfigError("MODO_EXECUCAO inválido. Use Producao, Homologacao ou Teste.")
    cfg["MODO_EXECUCAO"] = {"PRODUCAO":"Producao", "HOMOLOGACAO":"Homologacao", "TESTE":"Teste"}[modo_exec_norm]
    cfg.setdefault("MODO_GOLDEN_FILE", "Nao")
    cfg["MODO_GOLDEN_FILE"] = "Sim" if _is_sim(cfg.get("MODO_GOLDEN_FILE")) else "Nao"
    cfg["ADVERTENCIA_MODO_EXECUCAO"] = ""
    if cfg["MODO_EXECUCAO"] == "Producao" and cfg["MODO_GOLDEN_FILE"] == "Sim":
        cfg["MODO_GOLDEN_FILE"] = "Nao"
        cfg["ADVERTENCIA_MODO_EXECUCAO"] = "MODO_EXECUCAO=Producao prevaleceu sobre MODO_GOLDEN_FILE=Sim; golden file forçado para Nao."
    cfg.setdefault("validar_campos_obrigatorios_matriz", True)
    cfg.setdefault("validar_part_arbitramento_matriz", cfg.get("validar_participa_arbitr_matriz", True))

    cfg.setdefault("MODO_INTERATIVO", "NAO")
    cfg["MODO_INTERATIVO"] = "SIM" if clean_str(cfg.get("MODO_INTERATIVO")).upper() == "SIM" else "NAO"
    cfg.setdefault("perguntar_periodo_ao_iniciar", "Sim")
    cfg["perguntar_periodo_ao_iniciar"] = "Sim" if _is_sim(cfg.get("perguntar_periodo_ao_iniciar")) else "Nao"
    cfg["data_inicial_calculo"] = cfg.get("data_inicial_calculo") or cfg.get("periodo_base_inicio") or "01/01/2025"
    cfg["data_final_calculo"] = cfg.get("data_final_calculo") or cfg.get("periodo_base_fim") or "31/12/2025"
    # Mantem chaves antigas para preservar leitores existentes.
    cfg["periodo_base_inicio"] = cfg["data_inicial_calculo"]
    cfg["periodo_base_fim"] = cfg["data_final_calculo"]

    empresas = _empresas_grupo_cnpjs(cfg) or ["03408722000178", "03408722000330", "03408722000410", "03408722000763"]
    cfg["cnpjs_grupo"] = empresas
    if not cfg.get("empresas_grupo") or not isinstance(cfg.get("empresas_grupo"), list):
        cfg["empresas_grupo"] = empresas

    cfg.setdefault("CFOP_VENDA_CALCULO", ["5101","6101","5102","6102","5401","5403","5405","6401","6403","6404"])
    cfg.setdefault("CFOP_REVISAR", ["5110","6110","5116","5117","6116","6117","5404","6405"])
    cfg.setdefault("CFOP_NAO_CALCULA", ["5151","6151","5152","6152","5901","6901","5910","6910","5921","6921","5124","6124","5915","6915","5556","6556","5922","6922","5949","6949","7949","1201","2201","1411","2411"])
    cfg["CFOP_VENDA_CALCULO"] = [str(c).zfill(4) for c in cfg.get("CFOP_VENDA_CALCULO", [])]
    cfg["CFOP_REVISAR"] = [str(c).zfill(4) for c in cfg.get("CFOP_REVISAR", [])]
    cfg["CFOP_NAO_CALCULA"] = [str(c).zfill(4) for c in cfg.get("CFOP_NAO_CALCULA", [])]
    # Pre-computação dos frozensets usados em classify_cfop - evita recriação a cada linha processada
    cfg["_CFOP_VENDA_SET"]       = frozenset(cfg["CFOP_VENDA_CALCULO"])
    cfg["_CFOP_REVISAR_SET"]     = frozenset(cfg["CFOP_REVISAR"])
    cfg["_CFOP_NAO_SET"]         = frozenset(cfg["CFOP_NAO_CALCULA"])
    cfg["_CFOP_DEVOLUCAO_SET"]   = frozenset({"5201","5202","5410","5411","5660","5661","5662","6201","6202","6410","6411","6660","6661","6662","1201","1202","1410","1411","2201","2202","2410","2411"})
    cfg["_CFOP_TRANSFERENCIA_SET"] = frozenset({"5151","5152","5408","5409","6151","6152","6408","6409"})
    cfg["_CFOP_BONIFICACAO_SET"] = frozenset({"5910","6910","5929","6929"})
    cfg["_CNPJS_GRUPO_SET"]      = frozenset(normalize_cnpj(c) for c in (cfg.get("cnpjs_grupo") or []) if normalize_cnpj(c))
    cfg.setdefault("minimo_vendas_para_alerta", 3)
    cfg.setdefault("preco_acima_padrao_percentual", 50)
    cfg.setdefault("preco_muito_acima_padrao_percentual", 100)
    cfg.setdefault("justificativa_alerta_preco", "Parâmetro operacional de auditoria interna, sem natureza normativa.")
    cfg.setdefault("codigo_controle_arquivo", "SHA-256")
    cfg.setdefault("codigo_controle_zip_xml", "SHA-256_DO_XML_EXTRAIDO")
    cfg.setdefault("bloquear_qtd_calculo_zero_ou_negativa", True)
    cfg.setdefault("media_preco_tipo", "PONDERADA_POR_QTD_CALCULO")
    cfg.setdefault("formula_diferenca_percentual_preco", "(MAIOR_PRECO_UNITARIO - PRECO_MEDIO) / PRECO_MEDIO * 100")
    cfg.setdefault("calcula_revisar_entra_automaticamente", False)
    cfg.setdefault("tipo_item_fonte_prioridade", ["MATRIZ_CADASTRO", "REGRA_OPERACIONAL", "CONFIG_NCM", "DESCRICAO"])
    cfg.setdefault("usar_xml_como_movimentacao", True)
    cfg.setdefault("usar_zip_xml", True)
    cfg.setdefault("usar_leitor_xlsx_streaming", True)
    cfg.setdefault("permitir_documentos_em_input_raiz", False)
    cfg.setdefault("permitir_documentos_em_movimento_item", False)
    cfg.setdefault("segregar_calculo_por_cnpj_emitente", True)
    cfg.setdefault("deduzir_icms_do_valor_produto", False)
    cfg.setdefault("deduzir_ipi_do_valor_produto_bruto", False)
    cfg.setdefault("tratamento_ipi_valor_produto", "vprod_ja_sem_ipi")
    cfg.setdefault("formato_data", "DD/MM/AAAA")
    cfg.setdefault("fonte_excel", "Arial Narrow")
    cfg.setdefault("tamanho_fonte_excel", 10)
    cfg.setdefault("exibir_colunas_diagnostico", "Nao")
    cfg["exibir_colunas_diagnostico"] = "Sim" if _is_sim(cfg.get("exibir_colunas_diagnostico")) else "Nao"
    cfg.setdefault("max_rows_matriz", None)
    cfg.setdefault("tipo_produto_validos", ["PA", "MR"])
    cfg.setdefault("grupos_heuristica_tipo_produto", ["03408722"])
    # A partir da versão 1.2, o comportamento padrão é bloquear a execução
    # quando não houver cadastro completo na matriz de produtos.  Sem esse bloqueio,
    # itens podem ser processados sem validação de PART_ARBITRAMENTO, VINCULO_GRUPO
    # ou CATEGORIA_ITEM, o que gera divergências em auditorias externas.  O valor
    # pode ser sobrescrito explicitamente no arquivo de configuração para ambientes
    # de teste ou depuração.
    cfg.setdefault("bloquear_sem_cadastro_matriz", True)
    cfg.setdefault("validar_status_cadastro_matriz", True)
    cfg.setdefault("validar_participa_arbitr_matriz", True)
    cfg.setdefault("usar_fator_unidade_matriz_inventario", True)
    cfg.setdefault("matriz_sheets_validas", ["MATRIZ_CADASTRO", "MATRIZ_SIMILARIDADE", "MAPA_SIMILARIDADE"])
    cfg.setdefault("permitir_intercompany_por_raiz_cnpj", False)
    cfg.setdefault("ncms_arbitramento", NCM_ABA03_PERMITIDOS.copy())
    cfg.setdefault("ncms_inventario_pa", NCM_ABA03_PERMITIDOS.copy())
    cfg.setdefault("ncms_tipo_item_rir70", NCM_TIPO_ITEM_RIR70.copy())
    cfg.setdefault("usar_precos_referencia", True)
    cfg.setdefault("arquivo_precos_referencia", "precos_referencia.xlsx")
    cfg.setdefault("pasta_ajustes_analista", "input/ajustes")
    cfg.setdefault("arquivo_ajustes_analista", "AJUSTES_ANALISTA_RIR70.xlsx")
    cfg.setdefault("gerar_template_ajustes_analista", True)
    cfg.setdefault("criar_arquivo_ajustes_analista_se_ausente", True)
    cfg.setdefault("gerar_links_analista", True)
    cfg.setdefault("validar_criterio_custeio_rir70", True)
    cfg.setdefault("bloquear_custo_aquisicao_sem_rotina", True)
    cfg.setdefault("alertar_cnpj_raiz_grupo_nao_cadastrado", True)
    cfg.setdefault("gerar_diagnostico_matriz_rir70", True)
    cfg.setdefault("MODO_REPROCESSAMENTO", "COMPLETO")
    cfg.setdefault("preservar_ajustes_analista_existentes", True)

    # Em ambiente de produção, deduzir ICMS do valor do produto não é permitido
    # por violar o art. 308, par.1º do RIR/2018 e gerar base de cálculo incorreta.
    if cfg.get("MODO_EXECUCAO") == "Producao" and bool(cfg.get("deduzir_icms_do_valor_produto")):
        raise ConfigError(
            "deduzir_icms_do_valor_produto=True não é permitido em produção; o motor sempre inclui ICMS na base de cálculo conforme o RIR/2018."
        )
    # Normaliza o nome da chave de limite de meses retroativos.  Se a versão
    # minúscula for utilizada no arquivo de configuração, copia seu valor
    # para a versão em maiúsculas para consistência interna.
    if "max_meses_retroativos_preco" in cfg and "MAX_MESES_RETROATIVOS_PRECO" not in cfg:
        cfg["MAX_MESES_RETROATIVOS_PRECO"] = cfg["max_meses_retroativos_preco"]
    try:
        max_retro = int(cfg.get("MAX_MESES_RETROATIVOS_PRECO") or 12)
    except Exception:
        max_retro = 12
    cfg["MAX_MESES_RETROATIVOS_PRECO"] = max(0, max_retro)
    cfg["max_meses_retroativos_preco"] = cfg["MAX_MESES_RETROATIVOS_PRECO"]
    cfg.setdefault("validar_erros_excel_output", "Sim")
    cfg.setdefault("max_celulas_validacao_output", 0)
    cfg.setdefault("empresas_sem_movimento_confirmadas", [])
    cfg.setdefault("empresas_sem_estoque_confirmadas", [])
    cfg.setdefault("gerar_relatorio_sem_movimento", True)
    return cfg

def classify_cfop(cfop, config):
    code = normalize_cfop(cfop)
    if not code:
        return "CFOP_NAO_CLASSIFICADO", False, "CFOP_NAO_CLASSIFICADO"
    # Usa frozensets pré-computados por ensure_config; fallback inline se config vier sem eles (testes unitários)
    venda        = config.get("_CFOP_VENDA_SET")        or frozenset(str(c).zfill(4) for c in config.get("CFOP_VENDA_CALCULO", []))
    revisar      = config.get("_CFOP_REVISAR_SET")      or frozenset(str(c).zfill(4) for c in config.get("CFOP_REVISAR", []))
    nao          = config.get("_CFOP_NAO_SET")          or frozenset(str(c).zfill(4) for c in config.get("CFOP_NAO_CALCULA", []))
    devolucao    = config.get("_CFOP_DEVOLUCAO_SET")    or frozenset({"5201","5202","5410","5411","5660","5661","5662","6201","6202","6410","6411","6660","6661","6662","1201","1202","1410","1411","2201","2202","2410","2411"})
    transferencia= config.get("_CFOP_TRANSFERENCIA_SET") or frozenset({"5151","5152","5408","5409","6151","6152","6408","6409"})
    bonificacao  = config.get("_CFOP_BONIFICACAO_SET")  or frozenset({"5910","6910","5929","6929"})
    if code in venda:
        return "VENDA_VALIDA", True, "VENDA_VALIDA"
    if code in revisar:
        return "CFOP_REVISAR", False, "CFOP_NAO_CLASSIFICADO_COMO_VENDA"
    if code in nao:
        if code in transferencia: return "TRANSFERENCIA", False, "TRANSFERENCIA"
        if code in bonificacao: return "BONIFICACAO", False, "BONIFICACAO"
        if code in devolucao: return "DEVOLUCAO", False, "DEVOLUCAO"
        return "CFOP_NAO_E_VENDA", False, "CFOP_NAO_E_VENDA"
    if code in devolucao:
        return "DEVOLUCAO", False, "DEVOLUCAO"
    if code in transferencia:
        return "TRANSFERENCIA", False, "TRANSFERENCIA"
    if code in bonificacao:
        return "BONIFICACAO", False, "BONIFICACAO"
    if code.startswith(("5","6")):
        return "CFOP_REVISAR", False, "CFOP_NAO_CLASSIFICADO_COMO_VENDA"
    return "CFOP_NAO_E_VENDA", False, "CFOP_NAO_E_VENDA"

def extract_factor(descricao, unidade, config, manual_factor=None):
    if manual_factor is not None:
        return int(manual_factor), "OK", f"Fator manual: {manual_factor}."
    unit = norm_text(unidade)
    raw = f"{clean_str(descricao)} {clean_str(unidade)}"
    text = unicodedata.normalize("NFKD", raw).encode("ascii", "ignore").decode("ascii").upper()
    text = re.sub(r"[\-_]", " ", text)
    fmax = int(config.get("fator_unidade_maximo", 500) or 500)
    patterns = [
        r"\bPCT\s*(?:C\s*/\s*)?(\d{1,4})\s*(?:UN|UND|UNID|UNIDADES)?\b",
        r"\bPACOTE\s*(?:COM|C\s*/)?\s*(\d{1,4})\s*(?:UN|UND|UNID|UNIDADES)?\b",
        r"\bPACK\s*(?:C\s*/\s*)?(\d{1,4})\s*(?:UN|UND|UNID|UNIDADES)?\b",
        r"\bCX\s*(?:C\s*/\s*)?(\d{1,4})\s*(?:UN|UND|UNID|UNIDADES)?\b",
        r"\bCAIXA\s*(?:COM|C\s*/)?\s*(\d{1,4})\s*(?:UN|UND|UNID|UNIDADES)?\b",
        r"\bFARDO\s*(?:COM|C\s*/)?\s*(\d{1,4})\s*(?:UN|UND|UNID|UNIDADES)?\b",
        r"\bFD\s*(?:C\s*/\s*)?(\d{1,4})\s*(?:UN|UND|UNID|UNIDADES)?\b",
        r"\bC\s*/\s*(\d{1,4})\s*(?:UN|UND|UNID|UNIDADES)\b",
    ]
    found = []
    for pat in patterns:
        for m in re.finditer(pat, text):
            n = int(m.group(1))
            if 1 <= n <= fmax:
                found.append(n)
    unique = sorted(set(found))
    has_box = bool(re.search(r"\b(CX|CXS|CAIXA|FARDO|FD|PCT|PACOTE|PACK)\b", text))
    base_units = {norm_text(u) for u in config.get("unidades_base_fator_1", [])}
    if len(unique) == 1:
        return unique[0], "OK", f"FATOR_CAIXA_IDENTIFICADO_{unique[0]}"
    if len(unique) > 1:
        return None, "DIVERGENCIA_NA_QUANTIDADE", "Múltiplos fatores de caixa possíveis."
    if unit in {"KG","G","L","LT","ML","M","MT"}:
        return 1, "OK", "Unidade física sem caixa."
    if has_box and unit in base_units:
        return 1, "DESCRICAO_INDICA_CAIXA_VERIFICAR", "Descrição indica caixa, mas unidade indica unidade; mantida quantidade tributável informada."
    if has_box:
        return None, "SEM_FATOR_CAIXA", "Descrição/unidade indica caixa sem fator seguro."
    return 1, "OK", "Fator unitário adotado."

def _normalize_tipo_item_value(value):
    t = norm_text(value)
    if not t:
        return ""
    # Tokens curtos só valem por igualdade, para evitar que PA seja capturado dentro de palavras.
    if t in {"PA", "PROD ACAB", "PRODUTO ACABADO", "PRODUTO_ACABADO", "PRODUTO ABACADO", "PRODUTO_ABACADO"} or any(x in t for x in ["PRODUTO ACABADO", "PRODUTO_ACABADO", "PRODUTO ABACADO", "PRODUTO_ABACADO", "ACABADO"]):
        return "PRODUTO_ACABADO"
    if t in {"MR", "MERCADORIA REVENDA", "MERCADORIA_REVENDA", "MATERIAL REVENDA", "MATERIAL_REVENDA"} or any(x in t for x in ["MERCADORIA PARA REVENDA", "MERCADORIA_REVENDA", "MATERIAL_REVENDA", "REVENDA"]):
        return "MERCADORIA_REVENDA"
    if t in {"PE", "WIP"} or any(x in t for x in ["ELABORACAO", "ELABORADO", "SEMI ACABADO"]):
        return "PRODUTO_EM_ELABORACAO"
    if t in {"MP"} or any(x in t for x in ["MATERIA PRIMA", "MATERIA_PRIMA", "INSUMO"]):
        return "MATERIA_PRIMA"
    if t in {"EMB"} or any(x in t for x in ["EMBALAGEM", "ROTULO", "TAMPA", "GARRAFA VAZIA", "CAIXA VAZIA"]):
        return "EMBALAGEM"
    if t in {"NA", "ITEM NAO E PRODUTO ACABADO", "ITEM_NAO_E_PRODUTO_ACABADO", "NAO IDENTIFICADO", "NAO_IDENTIFICADO"} or any(x in t for x in ["ITEMNAOEPRODUTOACABADO", "NAOIDENTIFICADO"]):
        return "ITEM_NAO_E_PRODUTO_ACABADO"
    return ""

def _ncm_rir70_set(config):
    return {normalize_ncm(x) for x in (config.get("ncms_arbitramento") or NCM_ABA03_PERMITIDOS) if normalize_ncm(x)}

def _ncm_tipo_item_rir70_set(config):
    return {normalize_ncm(x) for x in (config.get("ncms_tipo_item_rir70") or NCM_TIPO_ITEM_RIR70) if normalize_ncm(x)}

def _descricao_indica_caixa(descricao):
    raw = unicodedata.normalize("NFKD", clean_str(descricao)).encode("ascii", "ignore").decode("ascii").upper()
    raw = re.sub(r"[\-_./]", " ", raw)
    compact = norm_text(descricao)
    return bool(re.search(r"(^|\s)(CX|CXS|CAIXA|FARDO|FD|PCT|PACOTE|PACK)(\s|\d|$)", raw)) or bool(re.search(r"(CX|CXS|FD|PCT|PACK)\d{1,4}", compact))

def _descricao_indica_unidade(descricao, unidade=""):
    u = norm_text(unidade)
    if u in {"UN", "UND", "UNID", "UNIDADE"}:
        return True
    if _descricao_indica_caixa(descricao):
        return False
    raw = unicodedata.normalize("NFKD", clean_str(descricao)).encode("ascii", "ignore").decode("ascii").upper()
    raw = re.sub(r"[\-_./]", " ", raw)
    compact = norm_text(descricao)
    return bool(re.search(r"(^|\s)(UN|UND|UNID|UNIDADE)(\s|$)", raw)) or bool(re.search(r"\d+(ML|L|LT)$", compact)) or bool(re.search(r"\d+(ML|L|LT)(UN|UND|UNID)?$", compact))

def _classificar_tipo_item_regra_operacional(codigo, ncm, descricao, unidade, config):
    cod = clean_str(codigo)
    n = normalize_ncm(ncm)
    ncm_ok = n in _ncm_tipo_item_rir70_set(config)
    if not ncm_ok:
        return "ITEM_NAO_E_PRODUTO_ACABADO", "REGRA_NCM_FORA_ESCOPO"
    if cod.upper().endswith("-F") and _descricao_indica_unidade(descricao, unidade):
        return "PRODUTO_ACABADO", "REGRA_CODIGO_F_NCM_DESCRICAO_UN"
    if not cod.upper().endswith("-F") and _descricao_indica_caixa(descricao):
        return "MERCADORIA_REVENDA", "REGRA_CODIGO_BASE_NCM_DESCRICAO_CX"
    return "NAO_IDENTIFICADO", "REGRA_OPERACIONAL_NAO_CONCLUSIVA"

def _matriz_sheets_validas(config):
    raw = config.get("matriz_sheets_validas")
    if raw:
        return {norm_text(x) for x in raw if clean_str(x)}
    # Sheets operacionais; demais abas administrativas são ignoradas para evitar overhead e leitura indevida.
    return {"MATRIZCADASTRO", "MATRIZSIMILARIDADE", "MAPASIMILARIDADE"}

def _deve_ler_sheet_matriz(sheet_name, config):
    sn = norm_text(sheet_name)
    if not sn or sn.startswith("LOG") or sn.startswith("_") or sn.startswith("RESUMO") or sn.startswith("REGRAS") or sn.startswith("INSTRUCOES"):
        return False
    validas = _matriz_sheets_validas(config)
    return sn in validas

def _max_rows_matriz(config):
    raw = config.get("max_rows_matriz", None)
    if raw in (None, "", 0, "0", "None", "NONE"):
        return None
    try:
        val = int(raw)
        return val if val > 0 else None
    except Exception:
        return None

def _paths_signature(paths):
    sig = []
    for path in paths:
        try:
            if path.exists():
                st = path.stat()
                sig.append((str(path.resolve()), st.st_mtime_ns, st.st_size))
        except Exception:
            sig.append((str(path), 0, 0))
    return tuple(sig)


def _matriz_produto_candidates(config):
    """Arquivos aceitos como fonte de verdade cadastral do produto."""
    aux_dir = BASE_DIR / clean_str(config.get("pasta_auxiliares", "input/auxiliares"))
    return [
        aux_dir / clean_str(config.get("arquivo_matriz_cadastro_produto", "MATRIZ_CADASTRO_PRODUTO_RIR70_SIMILARIDADE.xlsx")),
        aux_dir / "MATRIZ_CADASTRO_PRODUTO_RIR70_SIMILARIDADE.xlsx",
        aux_dir / "MAPA_SIMILARIDADE_PRODUTO.xlsx",
        aux_dir / "MATRIZ_CADASTRO_PRODUTO_RIR70_SIMILARIDADE_FINAL_PROCESSO.xlsx",
        aux_dir / "mapa_similaridade_produto.xlsx",
        BASE_DIR / "MATRIZ_CADASTRO_PRODUTO_RIR70_SIMILARIDADE.xlsx",
        BASE_DIR / "MAPA_SIMILARIDADE_PRODUTO.xlsx",
    ]


def _matriz_produto_meta_signature(meta):
    return tuple(clean_str(meta.get(k)) for k in [
        "CNPJ_RAIZ", "CODIGO", "TIPO_PRODUTO", "PART_ARBITRAMENTO", "PARTICIPA_ARBITR",
        "STATUS_CADASTRO", "VINCULO_GRUPO", "CATEGORIA_ITEM", "NCM", "CEST", "FATOR_UNIDADE",
        "GRUPO_SIMILARIDADE_HOMOLOGADO", "STATUS_HOMOLOGACAO", "PERMITE_FALLBACK_SIMILAR",
        "TIPO_ITEM", "FONTE_TIPO_ITEM", "CRITERIO_CUSTEIO_RIR70",
        "FUNDAMENTO_CRITERIO_CUSTEIO", "ORIGEM_PRODUTO", "CNPJ_FABRICANTE_GRUPO",
        "PRODUTO_EQUIVALENTE_GRUPO"
    ])


def _matriz_produto_cache(config):
    """Lookup cadastral amplo por produto.

    A matriz é a fonte primária para TIPO_PRODUTO, PART_ARBITRAMENTO/PARTICIPA_ARBITR, STATUS_CADASTRO,
    NCM e FATOR_UNIDADE. O lookup prioriza (CNPJ_RAIZ, CODIGO) e só usa chaves
    globais quando não há conflito entre raízes/códigos.
    """
    candidates = _matriz_produto_candidates(config)
    # A assinatura agora inclui um hash da configuração para invalidar o cache
    # sempre que o conteúdo do arquivo de configuração mudar. Isso evita que
    # dados antigos permaneçam no cache após alterações de parâmetros.
    #
    # Anteriormente usávamos ``json.dumps(config, sort_keys=True)`` diretamente
    # para calcular o hash. Contudo, isso pode falhar quando o ``config`` contém
    # tipos não serializáveis pelo JSON padrão, como ``Decimal``, ``set`` ou
    # ``frozenset``. Um fallback com ``hash(frozenset(config.items()))`` também
    # não é confiável, pois listas (e outras estruturas mutáveis) não são
    # hasheáveis. Para garantir robustez, usamos ``safe_json_dumps`` - que
    # converte tipos especiais para representações serializáveis - antes de
    # calcular o SHA‑256. Em caso de falha inesperada, recorremos a uma hash
    # simples baseada no ``repr`` do objeto, que sempre retorna uma string.
    try:
        cfg_json = safe_json_dumps(config, sort_keys=True)
        cfg_hash = hashlib.sha256(cfg_json.encode('utf-8')).hexdigest()
    except Exception:
        try:
            cfg_hash = hashlib.sha256(repr(config).encode('utf-8')).hexdigest()
        except Exception:
            # Último recurso: usar id(config) para diferenciar objetos
            cfg_hash = str(id(config))
    signature = (cfg_hash, _paths_signature(candidates), tuple(sorted(_matriz_sheets_validas(config))), _max_rows_matriz(config))
    if hasattr(_matriz_produto_cache, "_cache") and getattr(_matriz_produto_cache, "_signature", None) == signature:
        return _matriz_produto_cache._cache

    cache = {}
    buckets = defaultdict(list)
    max_rows = _max_rows_matriz(config)
    for path in candidates:
        if not path.exists():
            continue
        try:
            with pd.ExcelFile(path) as xl:
             for sh in xl.sheet_names:
                if not _deve_ler_sheet_matriz(sh, config):
                    continue
                try:
                    df = pd.read_excel(path, sheet_name=sh, dtype=object, nrows=max_rows, keep_default_na=False)
                    if max_rows and len(df) >= max_rows:
                        logger.warning("MATRIZ cadastro possivelmente truncada em %d linhas: %s [%s]", max_rows, path.name, sh)
                except Exception:
                    continue
                cols = {norm_text(c): c for c in df.columns}
                cod_col = cols.get("CODIGO") or cols.get("CODIGOPRODUTO") or cols.get("CODIGOITEM") or cols.get("SKU") or cols.get("CODIGOSEXEMPLO")
                base_col = cols.get("CODIGOBASE")
                cnpj_col = cols.get("CNPJRAIZ") or cols.get("CNPJBASE") or cols.get("RAIZCNPJ")
                if not cod_col:
                    continue

                def _row(row, name):
                    col = cols.get(norm_text(name))
                    return clean_str(row.get(col, "")) if col else ""

                for _, row in df.iterrows():
                    raw_cod = clean_str(row.get(cod_col, ""))
                    if not raw_cod:
                        continue
                    # Em MAPA_SIMILARIDADE, CODIGOS_EXEMPLO pode conter lista. Cada código recebe o mesmo meta de grupo.
                    cods = re.split(r"[,;|]", raw_cod) if cod_col == cols.get("CODIGOSEXEMPLO") else [raw_cod]
                    cnpj_raiz = only_digits(row.get(cnpj_col, ""))[:8] if cnpj_col else ""
                    tipo_item, tipo_col_usada = _pick_tipo_item_from_row(row, _tipo_item_source_cols(cols)) if _tipo_item_source_cols(cols) else ("", "")
                    tipo_produto = _row(row, "TIPO_PRODUTO").upper()
                    if tipo_produto not in {clean_str(x).upper() for x in config.get("tipo_produto_validos", ["PA", "MR"])}:
                        tipo_produto = ""
                    if not tipo_item and tipo_produto:
                        tipo_item = "PRODUTO_ACABADO" if tipo_produto == "PA" else "MERCADORIA_REVENDA" if tipo_produto == "MR" else ""
                        tipo_col_usada = "TIPO_PRODUTO"
                    fonte_tipo = _row(row, "FONTE_TIPO_ITEM") or (f"MATRIZ_AUXILIAR:{path.name}:{sh}:{tipo_col_usada}" if tipo_col_usada else f"MATRIZ_AUXILIAR:{path.name}:{sh}")
                    part_oficial = _row(row, "PART_ARBITRAMENTO")
                    part_legado = _row(row, "PARTICIPA_ARBITR")
                    part_usado = part_oficial if clean_str(part_oficial) else part_legado
                    part_status = _yes_no_status(part_usado, "")
                    legado_status = _yes_no_status(part_legado, "")
                    oficial_status = _yes_no_status(part_oficial, "")
                    conflito_part = bool(oficial_status and legado_status and oficial_status != legado_status)
                    meta_base = {
                        "CNPJ_RAIZ": cnpj_raiz,
                        "CODIGO_BASE": _row(row, "CODIGO_BASE"),
                        "DESCRICAO_TECNICA": _row(row, "DESCRICAO_TECNICA") or _row(row, "DESCRICAO_PRODUTO") or _row(row, "DESCRICAO"),
                        "TIPO_PRODUTO": tipo_produto,
                        "TIPO_PRODUTO_DISPLAY": _display_tipo_produto(tipo_produto),
                        "TIPO_ITEM": tipo_item,
                        "FONTE_TIPO_ITEM": fonte_tipo,
                        "PART_ARBITRAMENTO": part_usado.upper(),
                        "PART_ARBITRAMENTO_STATUS": part_status,
                        "PART_ARBITRAMENTO_LEGADO": part_legado.upper(),
                        "PARTICIPA_ARBITR": part_legado.upper(),
                        "PART_ARBITRAMENTO_CONFLITO": "Sim" if conflito_part else "Nao",
                        "CALCULA_MATRIZ": part_status,
                        "STATUS_CADASTRO": _row(row, "STATUS_CADASTRO").upper(),
                        "VINCULO_GRUPO": _yes_no_status(_row(row, "VINCULO_GRUPO"), _row(row, "VINCULO_GRUPO")),
                        "CATEGORIA_ITEM": _row(row, "CATEGORIA_ITEM"),
                        "GRUPO_SIMILARIDADE_HOMOLOGADO": _row(row, "GRUPO_SIMILARIDADE_HOMOLOGADO"),
                        "STATUS_HOMOLOGACAO": _row(row, "STATUS_HOMOLOGACAO").upper(),
                        "PERMITE_FALLBACK_SIMILAR": _row(row, "PERMITE_FALLBACK_SIMILAR").upper(),
                        "NCM": normalize_ncm(_row(row, "NCM")),
                        "CEST": normalize_cest(_row(row, "CEST")),
                        "FATOR_UNIDADE": _row(row, "FATOR_UNIDADE"),
                        "UNIDADE_MEDIDA": _row(row, "UNIDADE_MEDIDA"),
                        "CRITERIO_CUSTEIO_RIR70": _row(row, "CRITERIO_CUSTEIO_RIR70"),
                        "FUNDAMENTO_CRITERIO_CUSTEIO": _row(row, "FUNDAMENTO_CRITERIO_CUSTEIO"),
                        "ORIGEM_PRODUTO": _row(row, "ORIGEM_PRODUTO"),
                        "CNPJ_FABRICANTE_GRUPO": normalize_cnpj(_row(row, "CNPJ_FABRICANTE_GRUPO")),
                        "PRODUTO_EQUIVALENTE_GRUPO": _row(row, "PRODUTO_EQUIVALENTE_GRUPO"),
                        "FONTE_CADASTRO": f"{path.name}:{sh}",
                    }
                    for cod in cods:
                        exact = codigo_produto_key(cod)
                        normk = normalize_item_join_key(cod)
                        base = normalize_item_join_key(row.get(base_col)) if base_col else normalize_item_join_key(meta_base.get("CODIGO_BASE"))
                        if not exact and not normk:
                            continue
                        meta = dict(meta_base)
                        meta["CODIGO"] = clean_str(cod)
                        if cnpj_raiz and exact:
                            cache.setdefault(f"ROOT_EXACT|{cnpj_raiz}|{exact}", meta)
                        if cnpj_raiz and normk:
                            cache.setdefault(f"ROOT_NORM|{cnpj_raiz}|{normk}", meta)
                        if cnpj_raiz and base:
                            cache.setdefault(f"ROOT_BASE|{cnpj_raiz}|{base}", meta)
                        if exact:
                            buckets[f"EXACT|{exact}"].append(meta)
                        if normk:
                            buckets[f"NORM|{normk}"].append(meta)
                        if base:
                            buckets[f"BASE|{base}"].append(meta)
            try:
                xl.close()
            except Exception:
                pass
        except Exception:
            continue

    conflicts = set()
    for key, metas in buckets.items():
        distinct = {}
        for meta in metas:
            distinct[_matriz_produto_meta_signature(meta)] = meta
        if len(distinct) == 1:
            cache.setdefault(key, next(iter(distinct.values())))
        else:
            conflicts.add(key)
    cache["__CONFLICTS__"] = conflicts
    _matriz_produto_cache._cache = cache
    _matriz_produto_cache._signature = signature
    return cache


def _matriz_produto_meta(codigo, cnpj, config):
    cache = _matriz_produto_cache(config)
    exact = codigo_produto_key(codigo)
    normk = normalize_item_join_key(codigo)
    root = _root_cnpj(cnpj)
    requested_has_suffix = bool(re.search(r"[-_A-Z]", clean_str(codigo).upper()))
    keys = []
    if root and exact:
        keys.append(f"ROOT_EXACT|{root}|{exact}")
    if root and normk and not requested_has_suffix:
        keys.append(f"ROOT_NORM|{root}|{normk}")
    if root and normk:
        keys.append(f"ROOT_BASE|{root}|{normk}")
    if exact:
        keys.append(f"EXACT|{exact}")
    if normk and not requested_has_suffix:
        keys.append(f"NORM|{normk}")
    if normk:
        keys.append(f"BASE|{normk}")
    conflicts = cache.get("__CONFLICTS__", set())
    for key in keys:
        if key in conflicts:
            continue
        meta = cache.get(key)
        if meta:
            return meta
    return None


def _matriz_decimal(value):
    d = to_decimal(value)
    return d if isinstance(d, Decimal) and d > ZERO else None


def _join_adv(*parts):
    out = []
    for p in parts:
        for piece in re.split(r";", clean_str(p)):
            piece = clean_str(piece)
            if piece and piece not in out:
                out.append(piece)
    return ";".join(out)


def _yes_no_status(value, invalid_as=""):
    """Normaliza S/N, Sim/Nao, booleanos e equivalentes para Sim/Nao/Revisar."""
    t = norm_text(value)
    if t in {"S", "SIM", "TRUE", "1", "OK", "APROVADO", "ATIVO"}:
        return "Sim"
    if t in {"N", "NAO", "NO", "FALSE", "0", "NAOAPLICA", "NÃO"}:
        return "Nao"
    if t in {"R", "REV", "REVISAR", "PENDENTE"}:
        return "Revisar"
    return invalid_as

def _is_sim(value):
    return _yes_no_status(value) == "Sim"

def _is_nao(value):
    return _yes_no_status(value) == "Nao"

def _display_tipo_produto(value):
    t = norm_text(value)
    if t in {"PA", "PRODUTOACABADO"}:
        return "Produto Acabado"
    if t in {"MR", "MERCADORIAREVENDA"}:
        return "Mercadoria Revenda"
    return "Nao Identificado"

def _display_part_arbitramento(meta):
    return _yes_no_status((meta or {}).get("PART_ARBITRAMENTO"), "") or ""

def _categoria_norm(value):
    return norm_text(value)

def _perfil_empresa(cnpj, config):
    c = normalize_cnpj(cnpj)
    perfis = config.get("perfis_empresas") or {}
    if isinstance(perfis, dict):
        for k, v in perfis.items():
            if normalize_cnpj(k) == c:
                return clean_str(v)
    mapa = {
        "03408722000178": "Industria Produtora",
        "03408722000410": "Distribuidora",
        "03408722000763": "Industria Distribuidora",
        "03408722000330": "Deposito",
    }
    return mapa.get(c, "")

def _matrix_produto_columns_status(config):
    """Valida presença estrutural dos campos produtivos mínimos na matriz real."""
    required_any = [{"PARTARBITRAMENTO", "PARTICIPAARBITR"}, {"VINCULOGRUPO"}, {"CATEGORIAITEM"}]
    candidates = _matriz_produto_candidates(config)
    found_files = []
    checked_sheets = []
    missing_global = []
    for path in candidates:
        if not path.exists():
            continue
        found_files.append(path.name)
        try:
            with pd.ExcelFile(path) as xl:
             for sh in xl.sheet_names:
                if not _deve_ler_sheet_matriz(sh, config):
                    continue
                try:
                    cols_df = pd.read_excel(path, sheet_name=sh, nrows=0)
                except Exception:
                    continue
                cols = {norm_text(c) for c in cols_df.columns}
                checked_sheets.append(f"{path.name}:{sh}")
                missing = []
                for group in required_any:
                    if not (cols & group):
                        missing.append("PART_ARBITRAMENTO ou PARTICIPA_ARBITR" if "PARTARBITRAMENTO" in group else next(iter(group)))
                if not missing:
                    return {"ok": True, "found_files": found_files, "checked_sheets": checked_sheets, "missing": []}
                missing_global.extend(missing)
        except Exception:
            continue
    return {"ok": False, "found_files": found_files, "checked_sheets": checked_sheets, "missing": sorted(set(missing_global or ["MATRIZ_CADASTRO_PRODUTO_RIR70_SIMILARIDADE.xlsx não localizada"]))}

def _validate_matrix_for_mode(config):
    status = _matrix_produto_columns_status(config)
    config["_matriz_campos_obrigatorios_status"] = status
    if clean_str(config.get("MODO_EXECUCAO")).upper() == "PRODUCAO" and not status.get("ok"):
        raise ConfigError(
            "Campo obrigatório ausente na matriz para execução produtiva: "
            + ", ".join(status.get("missing") or [])
            + ". Campos mínimos: PART_ARBITRAMENTO ou PARTICIPA_ARBITR, VINCULO_GRUPO e CATEGORIA_ITEM."
        )
    return status


def _criterio_custeio_rir70(meta, config=None):
    """Retorna (criterio, fundamento, advertencia).

    Regra conservadora v1.3.6:
    - PA ou Produto Acabado: RIR70_70_MAIOR_PRECO;
    - MR/Mercadoria Revenda com VINCULO_GRUPO=Sim: RIR70_70_MAIOR_PRECO por equivalência de produto acabado do grupo;
    - MR sem vínculo claro: BLOQUEAR_REVISAR;
    - CUSTO_AQUISICAO informado explicitamente: bloqueia/revisa enquanto não existir rotina de custo de aquisição.
    """
    meta = meta or {}
    raw = norm_text(meta.get("CRITERIO_CUSTEIO_RIR70"))
    tipo = norm_text(meta.get("TIPO_PRODUTO"))
    vinculo = _yes_no_status(meta.get("VINCULO_GRUPO"), "")
    fundamento_raw = clean_str(meta.get("FUNDAMENTO_CRITERIO_CUSTEIO"))

    aliases = {
        "RIR7070MAIORPRECO": "RIR70_70_MAIOR_PRECO",
        "RIR70MAIORPRECO": "RIR70_70_MAIOR_PRECO",
        "70MAIORPRECO": "RIR70_70_MAIOR_PRECO",
        "CUSTOAQUISICAO": "CUSTO_AQUISICAO",
        "AQUISICAO": "CUSTO_AQUISICAO",
        "BLOQUEARREVISAR": "BLOQUEAR_REVISAR",
        "REVISAR": "BLOQUEAR_REVISAR",
        "NAOAPLICAVEL": "NAO_APLICAVEL",
        "NA": "NAO_APLICAVEL",
    }
    if raw:
        criterio = aliases.get(raw, clean_str(meta.get("CRITERIO_CUSTEIO_RIR70")).upper())
        fundamento = fundamento_raw or "Critério informado na matriz cadastral."
    elif tipo in {"PA", "PRODUTOACABADO"}:
        criterio = "RIR70_70_MAIOR_PRECO"
        fundamento = "Critério inferido: PA/Produto Acabado sem critério explícito na matriz."
    elif tipo in {"MR", "MERCADORIAREVENDA"} and vinculo == "Sim":
        criterio = "RIR70_70_MAIOR_PRECO"
        fundamento = "Critério inferido: MR vinculado ao grupo tratado como produto acabado equivalente do grupo."
    elif tipo in {"MR", "MERCADORIAREVENDA"}:
        criterio = "BLOQUEAR_REVISAR"
        fundamento = "MR sem VINCULO_GRUPO=Sim; não aplicar automaticamente 70% do maior preço."
    else:
        criterio = "BLOQUEAR_REVISAR"
        fundamento = "Tipo de produto sem critério de custeio seguro."

    adv = ""
    if not raw:
        adv = "CRITERIO_CUSTEIO_RIR70_INFERIDO"
    if criterio == "CUSTO_AQUISICAO":
        adv = "CRITERIO_CUSTEIO_CUSTO_AQUISICAO_REQUER_ROTINA_PROPRIA"
    if criterio == "BLOQUEAR_REVISAR":
        adv = "CRITERIO_CUSTEIO_RIR70_REQUER_REVISAO"
    return criterio, fundamento, adv


def _validar_meta_matriz_para_calculo(meta, config):
    """Retorna (bloqueia_calculo, motivo, advertencia) para elegibilidade cadastral."""
    if not meta:
        adv = "SEM_CADASTRO_MATRIZ"
        return bool(config.get("bloquear_sem_cadastro_matriz", False)), adv, adv

    status = norm_text(meta.get("STATUS_CADASTRO"))
    tipo_produto = norm_text(meta.get("TIPO_PRODUTO"))
    part_status = _yes_no_status(meta.get("PART_ARBITRAMENTO"), "")
    vinculo = _yes_no_status(meta.get("VINCULO_GRUPO"), "")
    ncm = normalize_ncm(meta.get("NCM"))
    criterio_custeio, fundamento_custeio, adv_custeio = _criterio_custeio_rir70(meta, config)

    if config.get("validar_criterio_custeio_rir70", True):
        if criterio_custeio == "BLOQUEAR_REVISAR":
            return True, "CRITERIO_CUSTEIO_RIR70_REQUER_REVISAO", adv_custeio
        if criterio_custeio == "CUSTO_AQUISICAO" and config.get("bloquear_custo_aquisicao_sem_rotina", True):
            return True, "CUSTO_AQUISICAO_REQUER_ROTINA_PROPRIA", adv_custeio

    # Adicionalmente, torna CATEGORIA_ITEM um campo obrigatório.  Ausências são
    # tratadas como pendências para revisão ou bloqueio, pois impedem a
    # correta aplicação das regras de CFOP e de fallback de inventário.  A
    # configuração `validar_categoria_item` pode ser usada para desativar
    # esta verificação em cenários de testes ou migração.
    if config.get("validar_categoria_item", True):
        cat = clean_str(meta.get("CATEGORIA_ITEM")) if meta else ""
        if not cat:
            return True, "CATEGORIA_ITEM_NAO_INFORMADA", "CATEGORIA_ITEM_NAO_INFORMADA"

    if meta.get("PART_ARBITRAMENTO_CONFLITO") == "Sim":
        return True, "DIVERGENCIA_PART_ARBITRAMENTO", "DIVERGENCIA_PART_ARBITRAMENTO"
    if config.get("validar_status_cadastro_matriz", True) and status and status not in {"ATIVO", "OK", "APROVADO"}:
        return True, "CADASTRO_INATIVO", f"STATUS_CADASTRO={meta.get('STATUS_CADASTRO')}"
    if config.get("validar_part_arbitramento_matriz", config.get("validar_participa_arbitr_matriz", True)):
        if not part_status:
            return True, "PART_ARBITRAMENTO_NAO_INFORMADO", "PART_ARBITRAMENTO_NAO_INFORMADO"
        if part_status == "Nao":
            return True, "PRODUTO_NAO_PARTICIPA_DO_ARBITRAMENTO", "PART_ARBITRAMENTO=Nao"
        if part_status == "Revisar":
            return True, "PART_ARBITRAMENTO_NAO_INFORMADO", "PART_ARBITRAMENTO=Revisar"
    if tipo_produto not in {"PA", "MR", "PRODUTOACABADO", "MERCADORIAREVENDA"}:
        return True, "TIPO_PRODUTO_NAO_IDENTIFICADO", f"TIPO_PRODUTO={meta.get('TIPO_PRODUTO')}"
    if tipo_produto in {"MR", "MERCADORIAREVENDA"} and vinculo != "Sim":
        return True, "VINCULO_GRUPO_NAO_INFORMADO", f"VINCULO_GRUPO={meta.get('VINCULO_GRUPO')}"
    if ncm and ncm not in _ncm_rir70_set(config):
        return True, "NCM_FORA_DO_ESCOPO", f"NCM={ncm}"
    return False, "", ""

def reset_tipo_item_cache():
    """Invalida cache de TIPO_ITEM_RIR70 para testes encadeados/Jupyter."""
    for fn in (_tipo_item_aux_cache, _matriz_produto_cache):
        if hasattr(fn, "_cache"):
            delattr(fn, "_cache")
        if hasattr(fn, "_signature"):
            delattr(fn, "_signature")

def reset_cest_cache():
    """Invalida cache de CEST para testes encadeados/Jupyter."""
    for fn in (_lookup_cest_por_codigo_descricao_ncm,):
        if hasattr(fn, "_cache"):
            delattr(fn, "_cache")
        if hasattr(fn, "_signature"):
            delattr(fn, "_signature")

def reset_matriz_caches():
    """Invalida todos os caches dependentes das matrizes auxiliares."""
    reset_tipo_item_cache()
    reset_cest_cache()
    if hasattr(_lookup_descricao_tecnica, "_cache"):
        delattr(_lookup_descricao_tecnica, "_cache")

def _tipo_item_source_cols(cols):
    # Prioridade por linha: se TIPO_ITEM_RIR70 estiver vazio, a linha ainda pode cair em TIPO_ITEM/TIPO_PRODUTO.
    out = []
    for key in ("TIPOITEMRIR70", "TIPOITEM", "TIPOPRODUTO"):
        col = cols.get(key)
        if col and col not in out:
            out.append(col)
    return out

def _pick_tipo_item_from_row(row, tipo_cols):
    for col in tipo_cols:
        raw = row.get(col)
        tipo = _normalize_tipo_item_value(raw)
        if not tipo and norm_text(raw) in {"NA", "NAOIDENTIFICADO", "ITEMNAOEPRODUTOACABADO"}:
            tipo = "ITEM_NAO_E_PRODUTO_ACABADO"
        if tipo:
            return tipo, col
    return "", ""

def _tipo_item_aux_cache(config):
    """Deriva cache de TIPO_ITEM_RIR70 a partir do _matriz_produto_cache já carregado.

    Elimina a segunda leitura independente dos mesmos arquivos XLSX de matriz.
    A assinatura é herdada do master cache para invalidação coerente.
    """
    master = _matriz_produto_cache(config)
    master_sig = getattr(_matriz_produto_cache, "_signature", None)

    if (hasattr(_tipo_item_aux_cache, "_cache")
            and getattr(_tipo_item_aux_cache, "_signature", None) == master_sig):
        return _tipo_item_aux_cache._cache

    cache = {}
    norm_sources: dict = {}

    for key, meta in master.items():
        if key.startswith("__") or not isinstance(meta, dict):
            continue
        tipo = meta.get("TIPO_ITEM")
        if not tipo:
            continue
        tipo_meta = {
            "TIPO_ITEM":            tipo,
            "ABREV_TIPO_ITEM_RIR70": "PA" if tipo == "PRODUTO_ACABADO" else ("MR" if tipo == "MERCADORIA_REVENDA" else "NA"),
            "FONTE_TIPO_ITEM":      meta.get("FONTE_TIPO_ITEM", ""),
            "REGRA_TIPO_ITEM":      meta.get("FONTE_TIPO_ITEM", ""),
            "STATUS_TIPO_ITEM":     "OK",
            "OBS_TIPO_ITEM":        "",
        }
        cache[key] = tipo_meta

    # Propaga conflitos de chave normalizada do master cache
    conflicts_master = master.get("__CONFLICTS__", set())
    cache["__NORM_CONFLICTS__"] = conflicts_master
    cache["__NORM_SOURCES__"]   = {}

    _tipo_item_aux_cache._cache     = cache
    _tipo_item_aux_cache._signature = master_sig
    return cache

def _tipo_item_from_matriz(codigo, config):
    cache = _tipo_item_aux_cache(config)
    exact = codigo_produto_key(codigo)
    normk = normalize_item_join_key(codigo)
    if exact and ("EXACT|" + exact) in cache:
        return cache["EXACT|" + exact]
    # Não colapsa sufixo operacional (-F, -FT etc.) para código base se não houver correspondência exata.
    requested_has_suffix = bool(re.search(r"[-_A-Z]", clean_str(codigo).upper()))
    conflicts = cache.get("__NORM_CONFLICTS__", set())
    if normk and not requested_has_suffix and normk not in conflicts and ("NORM|" + normk) in cache:
        return cache["NORM|" + normk]
    if normk and normk not in conflicts and ("BASE|" + normk) in cache:
        return cache["BASE|" + normk]
    return None

def _infer_tipo_item(ncm, descricao, config, codigo=None, unidade=None, cnpj=None, return_source=False):
    # 1) Matriz de cadastro por CNPJ_RAIZ + CODIGO: fonte oficial para TIPO_PRODUTO/TIPO_ITEM_RIR70.
    meta_produto = _matriz_produto_meta(codigo, cnpj, config)
    if meta_produto and meta_produto.get("TIPO_ITEM"):
        tipo = meta_produto.get("TIPO_ITEM") or "NAO_IDENTIFICADO"
        fonte = meta_produto.get("FONTE_TIPO_ITEM") or meta_produto.get("FONTE_CADASTRO") or "MATRIZ_CADASTRO_PRODUTO_RIR70_SIMILARIDADE"
        return (tipo, fonte) if return_source else tipo

    # 1.1) Compatibilidade: matriz saneada sem CNPJ_RAIZ ou lookup legado por código.
    # A leitura prioriza TIPO_ITEM_RIR70. Se essa coluna estiver vazia na linha, aceita TIPO_ITEM/TIPO_PRODUTO.
    meta = _tipo_item_from_matriz(codigo, config)
    if meta:
        tipo = meta.get("TIPO_ITEM") or "NAO_IDENTIFICADO"
        fonte = meta.get("FONTE_TIPO_ITEM") or "MATRIZ_CADASTRO_PRODUTO_RIR70_SIMILARIDADE"
        return (tipo, fonte) if return_source else tipo

    # 2) Regra operacional explícita do comando v5/v6 como fallback auditável.
    tipo_regra, fonte_regra = _classificar_tipo_item_regra_operacional(codigo, ncm, descricao, unidade or "", config)
    if tipo_regra in {"PRODUTO_ACABADO", "MERCADORIA_REVENDA", "ITEM_NAO_E_PRODUTO_ACABADO"}:
        return (tipo_regra, fonte_regra) if return_source else tipo_regra

    # 3) configuração explícita por NCM, apenas se o cadastro não resolveu.
    n = normalize_ncm(ncm)
    explicit_map = config.get("tipo_item_por_ncm", {}) or {}
    normalized_map = {normalize_ncm(k): _normalize_tipo_item_value(v) or clean_str(v) for k, v in explicit_map.items()}
    if n and n in normalized_map:
        val = normalized_map.get(n) or "NAO_IDENTIFICADO"
        return (val, "CONFIG_TIPO_ITEM_POR_NCM") if return_source else val

    # 4) descrição com termo inequívoco.
    desc_tipo = _normalize_tipo_item_value(descricao)
    if desc_tipo:
        return (desc_tipo, "INFERENCIA_DESCRICAO") if return_source else desc_tipo

    return ("NAO_IDENTIFICADO", "NAO_IDENTIFICADO") if return_source else "NAO_IDENTIFICADO"

def _tipo_item_calcula(tipo):
    return clean_str(tipo) in {"PRODUTO_ACABADO", "MERCADORIA_REVENDA"}

def _root_cnpj(cnpj):
    c = only_digits(cnpj)
    return c[:8] if len(c) >= 8 else ""

def _is_intercompany_cnpj(cnpj_emit, cnpj_dest, config):
    emit = normalize_cnpj(cnpj_emit)
    dest = normalize_cnpj(cnpj_dest)
    grupo = set(_empresas_grupo_cnpjs(config))
    if emit and dest and emit in grupo and dest in grupo and emit != dest:
        return True
    # Compatibilidade controlada: raiz só é usada se o config permitir expressamente.
    if config.get("permitir_intercompany_por_raiz_cnpj", False):
        return bool(_root_cnpj(emit) and _root_cnpj(emit) == _root_cnpj(dest))
    return False

def _is_same_group_root_unlisted(cnpj_emit, cnpj_dest, config):
    emit = normalize_cnpj(cnpj_emit)
    dest = normalize_cnpj(cnpj_dest)
    if not emit or not dest or emit == dest:
        return False
    grupo = set(_empresas_grupo_cnpjs(config))
    roots_grupo = {_root_cnpj(c) for c in grupo if _root_cnpj(c)}
    same_root = bool(_root_cnpj(emit) and _root_cnpj(emit) == _root_cnpj(dest))
    return same_root and _root_cnpj(emit) in roots_grupo and not (emit in grupo and dest in grupo)


def _motivo_cfop_especial(cfop, tipo_item, cnpj_emit, cnpj_dest, config, categoria_item=""):
    code = normalize_cfop(cfop)
    cat = _categoria_norm(categoria_item)
    if code in {"5151", "6151"}:
        if _is_intercompany_cnpj(cnpj_emit, cnpj_dest, config):
            return "TRANSFERENCIA_INTERCOMPANY"
        if config.get("alertar_cnpj_raiz_grupo_nao_cadastrado", True) and _is_same_group_root_unlisted(cnpj_emit, cnpj_dest, config):
            return "TRANSFERENCIA_MESMA_RAIZ_CNPJ_NAO_CADASTRADA"
        return "CFOP_TRANSFERENCIA_DESTINO_EXTERNO"
    if code in {"5910", "6910"}:
        return "BONIFICACAO"
    if code in {"5921", "6921"}:
        if not cat:
            return "CATEGORIA_ITEM_NAO_INFORMADA"
        if cat in {"VASILHAME", "SACARIA", "EMBALAGEM"}:
            return "DEVOLUCAO_VASILHAME_SACARIA"
    if code in {"5901", "6901"}:
        if not cat:
            return "CATEGORIA_ITEM_NAO_INFORMADA"
        if cat in {"MATERIAPRIMA", "EMBALAGEM", "REMESSA", "INDUSTRIALIZACAO"}:
            return "REMESSA_INDUSTRIALIZACAO"
    if code in {"5124", "6124"}:
        if not cat:
            return "CATEGORIA_ITEM_NAO_INFORMADA"
        if cat == "INDUSTRIALIZACAO":
            return "INDUSTRIALIZACAO_MO"
    if code in {"5915", "6915"}:
        if not cat:
            return "CATEGORIA_ITEM_NAO_INFORMADA"
        if cat == "CONSERTO":
            return "REMESSA_CONSERTO"
    if code in {"5556", "6556"}:
        if not cat:
            return "CATEGORIA_ITEM_NAO_INFORMADA"
        if cat == "USOCONSUMO":
            return "DEVOLUCAO_USO_CONSUMO"
    if code in {"5949", "6949", "7949"}:
        return "OUTRAS_SAIDAS"
    return ""

def build_record(row, mapping, config, manual_factors, ipi_column_present, row_origin=""):
    r = _legacy_build_record_rir70(row, mapping, config, manual_factors, ipi_column_present, row_origin)
    qtd_com = r.get("QTDE COM") if isinstance(r.get("QTDE COM"), Decimal) else to_decimal(r.get("QTDE COM"))
    qtd_trib = r.get("QTDE TRIB") if isinstance(r.get("QTDE TRIB"), Decimal) else to_decimal(r.get("QTDE TRIB"))
    qtd_trib_inf = r.get("QTDE TRIB Informada") if isinstance(r.get("QTDE TRIB Informada"), Decimal) else to_decimal(r.get("QTDE TRIB Informada"))
    fator = r.get("Fator Unidade")
    desc = clean_str(r.get("Descrição"))
    unid = clean_str(r.get("Unid. Comercial"))
    cest_idx = mapping.get("cest") if isinstance(mapping, dict) else None
    if cest_idx is not None and cest_idx < len(row):
        cest_lido = normalize_cest(row[cest_idx])
        if cest_lido:
            r["CEST"] = cest_lido
            r["FONTE_CEST"] = "INPUT_DIRETO"
    if not normalize_cest(r.get("CEST")):
        cest_lookup, fonte_cest = _lookup_cest_por_codigo_descricao_ncm(r.get("Código Item"), desc, r.get("NCM"), config)
        if cest_lookup:
            r["CEST"] = cest_lookup
            r["FONTE_CEST"] = fonte_cest
    matriz_meta = _matriz_produto_meta(r.get("Código Item"), r.get("CNPJ Emitente"), config)
    tipo_item, fonte_tipo_item = _infer_tipo_item(r.get("NCM"), desc, config, r.get("Código Item"), unid, r.get("CNPJ Emitente"), return_source=True)
    matriz_bloqueia, matriz_motivo, matriz_adv = _validar_meta_matriz_para_calculo(matriz_meta, config)
    categoria_item = clean_str((matriz_meta or {}).get("CATEGORIA_ITEM"))
    perfil_empresa = _perfil_empresa(r.get("CNPJ Emitente"), config)
    situacao_qtd = "OK"
    qtd_calc = qtd_trib
    fator_dec = Decimal(str(fator)) if fator not in (None, "", 0) else None
    has_box = bool(re.search(r"\b(CX|CXS|CAIXA|FARDO|FD|PCT|PACOTE|PACK)\b", norm_text(f"{desc} {unid}")))
    if qtd_com is None or qtd_com <= ZERO:
        situacao_qtd = "QUANTIDADE_INVALIDA"; qtd_calc = None
    elif fator_dec is None and has_box:
        situacao_qtd = "SEM_FATOR_CAIXA"; qtd_calc = None
    elif fator_dec and fator_dec > 1:
        esperado = q6(qtd_com * fator_dec)
        if qtd_trib_inf is None or qtd_trib_inf == qtd_com:
            qtd_calc = esperado; situacao_qtd = "CAIXA_CONVERTIDA"
        elif abs(qtd_trib_inf - esperado) <= Decimal("0.000100"):
            qtd_calc = qtd_trib_inf; situacao_qtd = "OK"
        else:
            qtd_calc = None; situacao_qtd = "DIVERGENCIA_NA_QUANTIDADE"
    elif has_box and fator_dec == Decimal("1"):
        qtd_calc = qtd_trib or qtd_com; situacao_qtd = "DESCRICAO_INDICA_CAIXA_VERIFICAR"
    else:
        qtd_calc = qtd_trib or qtd_com; situacao_qtd = "OK"
    if qtd_calc is not None:
        qtd_calc = q6(qtd_calc)
    if config.get("bloquear_qtd_calculo_zero_ou_negativa", True) and (qtd_calc is None or qtd_calc <= ZERO):
        situacao_qtd = "QUANTIDADE_INVALIDA"
        qtd_calc = None

    valor_total = r.get("_valor_total") if isinstance(r.get("_valor_total"), Decimal) else to_decimal(r.get("Valor Comercial"))
    unit = q6(valor_total / qtd_calc) if isinstance(valor_total, Decimal) and isinstance(qtd_calc, Decimal) and qtd_calc > ZERO else None

    participa = bool(r.get("_participa"))
    motivo = clean_str(r.get("_motivo") or r.get("Motivo Exclusão")) or "VENDA_VALIDA"
    cfop_tipo = clean_str(r.get("Tipo Operação"))
    if situacao_qtd in {"DIVERGENCIA_NA_QUANTIDADE", "SEM_FATOR_CAIXA", "QUANTIDADE_INVALIDA"}:
        participa = False; motivo = situacao_qtd
    cfop_motivo_especial = _motivo_cfop_especial(r.get("CFOP"), tipo_item, r.get("CNPJ Emitente"), r.get("CNPJ Destinatário"), config, categoria_item)
    if cfop_motivo_especial:
        participa = False; motivo = cfop_motivo_especial
    if not _tipo_item_calcula(tipo_item):
        participa = False; motivo = "ITEM_NAO_E_PRODUTO_ACABADO" if tipo_item not in {"NAO_IDENTIFICADO", "ITEM_NAO_E_PRODUTO_ACABADO"} else "TIPO_ITEM_NAO_IDENTIFICADO" if tipo_item == "NAO_IDENTIFICADO" else "ITEM_NAO_E_PRODUTO_ACABADO"
    if normalize_ncm(r.get("NCM")) and normalize_ncm(r.get("NCM")) not in _ncm_rir70_set(config):
        participa = False; motivo = "NCM_FORA_DO_ESCOPO"
    if normalize_item_join_key(r.get("Código Item")) == "650150" and matriz_meta:
        um_matriz = norm_text(matriz_meta.get("UNIDADE_MEDIDA"))
        fator_matriz = _matriz_decimal(matriz_meta.get("FATOR_UNIDADE"))
        if um_matriz in {"UN", "UND", "UNIDADE"} and fator_matriz == Decimal("6") and "CX6" in norm_text(desc):
            participa = False; motivo = "UNIDADE_MEDIDA_INCONSISTENTE"; matriz_adv = _join_adv(matriz_adv, "CODIGO_650150_UNIDADE_MEDIDA_INCONSISTENTE")
    if matriz_bloqueia and participa:
        participa = False
        motivo = matriz_motivo
    if unit is None or unit <= ZERO:
        participa = False
        if motivo == "VENDA_VALIDA": motivo = "VALOR_ZERO"

    r["TIPO_PRODUTO_MATRIZ"] = matriz_meta.get("TIPO_PRODUTO") if matriz_meta else ""
    r["TIPO_PRODUTO"] = _display_tipo_produto(matriz_meta.get("TIPO_PRODUTO") if matriz_meta else "")
    r["PART_ARBITRAMENTO"] = _display_part_arbitramento(matriz_meta)
    r["CALCULA_MATRIZ"] = matriz_meta.get("CALCULA_MATRIZ") if matriz_meta else ""
    r["PARTICIPA_ARBITR_MATRIZ"] = matriz_meta.get("PARTICIPA_ARBITR") if matriz_meta else ""
    r["STATUS_CADASTRO_MATRIZ"] = matriz_meta.get("STATUS_CADASTRO") if matriz_meta else ""
    r["VINCULO_GRUPO"] = (matriz_meta.get("VINCULO_GRUPO") if matriz_meta else "")
    r["CATEGORIA_ITEM"] = categoria_item
    criterio_custeio, fundamento_custeio, adv_custeio = _criterio_custeio_rir70(matriz_meta, config) if matriz_meta else ("BLOQUEAR_REVISAR", "Sem cadastro de matriz para definir critério de custeio.", "CRITERIO_CUSTEIO_RIR70_REQUER_REVISAO")
    r["CRITERIO_CUSTEIO_RIR70"] = criterio_custeio
    r["FUNDAMENTO_CRITERIO_CUSTEIO"] = fundamento_custeio
    r["PERFIL_EMPRESA"] = perfil_empresa
    r["FATOR_UNIDADE_MATRIZ"] = matriz_meta.get("FATOR_UNIDADE") if matriz_meta else ""
    r["NCM_MATRIZ"] = matriz_meta.get("NCM") if matriz_meta else ""
    r["Tipo Item"] = tipo_item
    r["TIPO_ITEM"] = tipo_item
    r["FONTE_TIPO_ITEM"] = fonte_tipo_item
    r["QTD_ORIGINAL"] = q4(qtd_com) if isinstance(qtd_com, Decimal) else None
    r["UNIDADE_ORIGINAL"] = unid
    r["FATOR_CAIXA"] = int(fator_dec) if fator_dec and fator_dec == int(fator_dec) else (fator if fator else "")
    r["QTD_CALCULO"] = qtd_calc
    r["SITUACAO_QTD"] = situacao_qtd
    r["VALOR_UNITARIO_CALCULO"] = unit
    r["Vlr Unitário Base RIR70"] = unit
    r["_participa"] = participa
    r["_motivo"] = "VENDA_VALIDA" if participa else motivo
    r["Participa Cálculo"] = "Sim" if participa else "Nao"
    # REVISAR não entra automaticamente no cálculo. Somente CFOP de revisão, sem outro bloqueio impeditivo, fica como REVISAR.
    revisar_motivos = {"CFOP_NAO_CLASSIFICADO_COMO_VENDA", "CFOP_TRANSFERENCIA_DESTINO_EXTERNO", "DIVERGENCIA_PART_ARBITRAMENTO", "PART_ARBITRAMENTO_NAO_INFORMADO", "VINCULO_GRUPO_NAO_INFORMADO", "CATEGORIA_ITEM_NAO_INFORMADA", "TIPO_PRODUTO_NAO_IDENTIFICADO", "UNIDADE_MEDIDA_INCONSISTENTE"}
    if participa:
        calcula_status = "Sim"
    elif motivo in revisar_motivos or (cfop_tipo == "CFOP_REVISAR" and motivo == "CFOP_NAO_CLASSIFICADO_COMO_VENDA"):
        calcula_status = "Revisar"
    else:
        calcula_status = "Nao"
    r["ENTRA_NO_CALCULO"] = calcula_status
    r["Motivo Exclusão"] = "" if participa else motivo
    r["MOTIVO"] = "VENDA_VALIDA" if participa else motivo
    adv_qtd = "DESCRICAO_INDICA_CAIXA_VERIFICAR" if situacao_qtd == "DESCRICAO_INDICA_CAIXA_VERIFICAR" else ""
    adv_motivo = motivo if not participa and motivo != "VENDA_VALIDA" else ""
    r["ADVERTENCIA"] = _join_adv(matriz_adv, adv_custeio, adv_qtd, adv_motivo)
    cnpj_key = r.get("CNPJ Emitente") if config.get("segregar_calculo_por_cnpj_emitente", True) else ""
    item_key = codigo_produto_key(r.get("Código Item")) or norm_text(r.get("Descrição"))[:80]
    r["_monthly_key"] = (cnpj_key, item_key, normalize_ncm(r.get("NCM")), r.get("Competência"))
    r["_group_key"] = (cnpj_key, item_key, normalize_ncm(r.get("NCM")))
    return r

def duplicate_fingerprint(r):
    chave = normalize_chave(r.get("Chave NF-e"))
    item = clean_str(r.get("_item_nf"))
    if re.fullmatch(r"\d{44}", chave):
        if item:
            return f"{chave}|{item}"
        # Fallback sem nItem: chave composta ampliada para reduzir colisão entre
        # linhas distintas da mesma NF com mesmo código/valor/quantidade.
        desc_norm = norm_text(clean_str(r.get("Descrição") or r.get("Descrição Produto") or ""))[:40]
        cfop = clean_str(r.get("CFOP") or "")
        gtin = clean_str(r.get("GTIN") or r.get("EAN") or "")
        logger.debug("duplicate_fingerprint: deduplicação frágil (sem nItem) chave=%s cod=%s", chave, clean_str(r.get("Código Item")))
        return "|".join([chave, clean_str(r.get("Código Item")), cfop, gtin, desc_norm,
                         str(r.get("_valor_total") or r.get("Valor Comercial") or ""),
                         str(r.get("QTD_CALCULO") or r.get("QTDE TRIB") or "")])
    dt = r.get("Data Emissão")
    return "|".join([dt.strftime("%Y%m%d") if isinstance(dt, datetime) else "",
                     clean_str(r.get("Número NF")), clean_str(r.get("Série")),
                     clean_str(r.get("CNPJ Emitente")), item, clean_str(r.get("Código Item")),
                     str(r.get("_valor_total") or ""), str(r.get("QTD_CALCULO") or "")])

def load_inputs(config):
    global RIR70_DUPLICATES
    files = discover_movement_files(config)
    xml_files = discover_xml_files(config)
    manual_factors = load_manual_factors(config)
    records, logs, duplicates = [], [], []
    if not files and not xml_files:
        logs.append(("SEM_ARQUIVO_MOVIMENTO", "-", "SEM_ARQUIVO_MOVIMENTO", "Importadas=0; fora_periodo=0; duplicadas=0; vazias=0; nenhum arquivo de movimento localizado."))
    document_master, _ = load_document_master(config, logs)
    fingerprints = {}
    periodo_inicio = parse_config_date(config.get("periodo_base_inicio"))
    periodo_fim = parse_config_date(config.get("periodo_base_fim"))
    if periodo_fim: periodo_fim = periodo_fim.replace(hour=23, minute=59, second=59)

    # Prioridade oficial: XML autorizado antes do DFE Excel de itens.
    load_xml_inputs(config, manual_factors, periodo_inicio, periodo_fim, fingerprints, records, duplicates, logs)

    for file_path in files:
        print(f"Lendo: {file_path.name}")
        if config.get("usar_leitor_xlsx_streaming", True) and file_path.suffix.lower() in {".xlsx", ".xlsm"}:
            try:
                from openpyxl import load_workbook as _load_wb_once
                wb_once = _load_wb_once(file_path, read_only=True, data_only=True)
                for sh in list(wb_once.sheetnames):
                    _process_sheet_stream(file_path, sh, config, manual_factors, periodo_inicio, periodo_fim, fingerprints, records, duplicates, logs, wb_open=wb_once)
                try:
                    wb_once.close()
                except Exception:
                    pass
                gc.collect()
            except Exception as e:
                logs.append((file_path.name, "-", "ERRO_LEITURA", str(e)))
        else:
            try:
                with pd.ExcelFile(file_path, engine=excel_engine(file_path)) as xl:
                 for sh in list(xl.sheet_names):
                    try:
                        df = pd.read_excel(file_path, sheet_name=sh, header=None, nrows=HEADER_SCAN, dtype=object, engine=excel_engine(file_path), keep_default_na=False)
                        hi = find_header_row(df, config.get("mapeamento_colunas"))
                        headers = make_unique_headers(df.iloc[hi["idx"]].tolist())
                        tipo = classify_file_headers(headers)
                        if tipo != "ITEM_FISCAL":
                            logs.append((file_path.name, sh, "NAO_UTILIZADO", f"Tipo={tipo}")); continue
                        mapping = {k: pick_column(headers, k, config.get("mapeamento_colunas")) for k in COLUMN_CANDIDATES}
                        score = sum(1 for k in ["chave","data_emissao","ncm","cfop","codigo_item","descricao","unid_comercial","valor_total_produto"] if mapping.get(k) is not None)
                        if score < config.get("score_minimo_colunas_fiscais", 5):
                            logs.append((file_path.name, sh, "NAO_UTILIZADO", f"Score={score}")); continue
                        ipi_col = mapping.get("valor_ipi") is not None
                        df2 = pd.read_excel(file_path, sheet_name=sh, header=hi["idx"], dtype=object, engine=excel_engine(file_path), keep_default_na=False)
                        df2.columns = make_unique_headers(list(df2.columns))
                        before = len(records); out = dup = empty = 0
                        for ridx, row in enumerate(df2.itertuples(index=False, name=None), start=hi["idx"] + 2):
                            s = process_record_row(row, mapping, config, manual_factors, ipi_col, file_path, sh, periodo_inicio, periodo_fim, fingerprints, records, duplicates, ridx)
                            if s == "VAZIA": empty += 1
                            elif s == "FORA_PERIODO": out += 1
                            elif s == "DUPLICATA": dup += 1
                        logs.append((file_path.name, sh, "PROCESSADA", f"Importadas={len(records)-before}; fora_periodo={out}; duplicadas={dup}; vazias={empty}; Score={score}; IPI={'SIM' if ipi_col else 'NAO'}"))
                        del df2; gc.collect()
                    except Exception as e:
                        logs.append((file_path.name, sh, "ERRO_LEITURA", str(e)))
            except Exception as e:
                logs.append((file_path.name, "-", "ERRO_LEITURA", str(e)))

    apply_document_master(records, document_master, logs, config)
    apply_party_directory(records, document_master, logs)
    if not records and not config.get("gerar_relatorio_sem_movimento", True):
        raise ValueError("Nenhuma linha no filtro de datas.")
    RIR70_DUPLICATES = duplicates
    return records, logs, duplicates

def _apply_cest_from_input(records):
    """Preenche CEST ausente usando chave CODIGO+DESCRICAO+NCM localizada nos próprios arquivos de entrada."""
    lookup = {}
    for r in records:
        cest = normalize_cest(r.get("CEST"))
        if not cest:
            continue
        key = (codigo_produto_key(r.get("Código Item")), norm_text(r.get("Descrição")), normalize_ncm(r.get("NCM")))
        if all(key) and key not in lookup:
            lookup[key] = cest
    for r in records:
        if normalize_cest(r.get("CEST")):
            continue
        key = (codigo_produto_key(r.get("Código Item")), norm_text(r.get("Descrição")), normalize_ncm(r.get("NCM")))
        if all(key) and key in lookup:
            r["CEST"] = lookup[key]
            r["FONTE_CEST"] = "INPUT_CODIGO_DESCRICAO_NCM"
        else:
            r["FONTE_CEST"] = "NAO_LOCALIZADO"

def enrich_records(records, config=None):
    cfg = config or DEFAULT_CONFIG
    _apply_cest_from_input(records)
    groups = defaultdict(list)
    for r in records:
        if r.get("_participa"):
            unit = r.get("VALOR_UNITARIO_CALCULO") or r.get("Vlr Unitário Base RIR70")
            qtd = r.get("QTD_CALCULO") or r.get("QTDE TRIB")
            if isinstance(unit, Decimal) and unit > ZERO and isinstance(qtd, Decimal) and qtd > ZERO:
                groups[r["_monthly_key"]].append(r)
    metrics = {}
    for k, rows in groups.items():
        pairs = []
        for r in rows:
            unit = r.get("VALOR_UNITARIO_CALCULO") or r.get("Vlr Unitário Base RIR70")
            qtd = r.get("QTD_CALCULO") or r.get("QTDE TRIB")
            if isinstance(unit, Decimal) and isinstance(qtd, Decimal) and qtd > ZERO:
                pairs.append((unit, qtd, r))
        if not pairs:
            continue
        units = [u for u, _, _ in pairs]
        max_unit = q6(max(units))
        total_qtd = sum((q for _, q, _ in pairs), ZERO)
        total_val = sum((u * q for u, q, _ in pairs), ZERO)
        avg = q6(total_val / total_qtd) if total_qtd > ZERO else None
        base = q6(max_unit * Decimal("0.70")) if isinstance(max_unit, Decimal) else None
        diff_pct = q6(((max_unit - avg) / avg) * Decimal("100")) if isinstance(max_unit, Decimal) and isinstance(avg, Decimal) and avg > ZERO else None
        ref_docs = []
        ref_dates = []
        for r in rows:
            u = r.get("VALOR_UNITARIO_CALCULO") or r.get("Vlr Unitário Base RIR70")
            if isinstance(u, Decimal) and max_unit is not None and q6(u) == max_unit:
                doc = clean_str(r.get("Número NF")) or clean_str(r.get("Chave NF-e"))
                if doc and doc not in ref_docs:
                    ref_docs.append(doc)
                    ref_dates.append(format_date_abnt(r.get("Data Emissão")))
        metrics[k] = {"avg": avg, "max": max_unit, "base": base, "diff": diff_pct, "cnt": len(pairs), "qtd": total_qtd, "docs": ref_docs, "dates": ref_dates}
    for r in records:
        m = metrics.get(r.get("_monthly_key"), {})
        r["PRECO_MEDIO"] = m.get("avg")
        r["Vlr Médio"] = m.get("avg")
        r["MAIOR_PRECO_UNITARIO"] = m.get("max")
        r["Maior Vlr Unitário"] = m.get("max")
        r["CUSTO_ARBITRADO_70"] = m.get("base")
        r["Base Arbitrada (70%)"] = m.get("base")
        r["VARIACAO_PRECO_PCT_INTERNA"] = m.get("diff")
        qtd = r.get("QTD_CALCULO") or r.get("QTDE TRIB")
        base = m.get("base")
        r["CMV Linha"] = q2(base * qtd) if isinstance(base, Decimal) and isinstance(qtd, Decimal) else None
        r["Participa Cálculo"] = "Sim" if r.get("_participa") else "Nao"
        r["Motivo Exclusão"] = "" if r.get("_participa") else clean_str(r.get("_motivo"))
    records.sort(key=lambda r: (r.get("_sort_comp", 999999), r.get("_sort_data", datetime.max), clean_str(r.get("CNPJ Emitente")), clean_str(r.get("Código Item"))))
    return records

# Colunas executivas: output enxuto para o analista. Campos tecnicos ficam ocultos por padrao.
MOV_COLS_ANALISTA = [
    "EMPRESA","PERFIL_EMPRESA","MES","DATA_EMISSAO","DOCUMENTO","CHAVE_NFE","CFOP",
    "CODIGO_PRODUTO","DESCRICAO_PRODUTO","TIPO_PRODUTO","Part_Arbitramento",
    "VINCULO_GRUPO","CATEGORIA_ITEM","CRITERIO_CUSTEIO_RIR70","NCM","CEST","VALOR_PRODUTO","QTD_ORIGINAL",
    "UM_ORIGINAL","FATOR_CAIXA","QTD_CALCULO","VALOR_UNITARIO","SITUACAO_QTD",
    "CALCULA","MOTIVO","ALERTA","ADVERTENCIA",
    "INTERVENCAO_ANALISTA","ACAO_ANALISTA"
]
# Ordem de colunas da aba 02_ARBITRAMENTO.
# A coluna DOCUMENTO_MAIOR_PRECO e ORIGEM_DO_PRECO foram movidas para o final da listagem,
# após JUSTIFICATIVA_ANALISTA, para cumprir a solicitação de posicionamento.  
ARB_COLS_ANALISTA = [
    "ID_ARBITRAMENTO","EMPRESA","PERFIL_EMPRESA","MES","CODIGO_PRODUTO","DESCRICAO_PRODUTO","TIPO_PRODUTO",
    "Part_Arbitramento","VINCULO_GRUPO","CATEGORIA_ITEM","CRITERIO_CUSTEIO_RIR70","NCM","CEST","MAIOR_PRECO_UNITARIO",
    "CUSTO_ARBITRADO_70","PRECO_MEDIO","VARIACAO_PRECO (%)",
    "MOTIVO","ALERTA","ADVERTENCIA",
    "INTERVENCAO_ANALISTA","ACAO_ANALISTA","ARQUIVO_AJUSTE","LINK_ABRIR_AJUSTE",
    "AJUSTE_ANALISTA_APLICADO","ID_AJUSTE","JUSTIFICATIVA_ANALISTA",
    "DOCUMENTO_MAIOR_PRECO","ORIGEM_DO_PRECO"
]
# Ordem de colunas da aba 03_INVENTARIO_VALORIZADO.  
# Ajustada para incluir quantidade vendida e CMV logo após o custo arbitrado,  
# além de deslocar SALDO_ESTOQUE para antes de VALOR_ESTOQUE_ARBITRADO.  
# FONTE_TIPO_ITEM é posicionada ao final após JUSTIFICATIVA_PRECO_REF para melhor leitura.
# Colunas operacionais da aba 03_INVENTARIO_VALORIZADO.
#
# Conforme as diretrizes de versão 1.3.3, removemos a coluna MAIOR_PRECO_REF
# da planilha de inventário valorizado.  Essa informação já é utilizada
# internamente para cálculo do custo arbitrado e sua exibição poderia causar
# interpretação equivocada, visto que a regra fiscal exige apenas o uso de 70%
# do maior preço de venda.  Além disso, a coluna DOCUMENTO_ORIGEM foi
# reposicionada para aparecer logo após JUSTIFICATIVA_PRECO_REF para
# oferecer melhor contexto operacional ao analista.
INV_COLS_ANALISTA = [
    "ID_INVENTARIO","EMPRESA","PERFIL_EMPRESA","MES","ALMOX","CODIGO_PRODUTO",
    "DESCRICAO_TECNICA","NCM","CEST","TIPO_PRODUTO",
    "Part_Arbitramento","VINCULO_GRUPO","CATEGORIA_ITEM","CRITERIO_CUSTEIO_RIR70",
    "UNIDADE","CUSTO_ARBITRADO_70","QUANTIDADE_VENDIDA","CMV",
    "SALDO_ESTOQUE","VALOR_ESTOQUE_ARBITRADO",
    # A coluna MAIOR_PRECO_REF foi removida a partir da versão 1.3.3.
    "ORIGEM_DO_PRECO","EMPRESA_ORIGEM_PRECO","PRODUTO_ORIGEM_PRECO",
    "DESCRICAO_PRODUTO","MES_ORIGEM_PRECO",
    # JUSTIFICATIVA_PRECO_REF e DOCUMENTO_ORIGEM aparecem em sequência.
    "JUSTIFICATIVA_PRECO_REF","DOCUMENTO_ORIGEM",
    "SITUACAO_PRECO","ALERTA","ADVERTENCIA",
    # ARQUIVO_AJUSTE e LINK_ABRIR_AJUSTE removidos da v1.3.5:
    # para todos os casos que geram INTERVENCAO=SIM, a pendência correspondente
    # já existe em 04_PENDENCIAS com informação mais completa (ABA_AJUSTE,
    # CAMPO_AJUSTE, TIPO_AJUSTE_SUGERIDO). O link da CAPA (LINK_ABRIR_AJUSTES)
    # cumpre o papel de acesso rápido sem ambiguidade funcional.
    "INTERVENCAO_ANALISTA","ACAO_ANALISTA",
    "AJUSTE_ANALISTA_APLICADO","ID_AJUSTE","JUSTIFICATIVA_ANALISTA",
    "JUSTIFICATIVA_FALLBACK","FONTE_TIPO_ITEM"
]
# Colunas tecnicas opcionais: somente aparecem quando exibir_colunas_diagnostico = Sim.
MOV_COLS_DIAGNOSTICO = ["TIPO_ITEM","FONTE_TIPO_ITEM","STATUS_CADASTRO_MATRIZ","CALCULA_MATRIZ","FATOR_UNIDADE_MATRIZ","NCM_MATRIZ","FUNDAMENTO_CRITERIO_CUSTEIO"]
ARB_COLS_DIAGNOSTICO = ["TIPO_ITEM","FONTE_TIPO_ITEM","STATUS_CADASTRO_MATRIZ","CALCULA_MATRIZ","FATOR_UNIDADE_MATRIZ","NCM_MATRIZ","FUNDAMENTO_CRITERIO_CUSTEIO"]
INV_COLS_DIAGNOSTICO = ["TIPO_ITEM","FONTE_TIPO_ITEM","STATUS_CADASTRO_MATRIZ","CALCULA_MATRIZ","FATOR_UNIDADE_MATRIZ","NCM_MATRIZ","FUNDAMENTO_CRITERIO_CUSTEIO"]

def _exibir_colunas_diagnostico(config):
    return _is_sim(config.get("exibir_colunas_diagnostico"))

def _cols_operacionais(base_cols, diag_cols, config):
    cols = list(base_cols)
    if _exibir_colunas_diagnostico(config):
        cols += [c for c in diag_cols if c not in cols]
    return cols
PEND_COLS_FINAL = [
    "ID_REFERENCIA","TIPO_REFERENCIA","EMPRESA","MES","CODIGO_PRODUTO","DESCRICAO_PRODUTO",
    "TIPO_PENDENCIA","NIVEL","INTERVENCAO_ANALISTA","ACAO_ANALISTA",
    "ARQUIVO_AJUSTE","ABA_AJUSTE","CAMPO_AJUSTE","TIPO_AJUSTE_SUGERIDO","ALTERA_CALCULO",
    "EVIDENCIA_OBRIGATORIA","RISCO_SE_NAO_CORRIGIR","LINK_ABRIR_AJUSTE",
    "DECISAO_ANALISTA","DATA_DECISAO","RESPONSAVEL_DECISAO","SITUACAO"
]
FONTES_COLS_FINAL = ["ARQUIVO","TIPO","EMPRESA","PERIODO","REGISTROS_LIDOS","REGISTROS_VALIDOS","REGISTROS_DUPLICADOS","REGISTROS_BLOQUEADOS_CFOP","REGISTROS_SEM_CADASTRO","LINHAS_LIDAS","LINHAS_USADAS","LINHAS_DESCARTADAS","LINHAS_DUPLICADAS_DESCARTADAS","MOTIVO_PRINCIPAL_DESCARTE","CODIGO_CONTROLE_ARQUIVO","ENCODING_USADO","SITUACAO"]


def _alerta_visual(advertencia="", situacao="", motivo=""):
    """Marcador visual estável para Excel.
    Evita emoji unicode solto, que costuma renderizar mal em Arial Narrow/Windows.
    A cor real é aplicada na formatação da célula.
    """
    texto = norm_text(f"{advertencia} {situacao} {motivo}")
    if not texto:
        return ""
    def hit(tokens):
        return any(norm_text(t) in texto for t in tokens)
    if hit(["BLOQUEADA", "SEM_PRECO", "QUANTIDADE_INVALIDA", "DIVERGENCIA_NA_QUANTIDADE", "SEM_FATOR_CAIXA", "ITEM_NAO_E_PRODUTO_ACABADO", "TIPO_ITEM_NAO_IDENTIFICADO", "PRECO_MUITO_ACIMA", "MUITO_ACIMA", "PENDENTE"]):
        return "● CRITICO"
    if hit(["PRECO_ACIMA", "OUTRA_EMPRESA", "REVISAR", "CFOP_NAO_CLASSIFICADO"]):
        return "● REVISAR"
    if hit(["POUCAS_VENDAS", "MES_ANTERIOR", "DEFASAGEM", "DESCRICAO_INDICA_CAIXA", "PRODUTO_SIMILAR", "PRECO_MEDIO_ZERO"]):
        return "● ATENCAO"
    if hit(["APROVADA", "OK", "SIM", "VENDA_VALIDA"]):
        return "● OK"
    return "● ALERTA"

def _lookup_descricao_tecnica(codigo, descricao_atual, config):
    """Busca DESCRICAO_TECNICA derivando do _matriz_produto_cache já carregado.

    Elimina a terceira leitura independente dos mesmos arquivos XLSX de matriz.
    A assinatura é herdada do master cache para invalidação coerente.
    """
    master = _matriz_produto_cache(config)
    master_sig = getattr(_matriz_produto_cache, "_signature", None)

    if (not hasattr(_lookup_descricao_tecnica, "_cache")
            or getattr(_lookup_descricao_tecnica, "_signature", None) != master_sig):
        cache: dict = {}
        seen_ids: set = set()
        for key, meta in master.items():
            if key.startswith("__") or not isinstance(meta, dict):
                continue
            mid = id(meta)
            if mid in seen_ids:
                continue
            seen_ids.add(mid)
            desc = clean_str(meta.get("DESCRICAO_TECNICA"))
            if not desc:
                continue
            # Popula cache com chaves globais (sem CNPJ) derivadas do campo CODIGO da meta,
            # para que o lookup por código funcione independentemente de qual CNPJ registrou.
            cod_meta = clean_str(meta.get("CODIGO"))
            if cod_meta:
                exact_k = codigo_produto_key(cod_meta)
                norm_k = normalize_item_join_key(cod_meta)
                if exact_k:
                    cache.setdefault(f"EXACT|{exact_k}", desc)
                if norm_k:
                    cache.setdefault(f"NORM|{norm_k}", desc)
        # Índice direto pelas chaves existentes no master que já carregam a descrição
        for key, meta in master.items():
            if key.startswith("__") or not isinstance(meta, dict):
                continue
            desc = clean_str(meta.get("DESCRICAO_TECNICA"))
            if desc:
                cache.setdefault(key, desc)
        _lookup_descricao_tecnica._cache     = cache
        _lookup_descricao_tecnica._signature = master_sig

    cache = _lookup_descricao_tecnica._cache
    exact = codigo_produto_key(codigo)
    normk = normalize_item_join_key(codigo)
    for key in [
        f"EXACT|{exact}" if exact else "",
        f"NORM|{normk}"  if normk else "",
    ]:
        if key and key in cache:
            return cache[key]
    return clean_str(descricao_atual)

def _lookup_cest_por_codigo_descricao_ncm(codigo, descricao, ncm, config):
    """Busca CEST por (CODIGO, DESCRICAO, NCM) derivando do _matriz_produto_cache já carregado.

    Elimina a quarta leitura independente dos mesmos arquivos XLSX de matriz.
    Cache construído a partir das metas já presentes no master cache,
    sem qualquer nova leitura de arquivo. Assinatura herdada do master.
    """
    master = _matriz_produto_cache(config)
    master_sig = getattr(_matriz_produto_cache, "_signature", None)

    if (not hasattr(_lookup_cest_por_codigo_descricao_ncm, "_cache")
            or getattr(_lookup_cest_por_codigo_descricao_ncm, "_signature", None) != master_sig):

        comp: dict = {}            # (cod_key, desc_norm, ncm_norm) -> cest
        loose: dict = {}           # (cod_key, ncm_norm) -> set[cest]
        seen_ids: set = set()

        for key, meta in master.items():
            if key.startswith("__") or not isinstance(meta, dict):
                continue
            mid = id(meta)
            if mid in seen_ids:
                continue
            seen_ids.add(mid)

            cest = normalize_cest(meta.get("CEST"))
            ncmv = normalize_ncm(meta.get("NCM"))
            if not cest or not ncmv:
                continue

            cod  = clean_str(meta.get("CODIGO"))
            desc = clean_str(meta.get("DESCRICAO_TECNICA"))
            ce   = codigo_produto_key(cod)      if cod  else ""
            cn   = normalize_item_join_key(cod) if cod  else ""
            dn   = norm_text(desc)              if desc else ""

            if ce and dn:
                comp.setdefault((ce, dn, ncmv), cest)
            if cn and dn:
                comp.setdefault((cn, dn, ncmv), cest)
            if ce:
                loose.setdefault((ce, ncmv), set()).add(cest)
            if cn:
                loose.setdefault((cn, ncmv), set()).add(cest)

        _lookup_cest_por_codigo_descricao_ncm._cache     = (comp, loose)
        _lookup_cest_por_codigo_descricao_ncm._signature = master_sig

    comp, loose = _lookup_cest_por_codigo_descricao_ncm._cache
    ce = codigo_produto_key(codigo)
    cn = normalize_item_join_key(codigo)
    dn = norm_text(descricao)
    nn = normalize_ncm(ncm)

    for k in [(ce, dn, nn), (cn, dn, nn)]:
        if all(k) and k in comp:
            return comp[k], "MATRIZ_INPUT_CODIGO_DESCRICAO_NCM"
    for k in [(ce, nn), (cn, nn)]:
        vals = loose.get(k, set()) if (k[0] and k[1]) else set()
        if len(vals) == 1:
            return next(iter(vals)), "MATRIZ_INPUT_CODIGO_NCM_UNICO"
    return "", ""


# =============================================================================
# CENTRAL DO ANALISTA - ajustes manuais auditáveis e links operacionais
# =============================================================================
AJUSTES_ANALISTA_COLS = [
    "ID_AJUSTE", "ID_REFERENCIA", "TIPO_REFERENCIA", "EMPRESA", "MES",
    "CODIGO_PRODUTO", "DESCRICAO_PRODUTO", "TIPO_AJUSTE", "CAMPO_ALVO",
    "ALTERA_CALCULO", "VALOR_ATUAL", "VALOR_AJUSTADO", "VALOR_REFERENCIA_INFORMADO",
    "JUSTIFICATIVA", "FONTE_COMPROVACAO", "CAMINHO_EVIDENCIA", "JUSTIFICATIVA_COMPARABILIDADE",
    "RESPONSAVEL", "RESPONSAVEL_VALIDACAO_FISCAL", "DATA_DECISAO", "STATUS_AJUSTE"
]
AJUSTES_ANALISTA_EDITAVEIS = {
    "TIPO_AJUSTE", "CAMPO_ALVO", "VALOR_AJUSTADO", "VALOR_REFERENCIA_INFORMADO", "JUSTIFICATIVA",
    "FONTE_COMPROVACAO", "CAMINHO_EVIDENCIA", "JUSTIFICATIVA_COMPARABILIDADE",
    "RESPONSAVEL", "RESPONSAVEL_VALIDACAO_FISCAL", "DATA_DECISAO", "STATUS_AJUSTE"
}
TIPOS_AJUSTE_VALIDOS = {
    "APROVAR_ALERTA_PRECO", "CORRIGIR_QTD_CALCULO", "CORRIGIR_FATOR_UNIDADE",
    "BLOQUEAR_ITEM", "CORRIGIR_CADASTRO_PRODUTO", "APROVAR_FALLBACK",
    "REJEITAR_FALLBACK", "INFORMAR_PRECO_REFERENCIA", "CORRIGIR_SALDO_INVENTARIO",
    "REVISAR_CLASSIFICACAO_OPERACAO"
}

def _ajuste_altera_calculo(tipo_ajuste):
    t = norm_text(tipo_ajuste)
    if t in {"APROVARALERTAPRECO", "APROVARFALLBACK", "REVISARCLASSIFICACAOOPERACAO"}:
        return "Nao"
    if t in {"INFORMARPRECOREFERENCIA", "CORRIGIRQTCALCULO", "CORRIGIRFATORUNIDADE", "CORRIGIRSALDOINVENTARIO", "BLOQUEARITEM", "REJEITARFALLBACK"}:
        return "Sim"
    if t == "CORRIGIRCADASTROPRODUTO":
        return "Indireto"
    return "Revisar"

MOTIVOS_EXCLUSAO_INFORMATIVOS = {
    "TRANSFERENCIA_INTERCOMPANY", "TRANSFERENCIA", "BONIFICACAO", "DEVOLUCAO_VAS_SAC",
    "REM_INDUSTRIALIZACAO", "INDUSTRIALIZACAO_MO", "REMESSA_CONSERTO",
    "DEVOLUCAO_USO_CONS", "OUTRAS_SAIDAS", "NCM_FORA_DO_ESCOPO",
    "PRODUTO_NAO_PARTICIPA_DO_ARBITRAMENTO", "ITEM_NAO_E_PRODUTO_ACABADO"
}
MOTIVOS_ACAO_MOVIMENTO = {
    "DIVERGENCIA_NA_QUANTIDADE", "SEM_FATOR_CAIXA", "QUANTIDADE_INVALIDA",
    "TIPO_ITEM_NAO_IDENTIFICADO", "CFOP_NAO_CLASSIFICADO_COMO_VENDA",
    "SEM_CADASTRO_MATRIZ", "CADASTRO_INATIVO", "PART_ARBITRAMENTO_NAO_INFORMADO",
    "DIVERGENCIA_PART_ARBITRAMENTO", "VINCULO_GRUPO_NAO_INFORMADO",
    "CATEGORIA_ITEM_NAO_INFORMADA", "TIPO_PRODUTO_NAO_IDENTIFICADO",
    "UNIDADE_MEDIDA_INCONSISTENTE", "CFOP_TRANSFERENCIA_DESTINO_EXTERNO",
    "TRANSFERENCIA_MESMA_RAIZ_CNPJ_NAO_CADASTRADA", "CRITERIO_CUSTEIO_RIR70_REQUER_REVISAO",
    "CUSTO_AQUISICAO_REQUER_ROTINA_PROPRIA", "VALOR_ZERO"
}

def _base_dir_path(value, default):
    raw = clean_str(value) or default
    path = Path(raw)
    if not path.is_absolute():
        path = BASE_DIR / path
    return path

def _ajustes_folder(config):
    return _base_dir_path(config.get("pasta_ajustes_analista"), "input/ajustes")

def _ajustes_file_path(config):
    return _ajustes_folder(config) / (clean_str(config.get("arquivo_ajustes_analista")) or "AJUSTES_ANALISTA_RIR70.xlsx")

def _template_ajustes_path(config):
    OUTPUT_DIR.mkdir(exist_ok=True)
    # O template de ajustes deixa de utilizar o prefixo "TEMPLATE_" para manter nomenclatura coerente com o arquivo operacional solicitado. O
    # timestamp continua presente para garantir unicidade em execuções paralelas.
    return OUTPUT_DIR / f"AJUSTES_ANALISTA_RIR70_{RUN_TS:%Y%m%d_%H%M%S}.xlsx"

def _file_uri(path):
    try:
        return Path(path).resolve().as_uri()
    except Exception:
        return clean_str(path)

def _link_ajuste(config):
    return _file_uri(_ajustes_file_path(config))

def _make_id(prefix, *parts):
    """Gera um identificador estável a partir de partes de texto.

    Para transparência em auditorias fiscais e para evitar colisões em
    ambientes paralelos, o algoritmo utiliza SHA-256 em vez de SHA-1 e
    produz um valor em hexadecimal. Quando todas as partes fornecidas
    estão vazias, gera-se um identificador único baseado em UUID.  O
    prefixo deve ser uma sigla que identifica a origem (ARB, INV, MOV).
    """
    # Normaliza cada parte para string; partes vazias tornam-se strings vazias.
    normalized_parts = [clean_str(p).upper() for p in parts]
    # Se todas as partes estão vazias, retorna um identificador único.
    if not any(normalized_parts):
        import uuid
        uid = uuid.uuid4().hex[:16].upper()
        return f"{prefix}_UNIQ_{uid}"
    raw = "|".join(normalized_parts)
    h = hashlib.sha256(raw.encode("utf-8")).hexdigest()[:16].upper()
    return f"{prefix}_{h}"

def _id_arbitramento(row):
    return _make_id("ARB", row.get("EMPRESA"), row.get("MES"), row.get("CODIGO_PRODUTO"), row.get("NCM"))

def _id_inventario(row):
    return _make_id("INV", row.get("EMPRESA"), row.get("MES"), row.get("ALMOX"), row.get("CODIGO_PRODUTO"), row.get("NCM"))

def _id_movimento_record(r):
    # Constrói identificador do item de movimentação.  Quando todos os
    # componentes relevantes estão vazios, _make_id retornará um
    # identificador único para evitar colisões entre diferentes itens
    # vazios (por exemplo, linhas sem código ou chave).
    return _make_id(
        "MOV",
        r.get("CNPJ Emitente"),
        r.get("Competência"),
        r.get("Chave NF-e"),
        r.get("_item_nf"),
        r.get("Código Item"),
        r.get("NCM"),
    )

def _nivel_por_advertencia(advertencia, situacao=""):
    txt = norm_text(f"{advertencia} {situacao}")
    if any(x in txt for x in ["CRITICO", "MUITOACIMA", "BLOQUEADA", "SEMPRECO", "INVALIDA", "AUSENTE", "NCMDIVERGENTE"]):
        return "CRITICO"
    if any(x in txt for x in ["REVISAR", "ACIMADOPADRAO", "OUTRAEMPRESA", "SIMILAR", "MANUAL", "MESANTERIOR", "ATENCAO", "POUCASVENDAS"]):
        return "REVISAR"
    return "REVISAR"

def _tipo_ajuste_sugerido(tipo_pendencia):
    t = norm_text(tipo_pendencia)
    if "CFOPTRANSFERENCIADESTINOEXTERNO" in t or "TRANSFERENCIAMESMARAIZCNPJNAOCADASTRADA" in t or "CFOPNAOCLASSIFICADOCOMOVENDA" in t:
        return "REVISAR_CLASSIFICACAO_OPERACAO"
    if "PRECOACIMADOPADRAO" in t or "PRECOMUITOACIMADOPADRAO" in t or "ALERTAPRECO" in t:
        return "APROVAR_ALERTA_PRECO"
    if "QUANTIDADE" in t or "QTDCALCULO" in t:
        return "CORRIGIR_QTD_CALCULO"
    if "FATOR" in t or "UNIDADE" in t:
        return "CORRIGIR_FATOR_UNIDADE"
    if "MESANTERIOR" in t:
        return "APROVAR_FALLBACK"
    if "OUTRAEMPRESA" in t or "SIMILAR" in t:
        return "APROVAR_FALLBACK"
    if "REFERENCIAMANUAL" in t or "SEMPRECO" in t or "PRECOLOCALIZADO" in t:
        return "INFORMAR_PRECO_REFERENCIA"
    if "NCMDIVERGENTE" in t or "NCM" in t:
        return "CORRIGIR_CADASTRO_PRODUTO"
    if "SALDO" in t or "INVENTARIO" in t:
        return "CORRIGIR_SALDO_INVENTARIO"
    if "CADASTRO" in t or "PARTARBITRAMENTO" in t or "VINCULOGRUPO" in t or "CATEGORIAITEM" in t or "TIPOPRODUTO" in t or "CRITERIOCUSTEIO" in t or "CUSTOAQUISICAO" in t:
        return "CORRIGIR_CADASTRO_PRODUTO"
    return "APROVAR_ALERTA_PRECO"

def _campo_ajuste_sugerido(tipo_pendencia):
    t = norm_text(tipo_pendencia)
    if "CFOP" in t or "TRANSFERENCIA" in t: return "CLASSIFICACAO_OPERACAO"
    if "CRITERIOCUSTEIO" in t or "CUSTOAQUISICAO" in t: return "CRITERIO_CUSTEIO_RIR70"
    if "QUANTIDADE" in t: return "QTD_CALCULO"
    if "FATOR" in t or "UNIDADE" in t: return "FATOR_UNIDADE"
    if "PARTARBITRAMENTO" in t: return "PART_ARBITRAMENTO"
    if "VINCULOGRUPO" in t: return "VINCULO_GRUPO"
    if "CATEGORIAITEM" in t: return "CATEGORIA_ITEM"
    if "NCMDIVERGENTE" in t or "NCM" in t: return "NCM"
    if "TIPOPRODUTO" in t or "CADASTRO" in t: return "MATRIZ_CADASTRO"
    if "SEMPRECO" in t or "REFERENCIAMANUAL" in t: return "MAIOR_PRECO_VENDA"
    if "MESANTERIOR" in t: return "ORIGEM_DO_PRECO"
    if "SALDO" in t: return "SALDO_ESTOQUE"
    return "SITUACAO_BASE"

def _arquivo_ajuste_sugerido(tipo_pendencia, config):
    t = norm_text(tipo_pendencia)
    if "NCMDIVERGENTE" in t:
        return "MATRIZ_CADASTRO_PRODUTO_RIR70_SIMILARIDADE.xlsx", "MATRIZ_CADASTRO", _file_uri(BASE_DIR / "input" / "auxiliares" / "MATRIZ_CADASTRO_PRODUTO_RIR70_SIMILARIDADE.xlsx")
    if "CADASTRO" in t or "PARTARBITRAMENTO" in t or "VINCULOGRUPO" in t or "CATEGORIAITEM" in t or "TIPOPRODUTO" in t:
        return "MATRIZ_CADASTRO_PRODUTO_RIR70_SIMILARIDADE.xlsx", "MATRIZ_CADASTRO", _file_uri(BASE_DIR / "input" / "auxiliares" / "MATRIZ_CADASTRO_PRODUTO_RIR70_SIMILARIDADE.xlsx")
    if "SIMILAR" in t:
        return "MAPA_SIMILARIDADE_PRODUTO.xlsx", "MAPA_SIMILARIDADE", _file_uri(BASE_DIR / "input" / "auxiliares" / "MAPA_SIMILARIDADE_PRODUTO.xlsx")
    if "SEMPRECO" in t or "REFERENCIAMANUAL" in t or "PRECOLOCALIZADO" in t:
        return "precos_referencia.xlsx", "PRECOS_REFERENCIA", _file_uri(BASE_DIR / "input" / "auxiliares" / "precos_referencia.xlsx")
    if "MOVIMENTO" in t:
        return "PASTA_MOVIMENTO_ITEM", "-", _file_uri(BASE_DIR / "input" / "movimento_item")
    if "INVENTARIO" in t:
        return "PASTA_INVENTARIO", "-", _file_uri(BASE_DIR / "input" / "inventario")
    return clean_str(config.get("arquivo_ajustes_analista")) or "AJUSTES_ANALISTA_RIR70.xlsx", "AJUSTES", _link_ajuste(config)

def _acao_analista(tipo_pendencia):
    t = norm_text(tipo_pendencia)
    if "CFOPTRANSFERENCIADESTINOEXTERNO" in t:
        return "Revisar se a transferência para destinatário externo deve ser tratada como operação fora do cálculo, venda, devolução ou erro cadastral. Usar TIPO_AJUSTE = REVISAR_CLASSIFICACAO_OPERACAO e anexar evidência."
    if "TRANSFERENCIAMESMARAIZCNPJNAOCADASTRADA" in t:
        return "CNPJ de mesma raiz do grupo não está cadastrado na lista fechada de empresas. Confirmar se faz parte do grupo; se sim, atualizar config/cnpjs_grupo antes de reprocessar."
    if "CRITERIOCUSTEIO" in t or "CUSTOAQUISICAO" in t:
        return "Revisar CRITERIO_CUSTEIO_RIR70 na matriz. Use RIR70_70_MAIOR_PRECO apenas para PA ou MR vinculado ao grupo; MR de terceiros exige CUSTO_AQUISICAO e rotina própria."
    if "PRECOACIMADOPADRAO" in t or "PRECOMUITOACIMADOPADRAO" in t:
        return "Validar documento do maior preço. Se a venda for real, preencher AJUSTES_ANALISTA_RIR70.xlsx com TIPO_AJUSTE = APROVAR_ALERTA_PRECO; se houver erro de quantidade/fator, informar o ajuste correspondente."
    if "QUANTIDADE" in t:
        return "Validar QTDE_COM, QTDE_TRIB e QTD_CALCULO no movimento/XML. Corrigir via AJUSTES_ANALISTA_RIR70.xlsx ou na origem e reprocessar."
    if "FATOR" in t or "UNIDADE" in t:
        return "Validar fator caixa/unidade. Corrigir em AJUSTES_ANALISTA_RIR70.xlsx ou fatores_manuais.xlsx e reprocessar."
    if "SEMPRECO" in t or "PRECOLOCALIZADO" in t:
        return "Informar preço de referência com fonte documental ou corrigir base de vendas/fallback. O motor relerá a informação na próxima execução."
    if "REFERENCIAMANUAL" in t:
        return "Validar a fonte do preço manual e manter evidência. Se correto, aprovar no arquivo de ajustes."
    if "MESANTERIOR" in t:
        return "Preço de mês anterior utilizado como fallback. Validar se o valor é comparável à competência atual. Se aceitável, aprovar em AJUSTES_ANALISTA_RIR70.xlsx com TIPO_AJUSTE = APROVAR_FALLBACK."
    if "OUTRAEMPRESA" in t or "SIMILAR" in t:
        return "Validar comparabilidade do preço/fallback. Se correto, aprovar no arquivo de ajustes; se incorreto, rejeitar fallback e informar preço referência."
    if "NCMDIVERGENTE" in t:
        return "NCM do inventário difere do NCM no movimento fiscal. Corrigir o campo NCM na MATRIZ_CADASTRO_PRODUTO_RIR70_SIMILARIDADE.xlsx e reprocessar."
    if "MOVIMENTO" in t:
        return "Incluir arquivo de movimento fiscal na pasta indicada ou confirmar ausência de movimento no config."
    if "INVENTARIO" in t:
        return "Incluir arquivo de inventário na pasta indicada ou confirmar ausência de estoque no config."
    if "CADASTRO" in t or "PARTARBITRAMENTO" in t or "VINCULOGRUPO" in t or "CATEGORIAITEM" in t or "TIPOPRODUTO" in t:
        return "Corrigir a matriz cadastral indicada e reprocessar o motor."
    return "Revisar a linha e preencher o arquivo de ajustes indicado se houver intervenção operacional."

def _evidencia_obrigatoria(tipo_pendencia):
    t = norm_text(tipo_pendencia)
    if "OUTRAEMPRESA" in t or "SIMILAR" in t:
        return "NF-e/relatório de venda + justificativa de comparabilidade."
    if "REFERENCIAMANUAL" in t or "SEMPRECO" in t:
        return "Documento de suporte do preço de referência em input/evidencias."
    if "CFOP" in t or "TRANSFERENCIA" in t:
        return "Documento fiscal e confirmação da natureza da operação."
    if "CADASTRO" in t or "NCM" in t or "CRITERIOCUSTEIO" in t:
        return "Matriz cadastral revisada e fonte fiscal/cadastral."
    return "Justificativa do responsável e fonte comprobatória quando aplicável."


def _risco_se_nao_corrigir(tipo_pendencia):
    t = norm_text(tipo_pendencia)
    if "REJEITARFALLBACK" in t or "SEMPRECO" in t or "PRECOLOCALIZADO" in t:
        return "Item pode ficar bloqueado/sem valorização na próxima execução."
    if "CFOP" in t or "TRANSFERENCIA" in t:
        return "Operação pode ser classificada incorretamente e afetar a base de preço."
    if "CRITERIOCUSTEIO" in t or "CUSTOAQUISICAO" in t:
        return "Risco de aplicar 70% do maior preço a item que exige outro critério de custeio."
    if "NCM" in t or "CADASTRO" in t:
        return "Cadastro incorreto pode bloquear cálculo ou gerar base fiscal indevida."
    return "Pendência permanecerá acionável e impedirá fechamento sem justificativa."


def _make_pendencia(tipo_ref, tipo_pend, empresa="", mes="", codigo="", descricao="", documento="", ncm="", config=None, situacao="PENDENTE"):
    cfg = config or DEFAULT_CONFIG
    id_ref = _make_id(tipo_ref[:3].upper() or "REF", empresa, mes, codigo, ncm, documento, tipo_pend)
    arq, aba, link = _arquivo_ajuste_sugerido(tipo_pend, cfg)
    return {
        "ID_REFERENCIA": id_ref,
        "TIPO_REFERENCIA": tipo_ref,
        "EMPRESA": empresa,
        "MES": mes,
        "CODIGO_PRODUTO": codigo,
        "DESCRICAO_PRODUTO": descricao,
        "TIPO_PENDENCIA": tipo_pend,
        "NIVEL": _nivel_por_advertencia(tipo_pend, situacao),
        "INTERVENCAO_ANALISTA": "SIM" if situacao == "PENDENTE" else "NAO",
        "DESCRICAO_PENDENCIA": clean_str(tipo_pend).replace("_", " ").title(),
        "ACAO_ANALISTA": _acao_analista(tipo_pend),
        "ARQUIVO_AJUSTE": arq,
        "ABA_AJUSTE": aba,
        "CAMPO_AJUSTE": _campo_ajuste_sugerido(tipo_pend),
        "TIPO_AJUSTE_SUGERIDO": _tipo_ajuste_sugerido(tipo_pend),
        "ALTERA_CALCULO": _ajuste_altera_calculo(_tipo_ajuste_sugerido(tipo_pend)),
        "EVIDENCIA_OBRIGATORIA": _evidencia_obrigatoria(tipo_pend),
        "RISCO_SE_NAO_CORRIGIR": _risco_se_nao_corrigir(tipo_pend),
        "LINK_ABRIR_AJUSTE": link,
        "DECISAO_ANALISTA": "",
        "DATA_DECISAO": "",
        "RESPONSAVEL_DECISAO": "",
        "SITUACAO": situacao,
    }

def _load_ajustes_analista(config):
    path = _ajustes_file_path(config)
    result = {}
    if not path.exists():
        return result
    try:
        df = pd.read_excel(path, sheet_name="AJUSTES", dtype=object, keep_default_na=False)
    except Exception as exc:
        logger.warning("Ajustes analista: falha ao ler %s: %s", path, exc)
        return result
    for _, row in df.iterrows():
        rec = {clean_str(c): clean_str(row.get(c, "")) for c in df.columns}
        idref = clean_str(rec.get("ID_REFERENCIA"))
        tipo = clean_str(rec.get("TIPO_AJUSTE")).upper()
        status = clean_str(rec.get("STATUS_AJUSTE")).upper()
        if not idref or tipo not in TIPOS_AJUSTE_VALIDOS or status not in {"APROVADO", "APROVADO PARA APLICAR", "APROVADO_PARA_APLICAR"}:
            continue
        if not clean_str(rec.get("JUSTIFICATIVA")) or not clean_str(rec.get("RESPONSAVEL")) or not clean_str(rec.get("DATA_DECISAO")):
            logger.warning("Ajuste analista ignorado por falta de evidência/responsável/data: %s", idref)
            continue
        result[idref] = rec
    return result

def _apply_operator_adjustments_to_arbitramento(arb_rows, config):
    ajustes = _load_ajustes_analista(config)
    for r in arb_rows:
        idr = _id_arbitramento(r)
        r["ID_ARBITRAMENTO"] = idr
        r.setdefault("AJUSTE_ANALISTA_APLICADO", "NAO")
        r.setdefault("ID_AJUSTE", "")
        r.setdefault("JUSTIFICATIVA_ANALISTA", "")
        adv = clean_str(r.get("ADVERTENCIA"))
        if idr in ajustes:
            aj = ajustes[idr]
            tipo = clean_str(aj.get("TIPO_AJUSTE")).upper()
            if tipo == "APROVAR_ALERTA_PRECO":
                r["SITUACAO_BASE"] = "APROVADA"
                r["ALERTA"] = "● OK"
                r["ADVERTENCIA"] = (adv + "_VALIDADO_PELO_ANALISTA") if adv else "VALIDADO_PELO_ANALISTA"
                r["INTERVENCAO_ANALISTA"] = "NAO"
                r["ACAO_ANALISTA"] = "Alerta de preço aprovado pelo analista com justificativa registrada."
                r["AJUSTE_ANALISTA_APLICADO"] = "SIM"
                r["ID_AJUSTE"] = clean_str(aj.get("ID_AJUSTE")) or _make_id("AJ", idr, tipo, aj.get("VALOR_AJUSTADO"))
                r["JUSTIFICATIVA_ANALISTA"] = clean_str(aj.get("JUSTIFICATIVA"))
                continue
        if clean_str(r.get("SITUACAO_BASE")) == "REVISAR" or norm_text(adv) in {"PRECOACIMADOPADRAO", "PRECOMUITOACIMADOPADRAO"} or "PRECOACIMA" in norm_text(adv):
            r["INTERVENCAO_ANALISTA"] = "SIM"
            r["ACAO_ANALISTA"] = _acao_analista(adv or "ALERTA_PRECO")
            r["ARQUIVO_AJUSTE"] = clean_str(config.get("arquivo_ajustes_analista")) or "AJUSTES_ANALISTA_RIR70.xlsx"
            r["LINK_ABRIR_AJUSTE"] = _link_ajuste(config)
        else:
            r["INTERVENCAO_ANALISTA"] = "NAO"
            r["ACAO_ANALISTA"] = "Nenhuma intervenção operacional obrigatória."
            r["ARQUIVO_AJUSTE"] = ""
            r["LINK_ABRIR_AJUSTE"] = ""
    return arb_rows

def _apply_operator_adjustments_to_inventario(inv_rows, config):
    ajustes = _load_ajustes_analista(config)
    for r in inv_rows:
        idr = _id_inventario(r)
        r["ID_INVENTARIO"] = idr
        r.setdefault("AJUSTE_ANALISTA_APLICADO", "NAO")
        r.setdefault("ID_AJUSTE", "")
        r.setdefault("JUSTIFICATIVA_ANALISTA", "")
        sit = clean_str(r.get("SITUACAO_PRECO"))
        origem = clean_str(r.get("ORIGEM_DO_PRECO"))
        adv = clean_str(r.get("ADVERTENCIA"))
        if idr in ajustes:
            aj = ajustes[idr]
            tipo = clean_str(aj.get("TIPO_AJUSTE")).upper()
            if tipo in {"APROVAR_FALLBACK", "INFORMAR_PRECO_REFERENCIA"}:
                r["SITUACAO_PRECO"] = "APROVADA" if tipo == "APROVAR_FALLBACK" else "REVISAR"
                r["ALERTA"] = "● OK" if tipo == "APROVAR_FALLBACK" else "● REVISAR"
                r["INTERVENCAO_ANALISTA"] = "NAO" if tipo == "APROVAR_FALLBACK" else "SIM"
                r["ACAO_ANALISTA"] = "Fallback/preço validado pelo analista com justificativa registrada."
                r["AJUSTE_ANALISTA_APLICADO"] = "SIM"
                r["ID_AJUSTE"] = clean_str(aj.get("ID_AJUSTE")) or _make_id("AJ", idr, tipo, aj.get("VALOR_AJUSTADO"))
                r["JUSTIFICATIVA_ANALISTA"] = clean_str(aj.get("JUSTIFICATIVA"))
                continue
        if sit in {"BLOQUEADA", "REVISAR"} or origem in {"PRECO_REFERENCIA_MANUAL", "PRECO_USADO_DE_OUTRA_EMPRESA", "PRECO_USADO_DE_PRODUTO_SIMILAR", "SEM_PRECO_LOCALIZADO"}:
            r["INTERVENCAO_ANALISTA"] = "SIM"
            r["ACAO_ANALISTA"] = _acao_analista(adv or origem)
            arq, _, link = _arquivo_ajuste_sugerido(adv or origem, config)
            r["ARQUIVO_AJUSTE"] = arq
            r["LINK_ABRIR_AJUSTE"] = link
        else:
            r["INTERVENCAO_ANALISTA"] = "NAO"
            r["ACAO_ANALISTA"] = "Nenhuma intervenção operacional obrigatória."
            r["ARQUIVO_AJUSTE"] = ""
            r["LINK_ABRIR_AJUSTE"] = ""
    return inv_rows

def _write_ajustes_xlsx(path, rows, config, title="AJUSTES ANALISTA RIR70"):
    path = Path(path); path.parent.mkdir(parents=True, exist_ok=True)
    font = config.get("fonte_excel", "Arial Narrow")
    size = int(config.get("tamanho_fonte_excel", 10) or 10)
    with pd.ExcelWriter(path, engine="xlsxwriter", datetime_format="dd/mm/yyyy", date_format="dd/mm/yyyy") as writer:
        wb = writer.book
        ws = wb.add_worksheet("AJUSTES"); writer.sheets["AJUSTES"] = ws
        ws.set_tab_color("#7F1D1D")
        ws.hide_gridlines(2)
        fmt_h      = wb.add_format({"font_name":font,"font_size":size,"bold":True,"font_color":"#FFFFFF","bg_color":"#7F1D1D","border":1,"border_color":"#5C1414","align":"center","valign":"vcenter","text_wrap":True})
        fmt_locked = wb.add_format({"font_name":font,"font_size":size,"border":1,"border_color":"#D9E2F3","bg_color":"#F2F2F2","font_color":"#444444","locked":True,"valign":"vcenter"})
        fmt_locked_alt = wb.add_format({"font_name":font,"font_size":size,"border":1,"border_color":"#D9E2F3","bg_color":"#EAEAEA","font_color":"#444444","locked":True,"valign":"vcenter"})
        # Formatos numéricos para colunas bloqueadas com 6 casas decimais (p.ex. VALOR_ATUAL)
        fmt_locked_num6 = wb.add_format({"font_name":font,"font_size":size,"border":1,"border_color":"#D9E2F3","bg_color":"#F2F2F2","font_color":"#444444","locked":True,"valign":"vcenter","num_format":"#,##0.000000"})
        fmt_locked_num6_alt = wb.add_format({"font_name":font,"font_size":size,"border":1,"border_color":"#D9E2F3","bg_color":"#EAEAEA","font_color":"#444444","locked":True,"valign":"vcenter","num_format":"#,##0.000000"})
        fmt_input  = wb.add_format({"font_name":font,"font_size":size,"border":1,"border_color":"#D4A017","bg_color":"#FFFBEB","font_color":"#1E293B","locked":False,"valign":"vcenter"})
        fmt_input_alt = wb.add_format({"font_name":font,"font_size":size,"border":1,"border_color":"#D4A017","bg_color":"#FEF9E7","font_color":"#1E293B","locked":False,"valign":"vcenter"})
        fmt_pend   = wb.add_format({"font_name":font,"font_size":size,"bold":True,"border":1,"border_color":"#F4B183","bg_color":"#FCE4D6","font_color":"#9C0006","locked":False,"valign":"vcenter","align":"center"})
        ws.write_row(0, 0, AJUSTES_ANALISTA_COLS, fmt_h)
        ws.set_row(0, 22)
        for ri, row in enumerate(rows, start=1):
            alt = (ri % 2 == 0)
            ws.set_row(ri, 16)
            for ci, col in enumerate(AJUSTES_ANALISTA_COLS):
                val = row.get(col, "")
                if col == "STATUS_AJUSTE" and clean_str(val).upper() == "PENDENTE":
                    ws.write(ri, ci, _fmt_value_for_excel(val), fmt_pend)
                elif col in AJUSTES_ANALISTA_EDITAVEIS:
                    ws.write(ri, ci, _fmt_value_for_excel(val), fmt_input_alt if alt else fmt_input)
                else:
                    # Aplica formato numérico de 6 casas decimais para VALOR_ATUAL, mantendo bloqueio de edição
                    if col == "VALOR_ATUAL":
                        ws.write(ri, ci, _fmt_value_for_excel(val), fmt_locked_num6_alt if alt else fmt_locked_num6)
                    else:
                        ws.write(ri, ci, _fmt_value_for_excel(val), fmt_locked_alt if alt else fmt_locked)
        widths = {"ID_AJUSTE":20,"ID_REFERENCIA":22,"TIPO_REFERENCIA":15,"EMPRESA":18,"MES":12,"CODIGO_PRODUTO":18,"DESCRICAO_PRODUTO":45,"TIPO_AJUSTE":28,"CAMPO_ALVO":24,"ALTERA_CALCULO":16,"VALOR_ATUAL":18,"VALOR_AJUSTADO":18,"VALOR_REFERENCIA_INFORMADO":24,"JUSTIFICATIVA":55,"FONTE_COMPROVACAO":40,"CAMINHO_EVIDENCIA":45,"JUSTIFICATIVA_COMPARABILIDADE":55,"RESPONSAVEL":22,"RESPONSAVEL_VALIDACAO_FISCAL":28,"DATA_DECISAO":15,"STATUS_AJUSTE":18}
        for ci, col in enumerate(AJUSTES_ANALISTA_COLS):
            ws.set_column(ci, ci, widths.get(col, 18))
        ws.freeze_panes(1, 0)
        ws.autofilter(0, 0, max(1, len(rows)), len(AJUSTES_ANALISTA_COLS)-1)
        if rows:
            tipo_col = AJUSTES_ANALISTA_COLS.index("TIPO_AJUSTE")
            status_col = AJUSTES_ANALISTA_COLS.index("STATUS_AJUSTE")
            ws.data_validation(1, tipo_col, max(1, len(rows)), tipo_col, {
                "validate": "list",
                "source": sorted(TIPOS_AJUSTE_VALIDOS),
                "input_title": "Tipo de Ajuste",
                "input_message": "Aprovar alerta/fallback não muda cálculo. Informar preço exige valor e evidência. Rejeitar fallback pode bloquear o item.",
                "error_title": "Tipo inválido",
                "error_message": "Selecione um valor da lista. Não é permitido digitar diretamente.",
                "error_type": "stop",
            })
            ws.data_validation(1, status_col, max(1, len(rows)), status_col, {
                "validate": "list",
                "source": ["APROVADO", "PENDENTE", "REJEITADO"],
                "input_title": "Status do Ajuste",
                "input_message": "APROVADO = ajuste validado e pronto para o motor processar. PENDENTE = aguarda decisão. REJEITADO = descartado.",
                "error_type": "stop",
            })
            for numeric_col_name in ["VALOR_AJUSTADO", "VALOR_REFERENCIA_INFORMADO"]:
                if numeric_col_name in AJUSTES_ANALISTA_COLS:
                    ncol = AJUSTES_ANALISTA_COLS.index(numeric_col_name)
                    ws.data_validation(1, ncol, max(1, len(rows)), ncol, {
                        "validate": "decimal", "criteria": ">=", "value": 0,
                        "input_title": "Valor numérico",
                        "input_message": "Preencha apenas quando o tipo de ajuste alterar cálculo. Aceita número positivo; use formato monetário normal do Excel.",
                        "error_title": "Valor inválido", "error_message": "Informe número maior ou igual a zero.", "error_type": "stop",
                    })

        # ====================================================================
        #  Instruções para o analista: nova planilha "INSTRUCOES"
        #
        #  Conforme solicitado, geramos uma segunda aba contendo orientações
        #  detalhadas sobre cada coluna editável. Isso ajuda o analista a
        #  entender quais tipos de informações devem ser preenchidas e quando
        #  fornecer justificativas ou evidências. O cabeçalho é sombreado e
        #  utiliza cores corporativas, e o texto das instruções ocupa a
        #  segunda coluna. Ajustamos larguras de coluna para melhorar
        #  legibilidade.
        instr_ws = wb.add_worksheet("INSTRUCOES")
        writer.sheets["INSTRUCOES"] = instr_ws
        instr_ws.set_tab_color("#5C1414")
        instr_ws.hide_gridlines(2)
        # Formatos para a planilha de instruções
        fmt_h_instr = wb.add_format({"font_name":font,"font_size":size,"bold":True,"font_color":"#FFFFFF","bg_color":"#7F1D1D","border":1,"border_color":"#5C1414","align":"center","valign":"vcenter"})
        fmt_col = wb.add_format({"font_name":font,"font_size":size,"border":1,"border_color":"#D9E2F3","bg_color":"#F2F2F2","valign":"top","text_wrap":True})
        fmt_desc = wb.add_format({"font_name":font,"font_size":size,"border":1,"border_color":"#D9E2F3","bg_color":"#F7F9FC","valign":"top","text_wrap":True})
        # Cabeçalho
        instr_ws.write_row(0, 0, ["Coluna", "Instrução de Preenchimento"], fmt_h_instr)
        instr_ws.set_row(0, 20)
        # Lista de instruções por coluna
        instrucciones = [
            ("VALOR_REFERENCIA_INFORMADO", "Informe o preço de referência quando não houver valor automático localizado ou quando houver discrepância significativa. Deve ser um número positivo com até 6 casas decimais, usando vírgula como separador decimal no Excel."),
            ("JUSTIFICATIVA", "Descreva de forma sucinta por que o ajuste é necessário. Explique o critério adotado e como o novo valor foi obtido."),
            ("FONTE_COMPROVACAO", "Indique a fonte de onde o valor foi retirado, por exemplo: contrato, nota fiscal, tabela de preços ou outro documento oficial."),
            ("CAMINHO_EVIDENCIA", "Informe o caminho da pasta ou arquivo onde a evidência está armazenada, permitindo a rastreabilidade do ajuste."),
            ("JUSTIFICATIVA_COMPARABILIDADE", "Para preços de produtos comparáveis, explique por que considera a comparação apropriada (p.ex., mesma categoria, características técnicas etc.)."),
            ("RESPONSAVEL", "Nome da pessoa que está realizando o ajuste. Use a forma completa com apenas a primeira letra de cada nome em maiúsculo."),
            ("RESPONSAVEL_VALIDACAO_FISCAL", "Nome do responsável pela validação fiscal. Utilize apenas a primeira letra de cada nome em maiúsculo."),
            ("DATA_DECISAO", "Data em que o ajuste foi analisado. Utilize o formato DD/MM/AAAA."),
            ("STATUS_AJUSTE", "Selecione na lista a situação do ajuste: APROVADO, PENDENTE ou REJEITADO. Apenas ajustes aprovados são processados pelo motor.")
        ]
        for ri, (col_name, desc) in enumerate(instrucciones, start=1):
            instr_ws.set_row(ri, 30)
            instr_ws.write(ri, 0, col_name, fmt_col)
            instr_ws.write(ri, 1, desc, fmt_desc)
        # Definir larguras de colunas
        instr_ws.set_column(0, 0, 28)
        instr_ws.set_column(1, 1, 95)
        # Adiciona comentário na célula A1 da planilha de ajustes para orientar o analista sobre qual arquivo usar. O
        # comentário explica que o template gerado no output é apenas uma cópia para auditoria e que o
        # arquivo que deve ser lido pelo motor está em input/ajustes/AJUSTES_ANALISTA_RIR70.xlsx.
        ws.write_comment(0, 0, "Arquivo operacional: salve e mantenha este arquivo em input/ajustes/AJUSTES_ANALISTA_RIR70.xlsx. O template no output é cópia de auditoria; não é o arquivo lido pelo motor.")
        # Protege a planilha de ajustes para evitar alterações acidentais nas células bloqueadas, mantendo liberação
        # apenas para células editáveis e recursos de filtro e formatação de linhas/colunas.
        ws.protect("", {"select_locked_cells": True, "select_unlocked_cells": True, "format_columns": True, "format_rows": True, "autofilter": True})

def _rows_template_from_pendencias(pend_rows):
    rows = []
    for p in pend_rows:
        if clean_str(p.get("INTERVENCAO_ANALISTA")) != "SIM" or clean_str(p.get("SITUACAO")) != "PENDENTE":
            continue
        idref = clean_str(p.get("ID_REFERENCIA"))
        tipo = clean_str(p.get("TIPO_AJUSTE_SUGERIDO")) or _tipo_ajuste_sugerido(p.get("TIPO_PENDENCIA"))
        # VALOR_ATUAL: preenchido com o valor de referência transportado do arb/inv row,
        # evitando que o analista precise consultar o output para saber o que está aprovando.
        valor_atual = clean_str(p.get("_VALOR_REFERENCIA")) or ""
        rows.append({
            "ID_AJUSTE": _make_id("AJ", idref, tipo),
            "ID_REFERENCIA": idref,
            "TIPO_REFERENCIA": p.get("TIPO_REFERENCIA"),
            "EMPRESA": p.get("EMPRESA"),
            "MES": p.get("MES"),
            "CODIGO_PRODUTO": p.get("CODIGO_PRODUTO"),
            "DESCRICAO_PRODUTO": p.get("DESCRICAO_PRODUTO"),
            "TIPO_AJUSTE": tipo,
            "CAMPO_ALVO": p.get("CAMPO_AJUSTE"),
            "ALTERA_CALCULO": _ajuste_altera_calculo(tipo),
            "VALOR_ATUAL": valor_atual,
            "VALOR_AJUSTADO": "",
            "VALOR_REFERENCIA_INFORMADO": "",
            "JUSTIFICATIVA": "",
            "FONTE_COMPROVACAO": "",
            "CAMINHO_EVIDENCIA": "",
            "JUSTIFICATIVA_COMPARABILIDADE": "",
            "RESPONSAVEL": "",
            "RESPONSAVEL_VALIDACAO_FISCAL": "",
            "DATA_DECISAO": "",
            "STATUS_AJUSTE": "PENDENTE",
        })
    return rows

def _merge_existing_ajustes(existing_path, new_rows):
    if not Path(existing_path).exists():
        return new_rows
    try:
        df = pd.read_excel(existing_path, sheet_name="AJUSTES", dtype=object, keep_default_na=False)
        existing = [{c: row.get(c, "") for c in df.columns} for _, row in df.iterrows()]
    except Exception:
        return new_rows
    ids = {clean_str(r.get("ID_REFERENCIA")) for r in existing if clean_str(r.get("ID_REFERENCIA"))}
    for r in new_rows:
        if clean_str(r.get("ID_REFERENCIA")) not in ids:
            existing.append(r)
    # Garante colunas na ordem atual
    out = []
    for r in existing:
        out.append({c: r.get(c, "") for c in AJUSTES_ANALISTA_COLS})
    return out

def _gerar_arquivos_ajustes_analista(pend_rows, config):
    if not _is_sim(config.get("gerar_template_ajustes_analista", True)):
        return None, None
    template_rows = _rows_template_from_pendencias(pend_rows)
    template_path = _template_ajustes_path(config)
    _write_ajustes_xlsx(template_path, template_rows, config)
    ajustes_path = _ajustes_file_path(config)
    if _is_sim(config.get("criar_arquivo_ajustes_analista_se_ausente", True)):
        rows_to_write = _merge_existing_ajustes(ajustes_path, template_rows) if _is_sim(config.get("preservar_ajustes_analista_existentes", True)) else template_rows
        _write_ajustes_xlsx(ajustes_path, rows_to_write, config)
    config["_template_ajustes_analista"] = str(template_path)
    config["_arquivo_ajustes_analista"] = str(ajustes_path)
    return template_path, ajustes_path

def build_movimentacao_rows(records):
    rows = []
    for r in records:
        adv = r.get("ADVERTENCIA")
        motivo = clean_str(r.get("MOTIVO"))
        calcula = r.get("ENTRA_NO_CALCULO")
        if clean_str(calcula) == "Sim":
            interv = "NAO"; acao = "Nenhuma intervenção operacional obrigatória."
        elif motivo in MOTIVOS_EXCLUSAO_INFORMATIVOS:
            interv = "INFORMATIVO"; acao = "Nenhuma. Operação excluída corretamente do cálculo."
        elif motivo in MOTIVOS_ACAO_MOVIMENTO:
            interv = "SIM"; acao = _acao_analista(motivo)
        else:
            interv = "NAO"; acao = "Nenhuma intervenção obrigatória; item não participante conforme regra operacional."
        rows.append({
            "EMPRESA": r.get("CNPJ Emitente"),
            "PERFIL_EMPRESA": r.get("PERFIL_EMPRESA"),
            "MES": r.get("Competência"),
            "DATA_EMISSAO": r.get("Data Emissão"),
            "DOCUMENTO": r.get("Número NF"),
            "CHAVE_NFE": r.get("Chave NF-e"),
            "CFOP": r.get("CFOP"),
            "CODIGO_PRODUTO": r.get("Código Item"),
            "DESCRICAO_PRODUTO": r.get("Descrição"),
            "TIPO_ITEM": r.get("TIPO_ITEM") or r.get("Tipo Item"),
            "FONTE_TIPO_ITEM": r.get("FONTE_TIPO_ITEM"),
            "TIPO_PRODUTO": r.get("TIPO_PRODUTO"),
            "STATUS_CADASTRO_MATRIZ": r.get("STATUS_CADASTRO_MATRIZ"),
            "Part_Arbitramento": r.get("PART_ARBITRAMENTO"),
            "CALCULA_MATRIZ": r.get("CALCULA_MATRIZ"),
            "VINCULO_GRUPO": r.get("VINCULO_GRUPO"),
            "CATEGORIA_ITEM": r.get("CATEGORIA_ITEM"),
            "CRITERIO_CUSTEIO_RIR70": r.get("CRITERIO_CUSTEIO_RIR70"),
            "FUNDAMENTO_CRITERIO_CUSTEIO": r.get("FUNDAMENTO_CRITERIO_CUSTEIO"),
            "FATOR_UNIDADE_MATRIZ": r.get("FATOR_UNIDADE_MATRIZ"),
            "NCM": r.get("NCM"),
            "NCM_MATRIZ": r.get("NCM_MATRIZ"),
            "CEST": r.get("CEST"),
            "VALOR_PRODUTO": r.get("_valor_total") or r.get("Valor Comercial"),
            "QTD_ORIGINAL": r.get("QTD_ORIGINAL"),
            "UM_ORIGINAL": r.get("UNIDADE_ORIGINAL"),
            "FATOR_CAIXA": r.get("FATOR_CAIXA"),
            "QTD_CALCULO": r.get("QTD_CALCULO"),
            "VALOR_UNITARIO": r.get("VALOR_UNITARIO_CALCULO"),
            "SITUACAO_QTD": r.get("SITUACAO_QTD"),
            "CALCULA": calcula,
            "MOTIVO": motivo,
            "ALERTA": _alerta_visual(adv, calcula, motivo),
            "ADVERTENCIA": adv,
            "INTERVENCAO_ANALISTA": interv,
            "ACAO_ANALISTA": acao,
        })
    return rows

def build_arbitramento_rows(records, config=None):
    cfg = config or DEFAULT_CONFIG
    grouped = {}
    for r in records:
        if not r.get("_participa"):
            continue
        k = r.get("_monthly_key")
        g = grouped.setdefault(k, {"rows": []})
        g["rows"].append(r)
    rows = []
    minimo = int(cfg.get("minimo_vendas_para_alerta", 3) or 3)
    pct1 = Decimal(str(cfg.get("preco_acima_padrao_percentual", 50)))
    pct2 = Decimal(str(cfg.get("preco_muito_acima_padrao_percentual", 100)))
    for g in grouped.values():
        gr = g["rows"]
        first = gr[0]
        valid_pairs = []
        for r in gr:
            unit = r.get("VALOR_UNITARIO_CALCULO")
            qtd = r.get("QTD_CALCULO") if isinstance(r.get("QTD_CALCULO"), Decimal) else to_decimal(r.get("QTD_CALCULO"))
            if isinstance(unit, Decimal) and isinstance(qtd, Decimal) and qtd > ZERO:
                valid_pairs.append((unit, qtd, r))
        if not valid_pairs:
            continue
        units = [u for u, _, _ in valid_pairs]
        maxv = q6(max(units))
        soma_qtd = sum((q for _, q, _ in valid_pairs), ZERO)
        soma_valor = sum((u * q for u, q, _ in valid_pairs), ZERO)
        avg = q6(soma_valor / soma_qtd) if soma_qtd > ZERO else None
        diff = q6(((maxv - avg) / avg) * Decimal("100")) if isinstance(avg, Decimal) and avg > ZERO else None
        base = q6(maxv * Decimal("0.70"))
        docs = []
        for r in gr:
            if isinstance(r.get("VALOR_UNITARIO_CALCULO"), Decimal) and q6(r.get("VALOR_UNITARIO_CALCULO")) == maxv:
                d = clean_str(r.get("Número NF")) or clean_str(r.get("Chave NF-e"))
                if d and d not in docs:
                    docs.append(d)
        sit = "APROVADA"; adv = ""
        if avg is None or avg <= ZERO:
            sit = "REVISAR"; adv = "PRECO_MEDIO_ZERO_SEM_COMPARACAO"
        elif len(valid_pairs) < minimo:
            adv = "POUCAS_VENDAS_NO_MES"
        if isinstance(diff, Decimal):
            if diff > pct2:
                sit = "REVISAR"; adv = "PRECO_MUITO_ACIMA_DO_PADRAO"
            elif diff > pct1:
                sit = "REVISAR"; adv = "PRECO_ACIMA_DO_PADRAO"
        # Arredondamento da variação de preço em percentuais inteiros para apresentação
        if isinstance(diff, Decimal):
            # Número inteiro (arredondado) da diferença percentual
            diff_int = diff.quantize(Decimal('1'), rounding=ROUND_HALF_UP)
            # Valor em percentagem para Excel (ex.: 71% será representado como 0.71).
            # Quando a diferença é zero (0%), não exibimos valor na célula (fica vazio).
            if diff_int <= 0:
                diff_pct_value = None
            else:
                diff_pct_value = (diff_int / Decimal('100'))
        else:
            diff_int = None
            diff_pct_value = None
        # Somatório de quantidade vendida para este agrupamento
        qtde_total_vendida = soma_qtd
        rows.append({
            "EMPRESA": first.get("CNPJ Emitente"),
            "PERFIL_EMPRESA": first.get("PERFIL_EMPRESA"),
            "MES": first.get("Competência"),
            "CODIGO_PRODUTO": first.get("Código Item"),
            "DESCRICAO_PRODUTO": first.get("Descrição"),
            "TIPO_ITEM": first.get("TIPO_ITEM") or first.get("Tipo Item"),
            "FONTE_TIPO_ITEM": first.get("FONTE_TIPO_ITEM"),
            "TIPO_PRODUTO": first.get("TIPO_PRODUTO"),
            "STATUS_CADASTRO_MATRIZ": first.get("STATUS_CADASTRO_MATRIZ"),
            "Part_Arbitramento": first.get("PART_ARBITRAMENTO"),
            "CALCULA_MATRIZ": first.get("CALCULA_MATRIZ"),
            "VINCULO_GRUPO": first.get("VINCULO_GRUPO"),
            "CATEGORIA_ITEM": first.get("CATEGORIA_ITEM"),
            "CRITERIO_CUSTEIO_RIR70": first.get("CRITERIO_CUSTEIO_RIR70"),
            "FUNDAMENTO_CRITERIO_CUSTEIO": first.get("FUNDAMENTO_CRITERIO_CUSTEIO"),
            "FATOR_UNIDADE_MATRIZ": first.get("FATOR_UNIDADE_MATRIZ"),
            "NCM": first.get("NCM"),
            "NCM_MATRIZ": first.get("NCM_MATRIZ"),
            "CEST": first.get("CEST"),
            "MAIOR_PRECO_UNITARIO": maxv,
            "CUSTO_ARBITRADO_70": base,
            "PRECO_MEDIO": avg,
            # Guarda o valor percentual para uso na planilha. Para Excel, usa-se formato de porcentagem (0.71 = 71%).
            "VARIACAO_PRECO (%)": diff_pct_value,
            "DOCUMENTO_MAIOR_PRECO": "; ".join(docs),
            "ORIGEM_DO_PRECO": "PRECO_DO_PROPRIO_PRODUTO_NO_MES",
            # A situação do resumo reflete a qualidade do preço usado, não a situação transacional individual.
            "SITUACAO_BASE": sit,
            "MOTIVO": "VENDA_VALIDA",
            "ALERTA": _alerta_visual(adv, sit, "VENDA_VALIDA"),
            "ADVERTENCIA": adv,
            # Quantidade vendida agregada no período (utilizado posteriormente para CMV no inventário)
            "QTD_VENDIDA": qtde_total_vendida,
        })
    rows.sort(key=lambda r: (competencia_sort_key(r.get("MES")), clean_str(r.get("EMPRESA")), clean_str(r.get("CODIGO_PRODUTO"))))
    return rows

def _add_inventory_record(agg, cod_original, comp, descricao, unidade, ncm, almox, qtde, fonte, cnpj_inv="", config=None):
    cod_original = normalize_code_text(cod_original)
    cod_key = codigo_produto_key(cod_original)
    if not cod_key or qtde is None:
        return False
    cnpj_inv = normalize_cnpj(cnpj_inv)
    almox_n = normalize_almox(almox) or "N/D"
    ncm_n = normalize_ncm(ncm)
    cfg = config or DEFAULT_CONFIG
    matriz_meta = _matriz_produto_meta(cod_original, cnpj_inv, cfg)
    fator_matriz = _matriz_decimal(matriz_meta.get("FATOR_UNIDADE")) if matriz_meta else None
    if cfg.get("usar_fator_unidade_matriz_inventario", True) and fator_matriz:
        fator_dec, status_fator, obs_fator = fator_matriz, "OK", "FATOR_UNIDADE_MATRIZ"
    else:
        fator_dec, status_fator, obs_fator = _inventario_factor_from_unit_or_desc(unidade, descricao, cfg)
    qtde_trib = qtde * fator_dec
    unid_trib = derive_unid_tributavel(unidade, int(fator_dec) if fator_dec else 1, descricao)
    # Almoxarifados da lista consolidada são somados em uma única chave para evitar duplicação de estoque.
    almox_consolidados = (
        {normalize_almox(a) for a in cfg.get("almoxarifados_inventario_pa", []) if normalize_almox(a)}
        if cfg.get("filtrar_almoxarifado_inventario_pa", True) else set()
    )
    almox_key = "CONSOLIDADO" if almox_n in almox_consolidados else almox_n
    key = (cnpj_inv, cod_key, comp, almox_key)
    rec = agg.get(key)
    if rec is None:
        rec = {"qtde": ZERO, "qtde_original": ZERO, "codigos_originais": Counter(), "descricoes": Counter(), "unidades": Counter(), "unidades_trib": Counter(), "fatores": Counter(), "ncms": Counter(), "almox": Counter(), "linhas": 0, "fontes": Counter()}
        agg[key] = rec
    rec["qtde"] += qtde_trib
    rec["qtde_original"] += qtde
    rec["codigos_originais"][clean_str(cod_original)] += 1
    if clean_str(descricao): rec["descricoes"][clean_str(descricao)] += 1
    if clean_str(unidade): rec["unidades"][clean_str(unidade)] += 1
    if clean_str(unid_trib): rec["unidades_trib"][clean_str(unid_trib)] += 1
    rec["fatores"][str(int(fator_dec))] += 1
    if ncm_n: rec["ncms"][ncm_n] += 1
    rec["almox"][almox_n] += 1
    rec["linhas"] += 1
    return True

def _ncm_mais_frequente_por_codigo(arb_rows):
    out = {}
    counts = defaultdict(Counter)
    for row in arb_rows:
        cod = codigo_produto_key(row.get("CODIGO_PRODUTO"))
        ncm = normalize_ncm(row.get("NCM"))
        if cod and ncm:
            counts[cod][ncm] += 1
    for cod, counter in counts.items():
        out[cod] = counter.most_common(1)[0][0]
    return out

def build_inventario_rows(arb_rows, inventory_map, config):
    ncm_ref_por_codigo = _ncm_mais_frequente_por_codigo(arb_rows)
    exact_current, group_current = {}, {}
    exact_history, group_history = defaultdict(list), defaultdict(list)
    by_code = defaultdict(list)
    # Pre-computado uma vez fora de todos os loops; evita recriação de set por item de inventário
    cnpjs_grupo_set = config.get("_CNPJS_GRUPO_SET") or frozenset(normalize_cnpj(c) for c in _empresas_grupo_cnpjs(config) if normalize_cnpj(c))
    for row in arb_rows:
        emp = normalize_cnpj(row.get("EMPRESA")); cod = codigo_produto_key(row.get("CODIGO_PRODUTO")); mes = clean_str(row.get("MES"))
        if not emp or not cod or not mes or not isinstance(row.get("CUSTO_ARBITRADO_70"), Decimal):
            continue
        exact_current[(emp, cod, mes)] = row
        group_current[(cod, mes, emp)] = row
        exact_history[(emp, cod)].append(row)
        group_history[cod].append(row)
        by_code[cod].append(row)
    for hist in list(exact_history.values()) + list(group_history.values()):
        hist.sort(key=lambda r: competencia_sort_key(r.get("MES")))
    def prev(hist, mes):
        # Usa o mês anterior disponível mais próximo, limitado por MAX_MESES_RETROATIVOS_PRECO.
        t = competencia_sort_key(mes); best = None
        for cand in hist:
            c = competencia_sort_key(cand.get("MES"))
            if c < t and _fallback_mes_dentro_limite(cand.get("MES"), mes, config):
                best = cand
            elif c >= t:
                break
        return best
    def adv_mes_anterior(cand, mes_destino, base="PRECO_DE_MES_ANTERIOR"):
        if not cand:
            return ""
        return _fmt_defasagem_preco(cand.get("MES"), mes_destino, base)
    sim_map = load_similarity_map(config) if config.get("permitir_fallback_base_produto_similar", True) else {}
    # Preços de referência manuais - último recurso antes de SEM_PRECO_LOCALIZADO
    ref_prices = load_precos_referencia(config)
    group_to_arbs = defaultdict(list)
    for cod, rows in by_code.items():
        meta = sim_map.get(cod)
        if meta and meta.get("grupo"):
            for r in rows:
                group_to_arbs[(meta.get("grupo"), r.get("MES"))].append(r)
                group_to_arbs[(meta.get("grupo"), "*")].append(r)
    rows = []
    for key, inv in inventory_map.items():
        if len(key) == 4:
            emp, cod, mes, almox = key
        else:
            emp, cod, mes = key; almox = "N/D"
        desc = _counter_main(inv.get("descricoes", {}))
        unid = _counter_main(inv.get("unidades_trib", {})) or _counter_main(inv.get("unidades", {}))
        ncm = _counter_main(inv.get("ncms", {}))
        cod_orig = _counter_main(inv.get("codigos_originais", {})) or cod
        matriz_meta_dest = _matriz_produto_meta(cod_orig, emp, config)
        ncm_matriz = normalize_ncm(matriz_meta_dest.get("NCM")) if matriz_meta_dest else ""
        ncm_ref = ncm_matriz or ncm_ref_por_codigo.get(codigo_produto_key(cod_orig) or cod)
        alerta_ncm = ""
        if ncm and ncm_ref and normalize_ncm(ncm) != normalize_ncm(ncm_ref):
            fonte_ncm_ref = "MATRIZ" if ncm_matriz else "MOVIMENTO"
            alerta_ncm = f"ALERTA_NCM_DIVERGENTE_INVENTARIO={normalize_ncm(ncm)}_{fonte_ncm_ref}={normalize_ncm(ncm_ref)}"
        matriz_bloqueia, matriz_motivo, matriz_adv = _validar_meta_matriz_para_calculo(matriz_meta_dest, config)
        saldo = inv.get("qtde")
        arb = exact_current.get((emp, cod, mes)); origem = "PRECO_DO_PROPRIO_PRODUTO_NO_MES"; adv = alerta_ncm
        if not arb:
            cand = prev(exact_history.get((emp, cod), []), mes)
            if cand:
                arb = cand; origem = "PRECO_DO_PROPRIO_PRODUTO_EM_MES_ANTERIOR"; adv = adv_mes_anterior(cand, mes)
        if not arb and config.get("permitir_fallback_base_inventario_entre_cnpjs", True) and emp in cnpjs_grupo_set:
            # O(k) lookup direto por empresa do grupo - sem varredura linear de group_current
            for other_emp in cnpjs_grupo_set:
                if other_emp == emp:
                    continue
                cand0 = group_current.get((cod, mes, other_emp))
                if cand0:
                    arb = cand0; origem = "PRECO_USADO_DE_OUTRA_EMPRESA"
                    adv = (adv + ";" if adv else "") + "PRECO_DE_OUTRA_EMPRESA_DO_GRUPO"; break
        if not arb and config.get("permitir_fallback_base_inventario_entre_cnpjs", True) and emp in cnpjs_grupo_set:
            cands_hist = [x for x in group_history.get(cod, []) if normalize_cnpj(x.get("EMPRESA")) != emp and normalize_cnpj(x.get("EMPRESA")) in cnpjs_grupo_set]
            cand = prev(cands_hist, mes)
            if cand:
                arb = cand; origem = "PRECO_USADO_DE_OUTRA_EMPRESA"
                adv_def = adv_mes_anterior(cand, mes)
                adv = (adv + ";" if adv else "") + "PRECO_DE_OUTRA_EMPRESA_DO_GRUPO" + (f";{adv_def}" if adv_def else "")
        if not arb and sim_map.get(cod):
            grupo = sim_map[cod].get("grupo")
            cands = [x for x in group_to_arbs.get((grupo, mes), []) if codigo_produto_key(x.get("CODIGO_PRODUTO")) != cod]
            if not cands:
                cands = [x for x in group_to_arbs.get((grupo, "*"), []) if codigo_produto_key(x.get("CODIGO_PRODUTO")) != cod and competencia_sort_key(x.get("MES")) < competencia_sort_key(mes) and _fallback_mes_dentro_limite(x.get("MES"), mes, config)]
            if cands:
                cands.sort(key=lambda r: (competencia_sort_key(r.get("MES")), clean_str(r.get("CODIGO_PRODUTO"))))
                arb = cands[-1]; origem = "PRECO_USADO_DE_PRODUTO_SIMILAR"; adv = (adv + ";" if adv else "") + "PRECO_DE_PRODUTO_SIMILAR"

        # ── Fallback final: preço de referência manual (precos_referencia.xlsx) ────
        # Acionado somente quando todos os demais fallbacks falharam.
        # O analista informa MAIOR_PRECO_VENDA + FONTE_PRECO; o motor aplica 70%.
        # Situação sempre marcada como REVISAR para obrigar revisão antes da entrega ao Fisco.
        ref_entry = None
        if not arb:
            ref_entry = _find_preco_referencia(cod_orig, ncm, mes, ref_prices, emp)
            if ref_entry:
                arb = {
                    "CUSTO_ARBITRADO_70":    ref_entry["base_70"],
                    "EMPRESA":               "",
                    "CODIGO_PRODUTO":        cod_orig,
                    "MES":                   mes,
                    "DOCUMENTO_MAIOR_PRECO": ref_entry["fonte"],
                    "NCM":                   ncm,
                }
                origem = "PRECO_REFERENCIA_MANUAL"
                adv    = _join_adv(adv, "PRECO_MANUAL_ANALISTA_AGUARDA_REVISAO")
        if arb:
            base = arb.get("CUSTO_ARBITRADO_70")
            valor = q2(base * saldo) if isinstance(base, Decimal) and isinstance(saldo, Decimal) else None
            # Preço de referência manual é sempre REVISAR - exige revisão formal antes do Fisco
            sit = "REVISAR" if origem == "PRECO_REFERENCIA_MANUAL" else ("APROVADA" if not adv else "REVISAR")
            emp_origem = arb.get("EMPRESA")
            prod_origem = arb.get("CODIGO_PRODUTO")
            mes_origem = arb.get("MES")
            doc_origem = arb.get("DOCUMENTO_MAIOR_PRECO")
            ncm_eff = ncm or arb.get("NCM")
            tipo, fonte_tipo = _infer_tipo_item(ncm_eff, desc, config, cod_orig, unid, emp, return_source=True)
        else:
            base = valor = None; sit = "BLOQUEADA"; origem = "SEM_PRECO_LOCALIZADO"; adv = (adv + ";" if adv else "") + "PRODUTO_SEM_PRECO_LOCALIZADO"
            emp_origem = prod_origem = mes_origem = doc_origem = ""; ncm_eff = ncm; tipo, fonte_tipo = _infer_tipo_item(ncm_eff, desc, config, cod_orig, unid, emp, return_source=True)
        if matriz_bloqueia:
            base = None
            valor = None
            sit = "BLOQUEADA"
            origem = matriz_motivo
            adv = _join_adv(adv, matriz_adv, matriz_motivo)
        if not _tipo_item_calcula(tipo):
            base = None
            valor = None
            sit = "BLOQUEADA"
            origem = "ITEM_NAO_E_PRODUTO_ACABADO"
            adv = _join_adv(adv, "ITEM_NAO_E_PRODUTO_ACABADO")
        desc_tecnica = _lookup_descricao_tecnica(cod_orig, desc, config)
        justificativa_fallback = ""
        maior_preco_ref = None
        justificativa_preco_ref = ""
        if origem == "PRECO_REFERENCIA_MANUAL" and ref_entry:
            maior_preco_ref = ref_entry.get("preco")
            justificativa_preco_ref = (
                f"PREÇO DE REFERÊNCIA MANUAL - Base: R$ {float(ref_entry['preco']):.4f}/un trib. "
                f"Fonte: {ref_entry['fonte'] or 'NÃO INFORMADA'}. "
                f"Justificativa: {ref_entry['just'] or 'Produto sem histórico de venda externa; preço inserido pelo analista.'}. "
                "OBRIGATÓRIO revisar comparabilidade e anexar documento de suporte antes da entrega ao Fisco."
            )
            justificativa_fallback = justificativa_preco_ref
        elif origem == "PRECO_USADO_DE_OUTRA_EMPRESA":
            if normalize_cnpj(emp_origem) == "03408722000178":
                justificativa_fallback = (
                    f"PREÇO OBTIDO POR FALLBACK DA EMPRESA {emp_origem} (FÁBRICA DO GRUPO). "
                    f"Conforme auditoria, este CNPJ tende a operar com transferências intercompany; preço utilizado apenas por ausência de histórico no CNPJ {emp} no período {mes}. "
                    "SUJEITO A QUESTIONAMENTO EM AUDITORIA FISCAL - revisar comparabilidade antes de entrega ao Fisco."
                )
            else:
                justificativa_fallback = (
                    f"Preço do próprio produto adotado de empresa do mesmo grupo ({emp_origem}) por ausência de histórico de vendas no CNPJ {emp} no período {mes}. "
                    "Não constitui preço de transferência por si só; é fallback operacional por produto idêntico, sujeito a validação de comparabilidade."
                )
        # Quantidade vendida: sempre do próprio CNPJ/produto/mês — nunca do fallback de preço.
        # O fallback de preço (mês anterior, outra empresa, preço manual) serve apenas para
        # valorar o estoque; a quantidade vendida deve refletir o volume real do período corrente.
        arb_qtd = exact_current.get((emp, cod, mes))
        qtde_vendida = ZERO
        if arb_qtd and isinstance(arb_qtd.get("QTD_VENDIDA"), Decimal):
            qtde_vendida = arb_qtd.get("QTD_VENDIDA")
        # CMV = quantidade vendida * custo arbitrado (base). Utiliza ZERO se qualquer valor estiver ausente.
        cmv_val = ZERO
        if isinstance(qtde_vendida, Decimal) and isinstance(base, Decimal):
            cmv_val = q6(qtde_vendida * base)
        # Monta o registro para a linha de inventário valorizado.  A chave MAIOR_PRECO_REF não é mais
        # exibida no output final (foi removida das colunas), mas é mantida localmente para eventual
        # diagnóstico interno.  DOCUMENTO_ORIGEM será reposicionado na ordem final de colunas.
        row = {
            "EMPRESA": emp,
            "MES": mes,
            "ALMOX": almox,
            "CODIGO_PRODUTO": cod_orig,
            "DESCRICAO_TECNICA": desc_tecnica,
            "NCM": ncm_eff,
            "NCM_MATRIZ": ncm_matriz,
            "CEST": matriz_meta_dest.get("CEST") if matriz_meta_dest else "",
            "TIPO_ITEM": tipo,
            "FONTE_TIPO_ITEM": fonte_tipo,
            "PERFIL_EMPRESA": _perfil_empresa(emp, config),
            "TIPO_PRODUTO": _display_tipo_produto(matriz_meta_dest.get("TIPO_PRODUTO") if matriz_meta_dest else ""),
            "STATUS_CADASTRO_MATRIZ": matriz_meta_dest.get("STATUS_CADASTRO") if matriz_meta_dest else "",
            "Part_Arbitramento": _display_part_arbitramento(matriz_meta_dest),
            "CALCULA_MATRIZ": matriz_meta_dest.get("CALCULA_MATRIZ") if matriz_meta_dest else "",
            "VINCULO_GRUPO": matriz_meta_dest.get("VINCULO_GRUPO") if matriz_meta_dest else "",
            "CATEGORIA_ITEM": matriz_meta_dest.get("CATEGORIA_ITEM") if matriz_meta_dest else "",
            "CRITERIO_CUSTEIO_RIR70": _criterio_custeio_rir70(matriz_meta_dest, config)[0] if matriz_meta_dest else "BLOQUEAR_REVISAR",
            "FUNDAMENTO_CRITERIO_CUSTEIO": _criterio_custeio_rir70(matriz_meta_dest, config)[1] if matriz_meta_dest else "Sem cadastro de matriz para definir critério de custeio.",
            "FATOR_UNIDADE_MATRIZ": matriz_meta_dest.get("FATOR_UNIDADE") if matriz_meta_dest else "",
            "UNIDADE": unid,
            "CUSTO_ARBITRADO_70": base,
            "QUANTIDADE_VENDIDA": qtde_vendida,
            "CMV": cmv_val,
            "SALDO_ESTOQUE": q4(saldo) if isinstance(saldo, Decimal) else saldo,
            "VALOR_ESTOQUE_ARBITRADO": valor,
            "ORIGEM_DO_PRECO": origem,
            "EMPRESA_ORIGEM_PRECO": emp_origem,
            "PRODUTO_ORIGEM_PRECO": prod_origem,
            "DESCRICAO_PRODUTO": desc,
            "MES_ORIGEM_PRECO": mes_origem,
            "DOCUMENTO_ORIGEM": doc_origem,
            "SITUACAO_PRECO": sit,
            "ALERTA": _alerta_visual(adv, sit, origem),
            "ADVERTENCIA": adv,
            "JUSTIFICATIVA_FALLBACK": justificativa_fallback,
            "JUSTIFICATIVA_PRECO_REF": justificativa_preco_ref,
        }
        # Adiciona a linha construída à lista
        rows.append(row)
    rows.sort(key=lambda r: (competencia_sort_key(r.get("MES")), clean_str(r.get("EMPRESA")), clean_str(r.get("CODIGO_PRODUTO")), clean_str(r.get("ALMOX"))))
    return rows

def _build_pendencias(records, arb_rows, inv_rows, config):
    """Gera somente pendências acionáveis.

    Exclusões operacionais normais (ex.: TRANSFERENCIA_INTERCOMPANY, bonificação,
    devolução normal, CFOP_NAO_CALCULA, NCM fora do escopo quando esperado) não
    entram na 04_PENDENCIAS. Elas permanecem rastreadas na 01_MOVIMENTO_VALIDADO
    como informativo, sem ação do analista.
    """
    pend = []
    emp_mov = {normalize_cnpj(r.get("CNPJ Emitente")) for r in records if r.get("CNPJ Emitente")}
    emp_inv = {normalize_cnpj(r.get("EMPRESA")) for r in inv_rows if r.get("EMPRESA")}
    sem_mov_conf = {normalize_cnpj(x) for x in config.get("empresas_sem_movimento_confirmadas", []) or []}
    sem_est_conf = {normalize_cnpj(x) for x in config.get("empresas_sem_estoque_confirmadas", []) or []}

    for emp in _empresas_grupo_cnpjs(config):
        if emp not in emp_mov and emp not in sem_mov_conf:
            pend.append(_make_pendencia(
                "MOVIMENTO", "EMPRESA_SEM_ARQUIVO_MOVIMENTO", empresa=emp, config=config,
                situacao="PENDENTE"
            ))
        if emp not in emp_inv and emp not in sem_est_conf:
            pend.append(_make_pendencia(
                "INVENTARIO", "EMPRESA_SEM_ARQUIVO_INVENTARIO", empresa=emp, config=config,
                situacao="PENDENTE"
            ))

    # Alertas estatísticos do maior preço: entram como pendência de validação,
    # mas não alteram cálculo até que o analista aprove ou corrija a causa.
    for r in arb_rows:
        if clean_str(r.get("AJUSTE_ANALISTA_APLICADO")) == "SIM":
            continue
        adv = clean_str(r.get("ADVERTENCIA"))
        sit = clean_str(r.get("SITUACAO_BASE"))
        if sit == "REVISAR" and adv:
            p = _make_pendencia(
                "ARBITRAMENTO", adv or "ALERTA_PRECO", empresa=r.get("EMPRESA"), mes=r.get("MES"),
                codigo=r.get("CODIGO_PRODUTO"), descricao=r.get("DESCRICAO_PRODUTO"),
                documento=r.get("DOCUMENTO_MAIOR_PRECO"), ncm=r.get("NCM"), config=config, situacao="PENDENTE"
            )
            # ID_REFERENCIA deve ser exatamente o ID da linha na 02_ARBITRAMENTO,
            # para o analista aprovar pelo arquivo AJUSTES_ANALISTA_RIR70.xlsx.
            p["ID_REFERENCIA"] = r.get("ID_ARBITRAMENTO") or _id_arbitramento(r)
            p["TIPO_AJUSTE_SUGERIDO"] = "APROVAR_ALERTA_PRECO"
            p["CAMPO_AJUSTE"] = "SITUACAO_BASE"
            p["ACAO_ANALISTA"] = _acao_analista(adv)
            p["_VALOR_REFERENCIA"] = clean_str(r.get("MAIOR_PRECO_UNITARIO"))
            pend.append(p)

    # Pendências de movimento: somente itens com ação real do analista.
    for r in records:
        mot = clean_str(r.get("MOTIVO"))
        if mot in MOTIVOS_ACAO_MOVIMENTO:
            p = _make_pendencia(
                "MOVIMENTO", mot, empresa=r.get("CNPJ Emitente"), mes=r.get("Competência"),
                codigo=r.get("Código Item"), descricao=r.get("Descrição"), documento=r.get("Número NF"),
                ncm=r.get("NCM"), config=config, situacao="PENDENTE"
            )
            p["ID_REFERENCIA"] = _id_movimento_record(r)
            pend.append(p)

    # Pendências de inventário/fallback.
    for r in inv_rows:
        if clean_str(r.get("AJUSTE_ANALISTA_APLICADO")) == "SIM":
            continue
        tipo_pend = ""
        if r.get("SITUACAO_PRECO") == "BLOQUEADA":
            tipo_pend = "PRODUTO_SEM_PRECO_LOCALIZADO"
        elif r.get("ORIGEM_DO_PRECO") == "PRECO_REFERENCIA_MANUAL":
            tipo_pend = "PRECO_REFERENCIA_MANUAL_UTILIZADO"
        elif r.get("ORIGEM_DO_PRECO") == "PRECO_USADO_DE_OUTRA_EMPRESA":
            tipo_pend = "PRECO_DE_OUTRA_EMPRESA_REQUER_REVISAO"
        elif r.get("ORIGEM_DO_PRECO") == "PRECO_USADO_DE_PRODUTO_SIMILAR":
            tipo_pend = "PRECO_DE_PRODUTO_SIMILAR_REQUER_REVISAO"
        if tipo_pend:
            p = _make_pendencia(
                "INVENTARIO", tipo_pend, empresa=r.get("EMPRESA"), mes=r.get("MES"),
                codigo=r.get("CODIGO_PRODUTO"), descricao=r.get("DESCRICAO_PRODUTO") or r.get("DESCRICAO_TECNICA"),
                documento=r.get("DOCUMENTO_ORIGEM"), ncm=r.get("NCM"), config=config, situacao="PENDENTE"
            )
            p["ID_REFERENCIA"] = r.get("ID_INVENTARIO") or _id_inventario(r)
            p["_VALOR_REFERENCIA"] = clean_str(r.get("CUSTO_ARBITRADO_70"))
            pend.append(p)

    # GAP-01: preço de mês anterior utilizado como fallback - exige aprovação formal.
    # INTERVENCAO=SIM era gerado em 03_INV mas não havia linha correspondente em 04_PEND
    # nem no TEMPLATE_AJUSTES, impedindo o fluxo de aprovação do analista.
    for r in inv_rows:
        if clean_str(r.get("AJUSTE_ANALISTA_APLICADO")) == "SIM":
            continue
        if clean_str(r.get("ORIGEM_DO_PRECO")) == "PRECO_DO_PROPRIO_PRODUTO_EM_MES_ANTERIOR":
            tipo_pend = "PRECO_USADO_DE_MES_ANTERIOR_REQUER_REVISAO"
            p = _make_pendencia(
                "INVENTARIO", tipo_pend, empresa=r.get("EMPRESA"), mes=r.get("MES"),
                codigo=r.get("CODIGO_PRODUTO"), descricao=r.get("DESCRICAO_PRODUTO") or r.get("DESCRICAO_TECNICA"),
                documento=r.get("DOCUMENTO_ORIGEM"), ncm=r.get("NCM"), config=config, situacao="PENDENTE"
            )
            p["ID_REFERENCIA"] = r.get("ID_INVENTARIO") or _id_inventario(r)
            p["TIPO_AJUSTE_SUGERIDO"] = "APROVAR_FALLBACK"
            p["CAMPO_AJUSTE"] = "ORIGEM_DO_PRECO"
            p["NIVEL"] = "REVISAR"
            p["_VALOR_REFERENCIA"] = clean_str(r.get("CUSTO_ARBITRADO_70"))
            pend.append(p)

    # GAP-02: divergência de NCM entre inventário e movimento - exige correção cadastral.
    # O alerta era gerado em 03_INV via adv, mas sem linha em 04_PEND nem no TEMPLATE.
    for r in inv_rows:
        if clean_str(r.get("AJUSTE_ANALISTA_APLICADO")) == "SIM":
            continue
        adv_inv = clean_str(r.get("ADVERTENCIA"))
        if "ALERTA_NCM_DIVERGENTE" in adv_inv.upper():
            tipo_pend = "ALERTA_NCM_DIVERGENTE_REQUER_VERIFICACAO"
            p = _make_pendencia(
                "INVENTARIO", tipo_pend, empresa=r.get("EMPRESA"), mes=r.get("MES"),
                codigo=r.get("CODIGO_PRODUTO"), descricao=r.get("DESCRICAO_PRODUTO") or r.get("DESCRICAO_TECNICA"),
                documento=r.get("DOCUMENTO_ORIGEM"), ncm=r.get("NCM"), config=config, situacao="PENDENTE"
            )
            p["ID_REFERENCIA"] = r.get("ID_INVENTARIO") or _id_inventario(r)
            p["TIPO_AJUSTE_SUGERIDO"] = "CORRIGIR_CADASTRO_PRODUTO"
            p["CAMPO_AJUSTE"] = "NCM"
            p["NIVEL"] = "CRITICO"
            p["DESCRICAO_PENDENCIA"] = adv_inv  # preserva o detalhe do NCM divergente
            p["_VALOR_REFERENCIA"] = clean_str(r.get("CUSTO_ARBITRADO_70"))
            pend.append(p)

    # Recalcula campos derivados quando o tipo de ajuste foi sobrescrito manualmente no fluxo.
    for p in pend:
        p["ALTERA_CALCULO"] = _ajuste_altera_calculo(p.get("TIPO_AJUSTE_SUGERIDO"))
        p["EVIDENCIA_OBRIGATORIA"] = p.get("EVIDENCIA_OBRIGATORIA") or _evidencia_obrigatoria(p.get("TIPO_PENDENCIA"))
        p["RISCO_SE_NAO_CORRIGIR"] = p.get("RISCO_SE_NAO_CORRIGIR") or _risco_se_nao_corrigir(p.get("TIPO_PENDENCIA"))

    # Limita repetições por ID_REFERENCIA + TIPO_PENDENCIA.
    seen = set(); out = []
    for p in pend:
        sig = (p.get("ID_REFERENCIA", ""), p.get("TIPO_PENDENCIA", ""))
        if sig not in seen:
            seen.add(sig); out.append(p)
    return out

def _extract_cnpj_from_filename(name):
    m = re.search(r"\d{14}", clean_str(name))
    return normalize_cnpj(m.group(0)) if m else ""

def _find_file_by_name(name):
    for folder in [INPUT_DIR, INPUT_DIR/"movimento_item", INPUT_DIR/"documentos", INPUT_DIR/"inventario", INPUT_DIR/"xml", INPUT_DIR/"auxiliares"]:
        if folder.exists():
            for p in folder.rglob("*"):
                if p.is_file() and p.name == name:
                    return p
    return None

def _parse_int_from_desc(desc, key):
    m = re.search(rf"{re.escape(key)}\s*=\s*(\d+)", clean_str(desc), flags=re.I)
    return int(m.group(1)) if m else 0

def _build_fontes_processadas(logs, records, config):
    rows = []
    periodo = f"{format_date_abnt(config.get('periodo_base_inicio'))} até {format_date_abnt(config.get('periodo_base_fim'))}"
    for arq, aba, status, desc in logs:
        if clean_str(arq) == "-":
            continue
        p = _find_file_by_name(clean_str(arq))
        imp = _parse_int_from_desc(desc, "Importadas") or _parse_int_from_desc(desc, "Importados") or _parse_int_from_desc(desc, "Linhas")
        fora = _parse_int_from_desc(desc, "fora_periodo")
        dup = _parse_int_from_desc(desc, "duplicadas")
        vaz = _parse_int_from_desc(desc, "vazias")
        tipo = "XML" if clean_str(aba).upper().startswith("XML") or clean_str(arq).lower().endswith((".xml", ".zip")) else ("INVENTARIO" if "INVENTARIO" in clean_str(status).upper() or "INVENTARIO" in clean_str(arq).upper() else "MOVIMENTO")
        situ = "OK" if clean_str(status).upper() in {"PROCESSADA", "XML_PROCESSADO", "INVENTARIO_PROCESSADO"} else ("REVISAR" if "IGNOR" in clean_str(status).upper() or "NAO_UTILIZADO" in clean_str(status).upper() else "PENDENTE")
        codigo_controle = RIR70_SOURCE_CONTROL_CODES.get(clean_str(arq)) or (sha256_file(p) if p else "")
        recs_arquivo = [r for r in records if clean_str(r.get("_row_origin")).split("|")[0] == clean_str(arq)]
        motivos_cfop = {"TRANSFERENCIA_INTERCOMPANY", "TRANSFERENCIA", "BONIFICACAO", "DEVOLUCAO", "CFOP_NAO_CLASSIFICADO_COMO_VENDA", "CFOP_NAO_E_VENDA"}
        reg_bloq_cfop = sum(1 for r in recs_arquivo if clean_str(r.get("MOTIVO")) in motivos_cfop or clean_str(r.get("Tipo Operação")) in {"TRANSFERENCIA", "CFOP_REVISAR", "CFOP_NAO_E_VENDA"})
        reg_sem_cad = sum(1 for r in recs_arquivo if "SEM_CADASTRO_MATRIZ" in clean_str(r.get("ADVERTENCIA")) or clean_str(r.get("MOTIVO")) == "SEM_CADASTRO_MATRIZ")
        lidas = imp + fora + dup + vaz
        rows.append({
            "ARQUIVO": arq, "TIPO": tipo, "EMPRESA": _extract_cnpj_from_filename(arq), "PERIODO": periodo,
            "REGISTROS_LIDOS": lidas, "REGISTROS_VALIDOS": imp, "REGISTROS_DUPLICADOS": dup,
            "REGISTROS_BLOQUEADOS_CFOP": reg_bloq_cfop, "REGISTROS_SEM_CADASTRO": reg_sem_cad,
            "LINHAS_LIDAS": lidas, "LINHAS_USADAS": imp, "LINHAS_DESCARTADAS": fora + dup + vaz,
            "LINHAS_DUPLICADAS_DESCARTADAS": dup, "MOTIVO_PRINCIPAL_DESCARTE": status,
            "CODIGO_CONTROLE_ARQUIVO": codigo_controle, "ENCODING_USADO": "N/A", "SITUACAO": situ
        })
    # Registra empresas configuradas sem arquivo de movimento para a rastreabilidade de abertura do processamento.
    emp_with_source = {normalize_cnpj(r.get("EMPRESA")) for r in rows if r.get("EMPRESA")}
    emp_mov = {normalize_cnpj(r.get("CNPJ Emitente")) for r in records if r.get("CNPJ Emitente")}
    sem_mov_conf = {normalize_cnpj(x) for x in config.get("empresas_sem_movimento_confirmadas", []) or []}
    for emp in _empresas_grupo_cnpjs(config):
        if emp not in emp_mov and emp not in emp_with_source:
            sit = "OK" if emp in sem_mov_conf else "PENDENTE"
            motivo = "SEM_MOVIMENTO_CONFIRMADO_NO_CONFIG" if emp in sem_mov_conf else "EMPRESA_SEM_ARQUIVO_MOVIMENTO"
            rows.append({"ARQUIVO": "SEM_ARQUIVO_MOVIMENTO", "TIPO": "MOVIMENTO", "EMPRESA": emp, "PERIODO": periodo, "REGISTROS_LIDOS": 0, "REGISTROS_VALIDOS": 0, "REGISTROS_DUPLICADOS": 0, "REGISTROS_BLOQUEADOS_CFOP": 0, "REGISTROS_SEM_CADASTRO": 0, "LINHAS_LIDAS": 0, "LINHAS_USADAS": 0, "LINHAS_DESCARTADAS": 0, "LINHAS_DUPLICADAS_DESCARTADAS": 0, "MOTIVO_PRINCIPAL_DESCARTE": motivo, "CODIGO_CONTROLE_ARQUIVO": "", "ENCODING_USADO": "N/A", "SITUACAO": sit})
    info_ref = config.get("_precos_referencia_info") or {}
    if info_ref.get("arquivo") and info_ref.get("situacao") not in {"NAO_LOCALIZADO", "DESABILITADO"}:
        arq_ref = clean_str(info_ref.get("arquivo"))
        if not any(clean_str(r.get("ARQUIVO")) == arq_ref for r in rows):
            rows.append({"ARQUIVO": arq_ref, "TIPO": "PRECO_REFERENCIA", "EMPRESA": "", "PERIODO": periodo,
                         "REGISTROS_LIDOS": int(info_ref.get("linhas_lidas") or 0), "REGISTROS_VALIDOS": int(info_ref.get("linhas_usadas") or 0),
                         "REGISTROS_DUPLICADOS": 0, "REGISTROS_BLOQUEADOS_CFOP": 0, "REGISTROS_SEM_CADASTRO": 0,
                         "LINHAS_LIDAS": int(info_ref.get("linhas_lidas") or 0), "LINHAS_USADAS": int(info_ref.get("linhas_usadas") or 0),
                         "LINHAS_DESCARTADAS": int(info_ref.get("linhas_descartadas") or 0), "LINHAS_DUPLICADAS_DESCARTADAS": 0,
                         "MOTIVO_PRINCIPAL_DESCARTE": info_ref.get("situacao"), "CODIGO_CONTROLE_ARQUIVO": info_ref.get("codigo_controle") or RIR70_SOURCE_CONTROL_CODES.get(arq_ref, ""),
                         "ENCODING_USADO": "N/A", "SITUACAO": "OK" if info_ref.get("situacao") == "OK" else "REVISAR"})
    return rows

def _fmt_value_for_excel(v):
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, datetime):
        return v
    return v

def _write_final_sheet(wb, ws, columns, rows, config, tab_kind="data"):
    font = config.get("fonte_excel", "Arial Narrow")
    size = int(config.get("tamanho_fonte_excel", 10) or 10)
    palette = {
        "capa": ("#0D2137", "#D9EAF7"),
        "data": ("#1A3A5C", "#EAF2F8"),
        "ok":   ("#1E5C8A", "#EAF2F8"),
        "warn": ("#14532D", "#EAF7EA"),
        "pend": ("#7F1D1D", "#FCE8E6"),
        "src":  ("#4C1D95", "#F3E8FF"),
    }
    header_fill, light_fill = palette.get(tab_kind, ("#1F4E78", "#EAF2F8"))
    fmt_h = wb.add_format({"font_name":font,"font_size":size,"bold":True,"font_color":"#FFFFFF","bg_color":header_fill,"border":1,"border_color":"#1A3A5C","align":"center","valign":"vcenter","text_wrap":True})
    fmt_txt = wb.add_format({"font_name":font,"font_size":size,"border":1,"border_color":"#D9E2F3","valign":"vcenter","text_wrap":False})
    fmt_alt = wb.add_format({"font_name":font,"font_size":size,"border":1,"border_color":"#D9E2F3","valign":"vcenter","bg_color":"#F7F9FC","text_wrap":False})
    fmt_ctr = wb.add_format({"font_name":font,"font_size":size,"border":1,"border_color":"#D9E2F3","valign":"vcenter","align":"center"})
    fmt_ctr_alt = wb.add_format({"font_name":font,"font_size":size,"border":1,"border_color":"#D9E2F3","valign":"vcenter","align":"center","bg_color":"#F7F9FC"})
    fmt_date = wb.add_format({"font_name":font,"font_size":size,"border":1,"border_color":"#D9E2F3","valign":"vcenter","align":"center","num_format":"dd/mm/yyyy"})
    fmt_date_alt = wb.add_format({"font_name":font,"font_size":size,"border":1,"border_color":"#D9E2F3","valign":"vcenter","align":"center","bg_color":"#F7F9FC","num_format":"dd/mm/yyyy"})
    fmt_num2 = wb.add_format({"font_name":font,"font_size":size,"border":1,"border_color":"#D9E2F3","valign":"vcenter","num_format":"#,##0.00"})
    fmt_num2_alt = wb.add_format({"font_name":font,"font_size":size,"border":1,"border_color":"#D9E2F3","valign":"vcenter","bg_color":"#F7F9FC","num_format":"#,##0.00"})
    fmt_num4     = wb.add_format({"font_name":font,"font_size":size,"border":1,"border_color":"#D9E2F3","valign":"vcenter","num_format":"#,##0.0000"})
    fmt_num4_alt = wb.add_format({"font_name":font,"font_size":size,"border":1,"border_color":"#D9E2F3","valign":"vcenter","bg_color":"#F7F9FC","num_format":"#,##0.0000"})
    fmt_num6     = wb.add_format({"font_name":font,"font_size":size,"border":1,"border_color":"#D9E2F3","valign":"vcenter","num_format":"#,##0.000000"})
    fmt_num6_alt = wb.add_format({"font_name":font,"font_size":size,"border":1,"border_color":"#D9E2F3","valign":"vcenter","bg_color":"#F7F9FC","num_format":"#,##0.000000"})
    fmt_pct = wb.add_format({"font_name":font,"font_size":size,"border":1,"border_color":"#D9E2F3","valign":"vcenter","align":"center","num_format":"0%"})
    fmt_pct_alt = wb.add_format({"font_name":font,"font_size":size,"border":1,"border_color":"#D9E2F3","valign":"vcenter","align":"center","bg_color":"#F7F9FC","num_format":"0%"})
    fmt_pct_alert = wb.add_format({"font_name":font,"font_size":size,"bold":True,"border":1,"border_color":"#F4B183","bg_color":"#FFF2CC","font_color":"#7F6000","valign":"vcenter","align":"center","num_format":"0%"})
    fmt_capa_status_ok   = wb.add_format({"font_name":font,"font_size":size,"bold":True,"border":1,"border_color":"#A9D18E","bg_color":"#E2F0D9","font_color":"#375623","valign":"vcenter","align":"center"})
    fmt_capa_status_warn = wb.add_format({"font_name":font,"font_size":size,"bold":True,"border":1,"border_color":"#FFD966","bg_color":"#FFF2CC","font_color":"#7F6000","valign":"vcenter","align":"center"})
    fmt_capa_status_crit = wb.add_format({"font_name":font,"font_size":size,"bold":True,"border":1,"border_color":"#F4B183","bg_color":"#FCE4D6","font_color":"#9C0006","valign":"vcenter","align":"center"})
    fmt_capa_sub_header  = wb.add_format({"font_name":font,"font_size":size,"bold":True,"border":1,"border_color":"#0D2137","bg_color":"#1A3A5C","font_color":"#D9EAF7","valign":"vcenter","align":"left","text_wrap":True})
    # Formato para células VALOR da CAPA com texto longo: quebra automática + alinhamento justificado.
    fmt_capa_long_val    = wb.add_format({"font_name":font,"font_size":size,"border":1,"border_color":"#D9E2F3","valign":"top","align":"justify","text_wrap":True})
    # Formato numérico inteiro para contadores na CAPA (separador de milhar, sem decimais)
    fmt_capa_int      = wb.add_format({"font_name":font,"font_size":size,"border":1,"border_color":"#D9E2F3","valign":"vcenter","align":"right","num_format":"#,##0"})
    fmt_capa_int_alt  = wb.add_format({"font_name":font,"font_size":size,"border":1,"border_color":"#D9E2F3","bg_color":"#F7F9FC","valign":"vcenter","align":"right","num_format":"#,##0"})
    # Contadores de alerta (> 0): destaque âmbar ou vermelho dependendo da coluna
    fmt_capa_int_warn = wb.add_format({"font_name":font,"font_size":size,"bold":True,"border":1,"border_color":"#FFD966","bg_color":"#FFF2CC","font_color":"#7F6000","valign":"vcenter","align":"right","num_format":"#,##0"})
    fmt_capa_int_crit = wb.add_format({"font_name":font,"font_size":size,"bold":True,"border":1,"border_color":"#F4B183","bg_color":"#FCE4D6","font_color":"#9C0006","valign":"vcenter","align":"right","num_format":"#,##0"})
    fmt_alert = wb.add_format({"font_name":font,"font_size":size,"bold":True,"border":1,"border_color":"#F4B183","bg_color":"#FFF2CC","font_color":"#7F6000","valign":"vcenter","align":"left"})
    fmt_icon = wb.add_format({"font_name":font,"font_size":size,"bold":True,"border":1,"border_color":"#BFBFBF","bg_color":"#F2F2F2","font_color":"#404040","valign":"vcenter","align":"center"})
    fmt_icon_crit = wb.add_format({"font_name":font,"font_size":size,"bold":True,"border":1,"border_color":"#C00000","bg_color":"#FCE4D6","font_color":"#9C0006","valign":"vcenter","align":"center"})
    fmt_icon_rev = wb.add_format({"font_name":font,"font_size":size,"bold":True,"border":1,"border_color":"#F4B183","bg_color":"#FFF2CC","font_color":"#C65911","valign":"vcenter","align":"center"})
    fmt_icon_warn = wb.add_format({"font_name":font,"font_size":size,"bold":True,"border":1,"border_color":"#FFD966","bg_color":"#FFF2CC","font_color":"#7F6000","valign":"vcenter","align":"center"})
    fmt_icon_ok = wb.add_format({"font_name":font,"font_size":size,"bold":True,"border":1,"border_color":"#A9D18E","bg_color":"#E2F0D9","font_color":"#375623","valign":"vcenter","align":"center"})
    fmt_ok = wb.add_format({"font_name":font,"font_size":size,"bold":True,"border":1,"border_color":"#A9D18E","bg_color":"#E2F0D9","font_color":"#375623","valign":"vcenter","align":"center"})
    fmt_gold6     = wb.add_format({"font_name":font,"font_size":size,"border":1,"border_color":"#D4A017","valign":"vcenter","align":"right","num_format":"R$ #,##0.000000","bg_color":"#FFFBEB","font_color":"#92600A","bold":True})
    fmt_gold6_alt = wb.add_format({"font_name":font,"font_size":size,"border":1,"border_color":"#D4A017","valign":"vcenter","align":"right","num_format":"R$ #,##0.000000","bg_color":"#FEF9E7","font_color":"#92600A","bold":True})
    fmt_link = wb.add_format({"font_name":font,"font_size":size,"border":1,"border_color":"#D9E2F3","font_color":"#0563C1","underline":1,"valign":"vcenter","align":"center"})
    # Link com alinhamento à esquerda - usado para LINK_ABRIR_AJUSTE e links na capa.
    fmt_link_left = wb.add_format({"font_name":font,"font_size":size,"border":1,"border_color":"#D9E2F3","font_color":"#0563C1","underline":1,"valign":"vcenter","align":"left"})
    # Alinhamento central: lista de colunas que devem ter texto centralizado.
    # Removido CHAVE_NFE desta lista para permitir alinhamento à esquerda conforme orientações de auditoria.
    center_cols = {"EMPRESA","MES","DATA_EMISSAO","CFOP","NCM","CEST","UM_ORIGINAL","SITUACAO_QTD","CALCULA","ALMOX","UNIDADE","EMPRESA_ORIGEM_PRECO","MES_ORIGEM_PRECO","SITUACAO_PRECO","SITUACAO","PERIODO","CODIGO_CONTROLE_ARQUIVO","ALERTA","SITUACAO_BASE","MOTIVO","FONTE_TIPO_ITEM","TIPO_PRODUTO","STATUS_CADASTRO_MATRIZ","Part_Arbitramento","CALCULA_MATRIZ","VINCULO_GRUPO","CATEGORIA_ITEM","CRITERIO_CUSTEIO_RIR70","PERFIL_EMPRESA","NCM_MATRIZ","INTERVENCAO_ANALISTA","AJUSTE_ANALISTA_APLICADO","TIPO_REFERENCIA","NIVEL","STATUS_AJUSTE","ALTERA_CALCULO"}

    ws.write_row(0, 0, columns, fmt_h)
    ws.set_row(0, 28)
    # Comentários discretos nos cabeçalhos: orientam o analista sobre o propósito das colunas.
    # Aba PENDENCIAS: colunas que o analista precisa preencher.
    if tab_kind == "pend":
        _pend_col_hints = {
            "TIPO_PENDENCIA":    "Código da pendência identificada pelo motor. Somente leitura.",
            "ALTERA_CALCULO":    "Indica se o ajuste correspondente impacta o valor calculado: Sim, Nao ou Indireto.",
            "INTERVENCAO_ANALISTA": "Sim = requer ação antes da próxima execução do motor.",
            "ACAO_ANALISTA":     "Instrução objetiva sobre o que fazer para resolver a pendência.",
            "DECISAO_ANALISTA":  "Preencha com a decisão tomada (ex.: aprovado, bloqueado, encaminhado).",
            "DATA_DECISAO":      "Data da decisão, formato DD/MM/AAAA.",
            "RESPONSAVEL_DECISAO": "Nome do responsável pela decisão fiscal.",
            "SITUACAO":          "Status atual da pendência: PENDENTE, RESOLVIDA ou DISPENSADA.",
        }
        for ci_h, col_h in enumerate(columns):
            if col_h in _pend_col_hints:
                ws.write_comment(0, ci_h, _pend_col_hints[col_h], {"x_scale": 1.5, "y_scale": 1.2})
    # Aba INVENTARIO_VALORIZADO: comentário no cabeçalho TIPO_PRODUTO.
    if tab_kind == "warn" and "TIPO_PRODUTO" in columns:
        tp_ci = columns.index("TIPO_PRODUTO")
        ws.write_comment(0, tp_ci,
            "Somente Produto Acabado e Mercadoria Revenda participam do cálculo RIR70.",
            {"x_scale": 2.0, "y_scale": 1.2})
    widths = [len(c) + 2 for c in columns]
    for ri, row in enumerate(rows, start=1):
        alt = (ri % 2 == 0)
        base_fmt = fmt_alt if alt else fmt_txt
        ctr_fmt = fmt_ctr_alt if alt else fmt_ctr
        date_fmt = fmt_date_alt if alt else fmt_date
        # CAPA: altura de linha explícita. Linhas com texto longo recebem altura maior para acomodar quebra.
        if tab_kind == "capa":
            indicador_norm = norm_text(str(row.get("INDICADOR", "")))
            if indicador_norm.startswith("PAINEL") or indicador_norm in {"CENTRAL DO ANALISTA"} or indicador_norm in {
                "CRITERIO APLICADO", "CRITERIO FISCAL ADOTADO",
                "ADVERTENCIA SOBRE PRECO DE OUTRA EMPRESA", "BASE LEGAL", "PROXIMO PASSO"
            }:
                ws.set_row(ri, 60)
            else:
                ws.set_row(ri, 17)
        for ci, col in enumerate(columns):
            v = row.get(col, "")
            # Para colunas específicas, normaliza textos usando apenas a primeira letra maiúscula. Para nomes
            # compostos por subpartes separadas por underscores (por exemplo, "RIR70_70_MAIOR_PRECO"), aplica
            # capitalização a cada parte, preservando os underscores. Isso melhora a legibilidade sem alterar
            # o conteúdo estrutural da coluna.
            if isinstance(v, str) and col in {"CRITERIO_CUSTEIO_RIR70", "INTERVENCAO_ANALISTA", "SITUACAO_PRECO"}:
                if v:
                    parts = v.split("_")
                    parts = [p[:1].upper() + p[1:].lower() if p else p for p in parts]
                    v = "_".join(parts)
            display = _fmt_value_for_excel(v)
            widths[ci] = min(max(widths[ci], len(clean_str(display)) + 3), 120)
            if (col.startswith("LINK_") or (tab_kind == "capa" and col == "VALOR" and norm_text(row.get("INDICADOR")).startswith("LINK"))) and clean_str(display):
                url = clean_str(display)
                texto_link = "Abrir Arquivo/Pasta"
                ind = norm_text(row.get("INDICADOR")) if tab_kind == "capa" else norm_text(col)
                if "AJUST" in ind:
                    texto_link = "Abrir Arquivo De Ajustes" if "PASTA" not in ind else "Abrir Pasta De Ajustes"
                elif "TEMPLATE" in ind:
                    texto_link = "Abrir Template De Ajustes"
                elif "AUXILIA" in ind:
                    texto_link = "Abrir Pasta De Auxiliares"
                elif "MOVIMENTO" in ind:
                    texto_link = "Abrir Pasta De Movimento"
                elif "INVENTARIO" in ind:
                    texto_link = "Abrir Pasta De Inventário"
                # Usa alinhamento à esquerda apenas para LINK_ABRIR_AJUSTE e links exibidos na capa
                link_fmt = fmt_link_left if col == "LINK_ABRIR_AJUSTE" or tab_kind == "capa" else fmt_link
                try:
                    ws.write_url(ri, ci, url, link_fmt, string=texto_link)
                except Exception:
                    ws.write(ri, ci, url, link_fmt)
            elif tab_kind == "capa" and ("CENTRAL DO ANALISTA" in norm_text(str(row.get("INDICADOR"))) or norm_text(str(row.get("INDICADOR"))).startswith("PAINEL")):
                # Destaque da linha Central do Analista. INDICADOR: fundo azul escuro.
                # VALOR: texto longo com quebra automática e alinhamento justificado.
                if col == "VALOR":
                    ws.write(ri, ci, display, fmt_capa_long_val)
                else:
                    ws.write(ri, ci, display, fmt_capa_sub_header)
            elif tab_kind == "capa" and col == "VALOR" and norm_text(row.get("INDICADOR")) == "SITUACAO FINAL":
                vn = norm_text(v)
                ws.write(ri, ci, display, fmt_capa_status_crit if "PENDENTE" in vn else fmt_capa_status_warn if "ADVERT" in vn else fmt_capa_status_ok)
            elif tab_kind == "capa" and col == "VALOR" and norm_text(row.get("INDICADOR")) == "STATUS_GERAL_ANALISTA":
                vn = norm_text(v)
                if vn == "INTERVENCAO NECESSARIA":
                    ws.write(ri, ci, display, fmt_capa_status_crit)
                else:
                    ws.write(ri, ci, display, fmt_capa_status_ok)
            elif tab_kind == "capa" and col == "VALOR" and "pendente de correcao" in norm_text(str(display)):
                ws.write(ri, ci, display, fmt_alert)
            elif tab_kind == "capa" and col == "VALOR" and isinstance(v, int) and norm_text(row.get("INDICADOR","")) in {
                    "TOTAL_COM_INTERVENCAO","TOTAL_CRITICO","EMPRESAS COM PENDENCIA","QUANTIDADE DE PENDENCIAS"}:
                # Contadores críticos: destacado em vermelho quando > 0
                fmt_n = fmt_capa_int_crit if v > 0 else (fmt_capa_int_alt if alt else fmt_capa_int)
                ws.write_number(ri, ci, v, fmt_n)
            elif tab_kind == "capa" and col == "VALOR" and isinstance(v, int) and norm_text(row.get("INDICADOR","")) in {
                    "TOTAL_REVISAR","TOTAL_INFORMATIVO","TOTAL DE VENDAS DESCARTADAS","LINHAS DUPLICADAS DESCARTADAS"}:
                # Contadores de revisão: destacado em âmbar quando > 0
                fmt_n = fmt_capa_int_warn if v > 0 else (fmt_capa_int_alt if alt else fmt_capa_int)
                ws.write_number(ri, ci, v, fmt_n)
            elif tab_kind == "capa" and col == "VALOR" and isinstance(v, int):
                # Outros contadores inteiros: #,##0 sem destaque
                ws.write_number(ri, ci, v, fmt_capa_int_alt if alt else fmt_capa_int)
            elif tab_kind == "capa" and col == "VALOR" and norm_text(row.get("INDICADOR")).startswith("EMPRESA"):
                # Alinha à esquerda e capitaliza apenas a primeira letra do texto das linhas Empresa nn.
                text = str(display)
                try:
                    text = text.capitalize()
                except Exception:
                    pass
                # Mantém formato condicional para pendências apenas para cor, mas com alinhamento à esquerda.
                if "PENDENTE" in norm_text(text):
                    ws.write(ri, ci, text, fmt_alert)
                else:
                    ws.write(ri, ci, text, base_fmt)
            elif tab_kind == "capa" and col == "VALOR" and norm_text(row.get("INDICADOR","")) in {
                    "CRITERIO APLICADO", "CRITERIO FISCAL ADOTADO",
                    "ADVERTENCIA SOBRE PRECO DE OUTRA EMPRESA", "BASE LEGAL", "PROXIMO PASSO", "PROXIMA ACAO DO OPERADOR", "PERIODO PROCESSADO"}:
                # Células da CAPA com texto longo: quebra automática + alinhamento justificado.
                ws.write(ri, ci, display, fmt_capa_long_val)
            elif col == "VARIACAO_PRECO (%)":
                # Grava como número com formato 0% - preserva ordenação, filtros e auditoria.
                # Célula vazia quando v é None ou <= 0.
                if v is None:
                    ws.write_blank(ri, ci, None, fmt_pct_alt if alt else fmt_pct)
                else:
                    try:
                        frac = float(Decimal(str(v)))
                    except Exception:
                        frac = 0.0
                    if frac <= 0:
                        ws.write_blank(ri, ci, None, fmt_pct_alt if alt else fmt_pct)
                    elif frac > 0.5:
                        ws.write_number(ri, ci, frac, fmt_pct_alert)
                    else:
                        ws.write_number(ri, ci, frac, fmt_pct_alt if alt else fmt_pct)
            elif col == "ORIGEM_DO_PRECO":
                # Converte para texto legível: remove underscores, primeira letra maiúscula.
                text = str(display) if display is not None else ""
                try:
                    text = text.replace("_", " ").lower().capitalize()
                except Exception:
                    pass
                # Para o caso especial SEM_PRECO_LOCALIZADO, aplica formatação crítica com ícone.
                if clean_str(v) == "SEM_PRECO_LOCALIZADO":
                    ws.write(ri, ci, text, fmt_icon_crit)
                else:
                    # Usa formatação padrão (texto ou alternado) para campos de origem do preço.
                    fmt_o = fmt_alt if alt else fmt_txt
                    ws.write(ri, ci, text, fmt_o)
            elif col == "ALERTA" and clean_str(v):
                val_norm = norm_text(v)
                if "CRITICO" in val_norm:
                    ws.write(ri, ci, display, fmt_icon_crit)
                elif "REVISAR" in val_norm:
                    ws.write(ri, ci, display, fmt_icon_rev)
                elif "ATENCAO" in val_norm or "ALERTA" in val_norm:
                    ws.write(ri, ci, display, fmt_icon_warn)
                elif "OK" in val_norm:
                    ws.write(ri, ci, display, fmt_icon_ok)
                else:
                    ws.write(ri, ci, display, fmt_icon)
            elif col == "NIVEL" and clean_str(v):
                val_norm = norm_text(v)
                if "CRITICO" in val_norm:
                    ws.write(ri, ci, display, fmt_icon_crit)
                elif "REVISAR" in val_norm:
                    ws.write(ri, ci, display, fmt_icon_rev)
                elif "INFORMATIVO" in val_norm or "INFO" in val_norm:
                    ws.write(ri, ci, display, fmt_icon)
                else:
                    ws.write(ri, ci, display, fmt_icon_warn)
            elif col == "ADVERTENCIA" and clean_str(v):
                # Exibe advertência com apenas a primeira letra maiúscula e alerta visual.
                text = str(display)
                try:
                    text = text.capitalize()
                except Exception:
                    pass
                ws.write(ri, ci, text, fmt_alert)
            elif col in {"ABA_AJUSTE", "CAMPO_AJUSTE", "TIPO_AJUSTE_SUGERIDO", "TIPO_PENDENCIA", "MOTIVO_PRINCIPAL_DESCARTE"}:
                # Exibe apenas a primeira letra maiúscula para campos de pendência.
                text = str(display)
                try:
                    text = text.capitalize()
                except Exception:
                    pass
                # Mantém alinhamento à esquerda.
                ws.write(ri, ci, text, base_fmt)
            elif col == "MOTIVO":
                # Motivo com apenas a primeira letra maiúscula, alinhado à esquerda.
                text = str(display)
                try:
                    text = text.capitalize()
                except Exception:
                    pass
                ws.write(ri, ci, text, base_fmt)
            elif col in {"INTERVENCAO_ANALISTA", "ACAO_ANALISTA"}:
                # INTERVENCAO_ANALISTA: centralizado conforme orientação de layout.
                # ACAO_ANALISTA: alinhamento à esquerda (texto descritivo).
                # Ambos destacam em âmbar quando valor = SIM.
                text = str(display)
                if col == "INTERVENCAO_ANALISTA":
                    ws.write(ri, ci, text, fmt_alert if norm_text(v) == "SIM" else ctr_fmt)
                else:
                    ws.write(ri, ci, text, fmt_alert if norm_text(v) == "SIM" else base_fmt)
            elif col in {"SITUACAO_BASE", "SITUACAO_PRECO", "SITUACAO", "CALCULA"} and clean_str(v) in {"APROVADA", "OK", "SIM"}:
                ws.write(ri, ci, display, fmt_ok)
            elif isinstance(v, datetime):
                ws.write_datetime(ri, ci, v, date_fmt)
            elif isinstance(v, Decimal):
                if col == "CUSTO_ARBITRADO_70":
                    ws.write_number(ri, ci, float(v), fmt_gold6_alt if alt else fmt_gold6)
                elif col in {"MAIOR_PRECO_UNITARIO","PRECO_MEDIO","VALOR_UNITARIO"}:
                    ws.write_number(ri, ci, float(v), fmt_num6_alt if alt else fmt_num6)
                elif "QTD" in col or "SALDO" in col:
                    ws.write_number(ri, ci, float(v), fmt_num4_alt if alt else fmt_num4)
                else:
                    ws.write_number(ri, ci, float(v), fmt_num2_alt if alt else fmt_num2)
            else:
                ws.write(ri, ci, display, ctr_fmt if col in center_cols else base_fmt)
    # Descompressão visual das colunas: largura pelo conteúdo, com mínimos específicos por coluna crítica.
    for ci, col in enumerate(columns):
        w = max(10, widths[ci])
        if tab_kind == "capa" and col == "VALOR":
            w = max(w, 85)
        if col in {"CHAVE_NFE", "CODIGO_CONTROLE_ARQUIVO"}:
            w = max(w, 58)
        if col in {"DESCRICAO_PRODUTO", "DESCRICAO_TECNICA", "DESCRICAO_PENDENCIA", "ACAO_NECESSARIA", "ACAO_ANALISTA", "MOTIVO_PRINCIPAL_DESCARTE", "ADVERTENCIA"}:
            w = max(w, 34)
        if col in {"ALERTA", "NIVEL"}:
            w = 14
        if col in {"FONTE_TIPO_ITEM", "ORIGEM_DO_PRECO", "ARQUIVO_AJUSTE", "TIPO_AJUSTE_SUGERIDO"}:
            w = max(w, 28)
        if col == "VARIACAO_PRECO (%)":
            w = max(w, 18)
        if col in {"EMPRESA", "PERIODO"}:
            w = max(w, 18)
        ws.set_column(ci, ci, min(w, 120))
    ws.freeze_panes(1, 0)
    ws.autofilter(0, 0, max(1, len(rows)), max(0, len(columns)-1))
    if rows:
        last = len(rows)
        # Realces executivos sem poluir a planilha.
        for col_name in ["SITUACAO_BASE", "SITUACAO_PRECO", "SITUACAO", "CALCULA"]:
            if col_name in columns:
                idx = columns.index(col_name)
                ws.conditional_format(1, idx, last, idx, {"type":"text","criteria":"containing","value":"REVISAR","format":wb.add_format({"bg_color":"#FFF2CC","font_color":"#7F6000","font_name":font,"font_size":size,"bold":True,"align":"center"})})
                ws.conditional_format(1, idx, last, idx, {"type":"text","criteria":"containing","value":"BLOQUEADA","format":wb.add_format({"bg_color":"#F4CCCC","font_color":"#990000","font_name":font,"font_size":size,"bold":True,"align":"center"})})
                ws.conditional_format(1, idx, last, idx, {"type":"text","criteria":"containing","value":"NAO","format":wb.add_format({"bg_color":"#F4CCCC","font_color":"#990000","font_name":font,"font_size":size,"bold":True,"align":"center"})})
                ws.conditional_format(1, idx, last, idx, {"type":"text","criteria":"containing","value":"APROVADA","format":wb.add_format({"bg_color":"#E2F0D9","font_color":"#375623","font_name":font,"font_size":size,"bold":True,"align":"center"})})
        if "ADVERTENCIA" in columns:
            idx = columns.index("ADVERTENCIA")
            ws.conditional_format(1, idx, last, idx, {"type":"no_blanks","format":fmt_alert})
        if "ALERTA" in columns:
            idx = columns.index("ALERTA")
            ws.conditional_format(1, idx, last, idx, {"type":"text","criteria":"containing","value":"CRITICO","format":fmt_icon_crit})
            ws.conditional_format(1, idx, last, idx, {"type":"text","criteria":"containing","value":"REVISAR","format":fmt_icon_rev})
            ws.conditional_format(1, idx, last, idx, {"type":"text","criteria":"containing","value":"ATENCAO","format":fmt_icon_warn})
            ws.conditional_format(1, idx, last, idx, {"type":"text","criteria":"containing","value":"OK","format":fmt_icon_ok})
        if "NIVEL" in columns:
            idx = columns.index("NIVEL")
            ws.conditional_format(1, idx, last, idx, {"type":"text","criteria":"containing","value":"CRITICO","format":fmt_icon_crit})
            ws.conditional_format(1, idx, last, idx, {"type":"text","criteria":"containing","value":"REVISAR","format":fmt_icon_rev})


def _empresa_status_rows(records, inv_rows, pend_rows, config):
    emp_info = {}
    for item in config.get("empresas_grupo", []) or []:
        if isinstance(item, dict):
            c = normalize_cnpj(item.get("cnpj")); raz = clean_str(item.get("razao_social")); filial = clean_str(item.get("filial")); uf = clean_str(item.get("uf"))
        else:
            c = normalize_cnpj(item); raz = filial = uf = ""
        if c:
            emp_info[c] = {"CNPJ": c, "RAZAO": raz, "FILIAL": filial, "UF": uf}
    for c in _empresas_grupo_cnpjs(config):
        emp_info.setdefault(c, {"CNPJ": c, "RAZAO": "", "FILIAL": "", "UF": ""})
    emp_mov = {normalize_cnpj(r.get("CNPJ Emitente")) for r in records if r.get("CNPJ Emitente")}
    emp_inv = {normalize_cnpj(r.get("EMPRESA")) for r in inv_rows if r.get("EMPRESA")}
    pend_by_emp = defaultdict(int)
    for p in pend_rows:
        if clean_str(p.get("SITUACAO")).upper() == "PENDENTE":
            pend_by_emp[normalize_cnpj(p.get("EMPRESA"))] += 1
    sem_mov_conf = {normalize_cnpj(x) for x in config.get("empresas_sem_movimento_confirmadas", []) or []}
    sem_est_conf = {normalize_cnpj(x) for x in config.get("empresas_sem_estoque_confirmadas", []) or []}
    out = []
    for c, info in sorted(emp_info.items()):
        mov = "OK" if c in emp_mov else ("SEM_MOVIMENTO_CONFIRMADO" if c in sem_mov_conf else "PENDENTE_MOVIMENTO")
        inv = "OK" if c in emp_inv else ("SEM_ESTOQUE_CONFIRMADO" if c in sem_est_conf else "PENDENTE_INVENTARIO")
        pend = pend_by_emp.get(c, 0)
        geral = "OK" if not pend and "PENDENTE" not in mov and "PENDENTE" not in inv else "PENDENTE"
        nome = " - ".join([x for x in [c, info.get("RAZAO"), ("FILIAL " + info.get("FILIAL")) if info.get("FILIAL") else "", info.get("UF")] if x])
        out.append({
            "EMPRESA": c,
            "STATUS_GERAL": geral,
            "DESCRICAO_STATUS": f"{nome} | MOVIMENTO: {mov} | INVENTARIO: {inv} | PENDENCIAS: {pend} | STATUS: {geral}",
        })
    return out

def _normalizar_linhas_para_colunas(rows, columns):
    normalized = []
    for r in rows:
        normalized.append({c: r.get(c, "") for c in columns})
    return normalized

def _expected_headers_map(config):
    return {
        "CAPA": ["INDICADOR", "VALOR"],
        "01_MOVIMENTO_VALIDADO": _cols_operacionais(MOV_COLS_ANALISTA, MOV_COLS_DIAGNOSTICO, config),
        "02_ARBITRAMENTO": _cols_operacionais(ARB_COLS_ANALISTA, ARB_COLS_DIAGNOSTICO, config),
        "03_INVENTARIO_VALORIZADO": _cols_operacionais(INV_COLS_ANALISTA, INV_COLS_DIAGNOSTICO, config),
        "04_PENDENCIAS": PEND_COLS_FINAL,
        "05_FONTES_PROCESSADAS": FONTES_COLS_FINAL,
    }

def write_workbook(records, logs, config):
    global RIR70_LAST_INVENTORY_ROWS, RIR70_LAST_ARB_ROWS
    output_path = _novo_output_path(config)
    mov_rows = build_movimentacao_rows(records)
    arb_rows = build_arbitramento_rows(records, config)
    arb_rows = _apply_operator_adjustments_to_arbitramento(arb_rows, config)
    RIR70_LAST_ARB_ROWS = arb_rows
    inventory_map, inv_logs = load_inventory_pa(config, logs)
    inv_rows = build_inventario_rows(arb_rows, inventory_map, config)
    inv_rows = _apply_operator_adjustments_to_inventario(inv_rows, config)
    RIR70_LAST_INVENTORY_ROWS = inv_rows
    pend_rows = _build_pendencias(records, arb_rows, inv_rows, config)
    # Gera/atualiza o arquivo único de ajustes antes da normalização das colunas.
    template_ajustes_path, ajustes_analista_path = _gerar_arquivos_ajustes_analista(pend_rows, config)
    fontes_rows = _build_fontes_processadas(logs, records, config)
    expected_headers = _expected_headers_map(config)
    mov_rows = _normalizar_linhas_para_colunas(mov_rows, expected_headers["01_MOVIMENTO_VALIDADO"])
    arb_rows = _normalizar_linhas_para_colunas(arb_rows, expected_headers["02_ARBITRAMENTO"])
    inv_rows = _normalizar_linhas_para_colunas(inv_rows, expected_headers["03_INVENTARIO_VALORIZADO"])
    pend_rows = _normalizar_linhas_para_colunas(pend_rows, expected_headers["04_PENDENCIAS"])
    fontes_rows = _normalizar_linhas_para_colunas(fontes_rows, expected_headers["05_FONTES_PROCESSADAS"])

    inicio = parse_config_date(config.get("periodo_base_inicio")); fim = parse_config_date(config.get("periodo_base_fim"))
    meses = _month_range(inicio, fim)
    situ_final = "PENDENTE DE CORRECAO" if pend_rows else ("OK COM ADVERTENCIAS" if any(clean_str(r.get("ADVERTENCIA")) for r in arb_rows + inv_rows) else "OK")
    empresa_status = _empresa_status_rows(records, inv_rows, pend_rows, config)
    qtd_pendentes = sum(1 for p in pend_rows if clean_str(p.get("SITUACAO")).upper() == "PENDENTE")
    total_intervencao = sum(1 for p in pend_rows if clean_str(p.get("INTERVENCAO_ANALISTA")) == "SIM")
    total_critico = sum(1 for p in pend_rows if clean_str(p.get("NIVEL")) == "CRITICO")
    total_revisar = sum(1 for p in pend_rows if clean_str(p.get("NIVEL")) == "REVISAR")
    total_atencao = sum(1 for p in pend_rows if clean_str(p.get("NIVEL")) == "ATENCAO")
    total_acionavel = total_critico + total_revisar + total_atencao
    total_informativo = sum(1 for r in mov_rows if clean_str(r.get("INTERVENCAO_ANALISTA")) == "INFORMATIVO")
    situ_final = "PENDENTE DE CORRECAO" if qtd_pendentes else ("OK COM ADVERTENCIAS" if any(clean_str(r.get("ADVERTENCIA")) for r in arb_rows + inv_rows) else "OK")
    # Mensagem de próximo passo: orienta intervenção do analista quando há pendências ou confirma ausência de ações necessárias
    prox_passo = (
        "Existem itens que exigem intervenção. Clique em LINK_ABRIR_AJUSTES, preencha as células amarelas do arquivo operacional e execute novamente o motor." if total_intervencao
        else "Não há intervenção pendente. Arquivo apto para revisão final."
    )
    # Instruções para o analista conforme modelo corporativo: enumera passo a passo de forma clara. O texto
    # explicita o link a ser clicado, as colunas que devem ser editadas no arquivo operacional e reforça que o
    # template no output é somente para auditoria. Ao final, alerta que o analista não deve editar as abas de
    # saída. Para manter aderência ao layout exemplo, o texto contém travessões longos e conjuga "PASSO" com
    # numeral, tudo em português.
    instrucao_analista = (
        "PASSO 1 - Clique em LINK_ABRIR_AJUSTES e preencha somente o arquivo operacional input/ajustes/AJUSTES_ANALISTA_RIR70.xlsx. "
        "PASSO 2 - Preencha as células amarelas: TIPO_AJUSTE, VALOR_AJUSTADO ou VALOR_REFERENCIA_INFORMADO quando aplicável, JUSTIFICATIVA, FONTE_COMPROVACAO, CAMINHO_EVIDENCIA, RESPONSAVEL, DATA_DECISAO e STATUS_AJUSTE = APROVADO. "
        "PASSO 3 - Salve o próprio arquivo operacional e execute o motor novamente. "
        "IMPORTANTE: o TEMPLATE no output é somente cópia auditável; se preencher e apertar Ctrl+S nele, o motor não lerá o ajuste. Não edite as abas de output."
    )
    capa_rows = [
        {"INDICADOR":"PAINEL DO SISTEMA", "VALOR":"Motor de Arbitramento RIR70 - rotina operacional guiada"},
        {"INDICADOR":"STATUS_GERAL_ANALISTA", "VALOR":"INTERVENCAO NECESSARIA" if total_intervencao else "SEM INTERVENCAO PENDENTE"},
        {"INDICADOR":"PERIODO_PROCESSADO", "VALOR":f"{format_date_abnt(inicio)} ate {format_date_abnt(fim)}"},
        {"INDICADOR":"PROXIMA_ACAO_DO_OPERADOR", "VALOR":prox_passo},
        {"INDICADOR":"PAINEL DE LINKS", "VALOR":"Use os links abaixo para abrir ajustes, evidencias, output, log e metadados."},
        {"INDICADOR":"Central do Analista", "VALOR": instrucao_analista},
        {"INDICADOR":"TOTAL_COM_INTERVENCAO", "VALOR":total_intervencao},
        {"INDICADOR":"TOTAL_CRITICO", "VALOR":total_critico},
        {"INDICADOR":"TOTAL_REVISAR", "VALOR":total_revisar},
        {"INDICADOR":"TOTAL_ATENCAO_COMPATIBILIDADE", "VALOR":total_atencao},
        {"INDICADOR":"TOTAL_ACIONAVEL", "VALOR":total_acionavel},
        {"INDICADOR":"TOTAL_INFORMATIVO", "VALOR":total_informativo},
        {"INDICADOR":"ARQUIVO_DE_AJUSTES", "VALOR":clean_str(config.get("_arquivo_ajustes_analista"))},
        {"INDICADOR":"LINK_ABRIR_AJUSTES", "VALOR":_file_uri(config.get("_arquivo_ajustes_analista") or _ajustes_file_path(config))},
        {"INDICADOR":"LINK_ABRIR_PASTA_AJUSTES", "VALOR":_file_uri(_ajustes_folder(config))},
        {"INDICADOR":"LINK_ABRIR_TEMPLATE_AJUSTES_COPIA_NAO_PREENCHER", "VALOR":_file_uri(config.get("_template_ajustes_analista") or _template_ajustes_path(config))},
        {"INDICADOR":"LINK_ABRIR_PASTA_AUXILIARES", "VALOR":_file_uri(BASE_DIR / "input" / "auxiliares")},
        {"INDICADOR":"LINK_ABRIR_PASTA_MOVIMENTO", "VALOR":_file_uri(BASE_DIR / "input" / "movimento_item")},
        {"INDICADOR":"LINK_ABRIR_PASTA_INVENTARIO", "VALOR":_file_uri(BASE_DIR / "input" / "inventario")},
        {"INDICADOR":"PROXIMO_PASSO", "VALOR":prox_passo},
        {"INDICADOR":"Nome do motor", "VALOR":"Motor de Arbitramento RIR70"},
        {"INDICADOR":"Data e hora do processamento", "VALOR":datetime.now().strftime("%d/%m/%Y %H:%M:%S")},
        {"INDICADOR":"Versao do comando", "VALOR":SCRIPT_VER},
        {"INDICADOR":"Versao do script", "VALOR":SCRIPT_VER},
        {"INDICADOR":"MODO_EXECUCAO", "VALOR":config.get("MODO_EXECUCAO")},
        {"INDICADOR":"MODO_GOLDEN_FILE", "VALOR":config.get("MODO_GOLDEN_FILE")},
        {"INDICADOR":"GOVERNANCA_OPERACIONAL", "VALOR":config.get("_governanca_operacional")},
        {"INDICADOR":"REGRA_MOTOR_VFINAL", "VALOR":config.get("regra_motor_vfinal")},
        {"INDICADOR":"PERIODO_ORIGEM", "VALOR":config.get("periodo_definido_interativamente")},
        {"INDICADOR":"LINK_ABRIR_OUTPUT_FINAL", "VALOR":_file_uri(output_path)},
        {"INDICADOR":"LINK_ABRIR_LOG_PROCESSAMENTO", "VALOR":_file_uri(LOG_PATH)},
        {"INDICADOR":"LINK_ABRIR_METADADOS", "VALOR":_file_uri(META_PATH)},
        {"INDICADOR":"exibir_colunas_diagnostico", "VALOR":config.get("exibir_colunas_diagnostico")},
        {"INDICADOR":"ADVERTENCIA_MODO_EXECUCAO", "VALOR":config.get("ADVERTENCIA_MODO_EXECUCAO")},
        {"INDICADOR":"Matriz campos obrigatorios", "VALOR":"OK" if (config.get("_matriz_campos_obrigatorios_status") or {}).get("ok") else "Pendente De Correcao"},
        {"INDICADOR":"Matriz pendencias estruturais", "VALOR":"; ".join((config.get("_matriz_campos_obrigatorios_status") or {}).get("missing") or [])},
        {"INDICADOR":"Período informado", "VALOR":f"{format_date_abnt(inicio)} até {format_date_abnt(fim)}"},
        {"INDICADOR":"Forma de cálculo", "VALOR":"mensal, por empresa e produto"},
        {"INDICADOR":"Meses processados", "VALOR":"; ".join(meses)},
        {"INDICADOR":"Situação final", "VALOR":situ_final},
        {"INDICADOR":"Empresas com pendência", "VALOR":sum(1 for e in empresa_status if "PENDENTE" in norm_text(e.get("STATUS_GERAL")))},
        {"INDICADOR":"Total de vendas lidas", "VALOR":len(records)},
        {"INDICADOR":"Total de vendas usadas no cálculo", "VALOR":sum(1 for r in records if r.get('_participa'))},
        {"INDICADOR":"Total de vendas descartadas", "VALOR":sum(1 for r in records if not r.get('_participa'))},
        {"INDICADOR":"Linhas duplicadas descartadas", "VALOR":len(RIR70_DUPLICATES)},
        {"INDICADOR":"Produtos com preço calculado", "VALOR":len(arb_rows)},
        {"INDICADOR":"Produtos sem preço localizado", "VALOR":sum(1 for r in inv_rows if r.get('ORIGEM_DO_PRECO') == 'SEM_PRECO_LOCALIZADO')},
        {"INDICADOR":"Produtos com preço usado de mês anterior", "VALOR":sum(1 for r in inv_rows if r.get('ORIGEM_DO_PRECO') == 'PRECO_DO_PROPRIO_PRODUTO_EM_MES_ANTERIOR')},
        {"INDICADOR":"Produtos com preço usado de outra empresa", "VALOR":sum(1 for r in inv_rows if r.get('ORIGEM_DO_PRECO') == 'PRECO_USADO_DE_OUTRA_EMPRESA')},
        {"INDICADOR":"Produtos com preço usado de produto similar", "VALOR":sum(1 for r in inv_rows if r.get('ORIGEM_DO_PRECO') == 'PRECO_USADO_DE_PRODUTO_SIMILAR')},
        {"INDICADOR":"Produtos com preço de referência manual", "VALOR":sum(1 for r in inv_rows if r.get('ORIGEM_DO_PRECO') == 'PRECO_REFERENCIA_MANUAL')},
        {"INDICADOR":"Quantidade de pendências", "VALOR":qtd_pendentes},
    ]
    for idx, e in enumerate(empresa_status, start=1):
        capa_rows.append({"INDICADOR":f"Empresa {idx:02d}", "VALOR":e.get("DESCRICAO_STATUS")})
    capa_rows.extend([
        {"INDICADOR":"Base legal", "VALOR":"RIR/2018, art. 308; Decreto-Lei nº 1.598/1977, art. 14, par.3º; PN CST nº 14/1981."},
        {"INDICADOR":"Critério aplicado", "VALOR":"Custo arbitrado por CRITERIO_CUSTEIO_RIR70: PA e MR vinculado ao grupo usam 70% do maior preço de venda; MR de terceiros exige CUSTO_AQUISICAO/rotina própria e não deve ser convertido cegamente em 70% do maior preço."},
        {"INDICADOR":"Critério fiscal adotado", "VALOR":"Preço de venda bruto, com ICMS incluído. Critério conservador adotado pela política da empresa para este motor. Revisível mediante orientação jurídica formal."},
        {"INDICADOR":"Advertência sobre preço de outra empresa", "VALOR":"Preços usados de outra empresa do grupo exigem verificação de comparabilidade comercial."},
    ])

    with pd.ExcelWriter(output_path, engine="xlsxwriter", datetime_format="dd/mm/yyyy", date_format="dd/mm/yyyy") as writer:
        wb = writer.book
        wb.set_properties({"title":"Motor de Arbitramento RIR70", "subject":"RIR70", "author":clean_str(config.get("autor_excel")) or "101 BRASIL INDUSTRIA DE BEBIDAS LTDA"})
        sheets = [
            ("CAPA", expected_headers["CAPA"], capa_rows, "capa"),
            ("01_MOVIMENTO_VALIDADO", expected_headers["01_MOVIMENTO_VALIDADO"], mov_rows, "data"),
            ("02_ARBITRAMENTO", expected_headers["02_ARBITRAMENTO"], arb_rows, "ok"),
            ("03_INVENTARIO_VALORIZADO", expected_headers["03_INVENTARIO_VALORIZADO"], inv_rows, "warn"),
            ("04_PENDENCIAS", expected_headers["04_PENDENCIAS"], pend_rows, "pend"),
            ("05_FONTES_PROCESSADAS", expected_headers["05_FONTES_PROCESSADAS"], fontes_rows, "src"),
        ]
        tab_colors = {
            "CAPA":                    "#0D2137",
            "01_MOVIMENTO_VALIDADO":   "#1A3A5C",
            "02_ARBITRAMENTO":         "#1E5C8A",
            "03_INVENTARIO_VALORIZADO":"#14532D",
            "04_PENDENCIAS":           "#7F1D1D",
            "05_FONTES_PROCESSADAS":   "#4C1D95",
        }
        for name, cols, rows, kind in sheets:
            ws = wb.add_worksheet(name); writer.sheets[name] = ws
            tc = tab_colors.get(name)
            if tc:
                ws.set_tab_color(tc)
            ws.hide_gridlines(2)
            _write_final_sheet(wb, ws, cols, rows, config, kind)
    validate_generated_xlsx_package(output_path, expected_headers=expected_headers, config=config)
    print(f"OK Excel gerado e validado estruturalmente: {output_path}")

def _cell_text_xlsx(cell, shared_strings):
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    t = cell.attrib.get("t", "")
    if t == "inlineStr":
        return "".join(x.text or "" for x in cell.findall(f".//{ns}t"))
    v = cell.find(f"{ns}v")
    raw = v.text if v is not None and v.text is not None else ""
    if t == "s":
        try:
            idx = int(raw)
            return shared_strings[idx] if 0 <= idx < len(shared_strings) else ""
        except Exception:
            return raw
    return raw

def _read_sheet_header_and_errors(zf, sheet_path, shared_strings, max_cells=None):
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    header = []
    errors = []
    cells_seen = 0
    error_literals = {"#NOME?", "#NAME?", "#VALOR!", "#VALUE!", "#REF!", "#DIV/0!", "#N/D", "#N/A", "#NULL!", "#NUM!"}
    with zf.open(sheet_path) as fh:
        for event, elem in ET.iterparse(fh, events=("end",)):
            if elem.tag != ns + "row":
                continue
            row_ref = elem.attrib.get("r", "")
            values = []
            for c in elem.findall(ns + "c"):
                cells_seen += 1
                value = _cell_text_xlsx(c, shared_strings)
                if row_ref == "1":
                    col_idx = _xlsx_col_index(c.attrib.get("r", "A1"))
                    while len(values) <= col_idx:
                        values.append("")
                    values[col_idx] = value
                if c.attrib.get("t") == "e" or value in error_literals:
                    errors.append(f"{sheet_path}:{c.attrib.get('r','?')}={value}")
                    if len(errors) >= 20:
                        return values if row_ref == "1" else header, errors, cells_seen
                if max_cells and cells_seen >= max_cells:
                    break
            if row_ref == "1":
                header = [clean_str(v) for v in values]
            elem.clear()
            if max_cells and cells_seen >= max_cells:
                break
    return header, errors, cells_seen

def validate_generated_xlsx_package(path, expected_headers=None, config=None):
    required = ["CAPA", "01_MOVIMENTO_VALIDADO", "02_ARBITRAMENTO", "03_INVENTARIO_VALORIZADO", "04_PENDENCIAS", "05_FONTES_PROCESSADAS"]
    expected_headers = expected_headers or {}
    config = config or {}
    with zipfile.ZipFile(path, "r") as zf:
        wbxml = ET.fromstring(zf.read("xl/workbook.xml"))
        relxml = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rels = {r.attrib.get("Id"): r.attrib.get("Target") for r in relxml}
        ns = {"a":"http://schemas.openxmlformats.org/spreadsheetml/2006/main", "r":"http://schemas.openxmlformats.org/officeDocument/2006/relationships"}
        sheet_to_path = {}
        names = []
        for s in wbxml.findall(".//a:sheet", ns):
            name = s.attrib.get("name")
            rid = s.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            target = rels.get(rid, "")
            path_rel = target.lstrip("/")
            if path_rel and not path_rel.startswith("xl/"):
                path_rel = "xl/" + path_rel
            if path_rel not in zf.namelist() and path_rel:
                path_rel = "xl/worksheets/" + Path(path_rel).name
            names.append(name)
            sheet_to_path[name] = path_rel
        if names != required:
            raise ValueError(f"Abas fora do padrão. Esperado={required}; Gerado={names}")
        shared_strings = _xlsx_shared_strings(zf)
        try:
            max_cells = int(config.get("max_celulas_validacao_output") or 0) or None
        except Exception:
            max_cells = None
        total_seen = 0
        for sheet in required:
            sh_path = sheet_to_path.get(sheet)
            if not sh_path or sh_path not in zf.namelist():
                raise ValueError(f"Aba obrigatória sem XML interno: {sheet}")
            limit_remain = None if max_cells is None else max(1, max_cells - total_seen)
            header, errors, seen = _read_sheet_header_and_errors(zf, sh_path, shared_strings, limit_remain)
            total_seen += seen
            expected = expected_headers.get(sheet)
            if expected and header[:len(expected)] != expected:
                raise ValueError(f"Cabeçalho inválido em {sheet}. Esperado={expected}; Gerado={header[:len(expected)]}")
            if _is_sim(config.get("validar_erros_excel_output", "Sim")) and errors:
                raise ValueError("Erros de Excel detectados no output: " + "; ".join(errors[:20]))
            if max_cells is not None and total_seen >= max_cells:
                logger.warning("Validação do output interrompida por max_celulas_validacao_output=%s", max_cells)
                break
    return True

def write_metadata(records, config):
    meta = {
        "versao_script": SCRIPT_VER,
        "versao_comando": SCRIPT_VER,
        "modo_execucao": config.get("MODO_EXECUCAO"),
        "modo_golden_file": config.get("MODO_GOLDEN_FILE"),
        "arquivo_xlsx": str(OUTPUT_PATH),
        "arquivo_log": str(LOG_PATH),
        "regra_motor_vfinal": config.get("regra_motor_vfinal"),
        "periodo_definido_interativamente": config.get("periodo_definido_interativamente"),
        "sha256_xlsx": sha256_file(OUTPUT_PATH),
        "data_hora_execucao": datetime.now().isoformat(timespec="seconds"),
        "periodo_inicio": config.get("periodo_base_inicio"),
        "periodo_fim": config.get("periodo_base_fim"),
        "criterio": "RIR/2018, art. 308: 70% do maior preço de venda mensal de produto acabado, preço bruto com ICMS incluído; preço médio de alerta ponderado por QTD_CALCULO.",
        "abas_obrigatorias": ["CAPA", "01_MOVIMENTO_VALIDADO", "02_ARBITRAMENTO", "03_INVENTARIO_VALORIZADO", "04_PENDENCIAS", "05_FONTES_PROCESSADAS"],
        "linhas_movimento": len(records),
        "linhas_participantes": sum(1 for r in records if r.get("_participa")),
        "duplicidades_descartadas": len(RIR70_DUPLICATES),
    }
    META_PATH.write_text(safe_json_dumps(meta, indent=2), encoding="utf-8")



# =============================================================================
# OTIMIZACAO DE LEITURA XLSX - colunas selecionadas somente
# =============================================================================

def _fast_xlsx_rows_selected(file_path, sheet_name, selected_idxs, max_rows=None):
    selected_idxs = {int(i) for i in selected_idxs if i is not None and int(i) >= 0}
    max_idx = max(selected_idxs) if selected_idxs else 0
    sheet_map = dict(_xlsx_sheet_paths(file_path))
    sheet_rel = sheet_map.get(sheet_name)
    if not sheet_rel:
        return
    strings = _xlsx_shared_strings_cached(file_path)
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    with zipfile.ZipFile(file_path) as zf:
        with zf.open(sheet_rel) as fh:
            context = ET.iterparse(fh, events=("end",))
            yielded = 0
            for event, elem in context:
                if elem.tag != ns + "row":
                    continue
                row_vals = [""] * (max_idx + 1)
                has_val = False
                for c in elem.findall(ns + "c"):
                    ref = c.attrib.get("r", "")
                    m = re.match(r"([A-Z]+)", ref)
                    if not m:
                        continue
                    idx = _xlsx_col_index(m.group(1))
                    if idx not in selected_idxs:
                        continue
                    t = c.attrib.get("t")
                    v = c.find(ns + "v")
                    val = ""
                    if t == "inlineStr":
                        is_el = c.find(ns + "is")
                        if is_el is not None:
                            texts = [tx.text or "" for tx in is_el.iter(ns + "t")]
                            val = "".join(texts)
                    elif v is not None and v.text is not None:
                        raw = v.text
                        if t == "s":
                            try:
                                val = strings[int(raw)]
                            except Exception:
                                val = raw
                        else:
                            val = raw
                    if idx <= max_idx:
                        row_vals[idx] = val
                        if clean_str(val):
                            has_val = True
                if has_val:
                    yield row_vals
                    yielded += 1
                    if max_rows and yielded >= max_rows:
                        elem.clear(); break
                elem.clear()


def _file_date_range_from_name(file_name):
    txt = clean_str(file_name)
    m = re.search(r"de\s+(\d{2})[-_/](\d{2})[-_/](\d{4})\s+a\s+(\d{2})[-_/](\d{2})[-_/](\d{4})", txt, flags=re.I)
    if not m:
        return None, None
    d1, m1, y1, d2, m2, y2 = m.groups()
    try:
        return datetime(int(y1), int(m1), int(d1)), datetime(int(y2), int(m2), int(d2), 23, 59, 59)
    except Exception:
        return None, None

def _file_overlaps_period(path, config):
    inicio = parse_config_date(config.get("periodo_base_inicio"))
    fim = parse_config_date(config.get("periodo_base_fim"))
    if fim:
        fim = fim.replace(hour=23, minute=59, second=59)
    a, b = _file_date_range_from_name(getattr(path, "name", path))
    if not a or not b or not inicio or not fim:
        return True
    return not (b < inicio or a > fim)

def discover_movement_files(config):
    files = []
    if config.get("usar_subpastas_input", True):
        files.extend([p for p in _excel_files_in_dir(config.get("pasta_movimentacao", "input/movimento_item")) if not _is_excluded_movement_file(p, config)])
    if config.get("permitir_movimentacao_em_input_raiz", True):
        inv_pref = config.get("arquivos_inventario_prefixos", ["INVENTARIO", "ESTOQUE", "SALDO"])
        for p in discover_files():
            if _is_manual_factor_file(p, config) or _filename_matches_prefix(p, inv_pref) or _is_excluded_movement_file(p, config):
                continue
            files.append(p)
    seen = set(); out = []
    for p in files:
        rp = str(p.resolve())
        if rp not in seen and _file_overlaps_period(p, config):
            seen.add(rp); out.append(p)
    priority = config.get("arquivos_movimentacao_priorizar_contem", [])
    preferred = [p for p in sorted(out) if _filename_contains_any(p, priority)] if priority else []
    return preferred if preferred else sorted(out)

def discover_document_files(config):
    folders = []
    raw = config.get("pasta_documentos", "input/documentos")
    p = Path(raw); folders.append(p if p.is_absolute() else BASE_DIR / p)
    if config.get("permitir_documentos_em_input_raiz", False):
        folders.append(INPUT_DIR)
    files = []
    for folder in folders:
        if folder.exists():
            files.extend(_excel_files_in_dir(folder))
    seen = set(); out = []
    for f in files:
        rp = str(f.resolve())
        if rp not in seen and not _is_manual_factor_file(f, config) and not _is_excluded_movement_file(f, config) and _file_overlaps_period(f, config):
            seen.add(rp); out.append(f)
    return sorted(out)



# =============================================================================
# LEITOR XLSX READ_ONLY - mais rápido para DFE Excel largo
# =============================================================================

def _process_sheet_stream(file_path, sheet_name, config, manual_factors, periodo_inicio, periodo_fim, fingerprints, records, duplicates, logs, wb_open=None):
    """Processa uma aba de um XLSX usando openpyxl read_only.

    Se wb_open for fornecido (workbook já aberto por load_inputs), reutiliza-o sem
    nova abertura de disco - padrão correto quando há múltiplas abas no mesmo arquivo.
    """
    _wb_local = None
    try:
        from openpyxl import load_workbook
        if wb_open is not None:
            ws_obj = wb_open[sheet_name] if sheet_name in wb_open.sheetnames else None
            if ws_obj is None:
                return
            ws = ws_obj
        else:
            _wb_local = load_workbook(file_path, read_only=True, data_only=True)
            ws = _wb_local[sheet_name]
        preview_rows = []
        for row in ws.iter_rows(min_row=1, max_row=HEADER_SCAN, values_only=True):
            preview_rows.append(list(row))
        if not preview_rows:
            logs.append((file_path.name, sheet_name, "NAO_UTILIZADO", "Aba vazia."))
            if _wb_local: _wb_local.close()
            return
        preview = pd.DataFrame(preview_rows)
        hi = find_header_row(preview, config.get("mapeamento_colunas"))
        headers = make_unique_headers(preview.iloc[hi["idx"]].tolist())
        tipo = classify_file_headers(headers)
        if tipo != "ITEM_FISCAL":
            logs.append((file_path.name, sheet_name, "NAO_UTILIZADO", f"Tipo={tipo}"))
            if _wb_local: _wb_local.close()
            return
        mapping = {k: pick_column(headers, k, config.get("mapeamento_colunas")) for k in COLUMN_CANDIDATES}
        score = sum(1 for k in ["chave", "data_emissao", "ncm", "cfop", "codigo_item", "descricao", "unid_comercial", "valor_total_produto"] if mapping.get(k) is not None)
        if score < config.get("score_minimo_colunas_fiscais", 5):
            logs.append((file_path.name, sheet_name, "NAO_UTILIZADO", f"Score={score}"))
            if _wb_local: _wb_local.close()
            return
        ipi_col = mapping.get("valor_ipi") is not None
        before = len(records); out = dup = empty = 0; total = 0
        start_row = hi["idx"] + 2
        for ridx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
            total += 1
            status = process_record_row(row, mapping, config, manual_factors, ipi_col, file_path, sheet_name, periodo_inicio, periodo_fim, fingerprints, records, duplicates, ridx)
            if status == "VAZIA": empty += 1
            elif status == "FORA_PERIODO": out += 1
            elif status == "DUPLICATA": dup += 1
        if _wb_local: _wb_local.close()
        imp = len(records) - before
        logs.append((file_path.name, sheet_name, "PROCESSADA", f"Importadas={imp}; fora_periodo={out}; duplicadas={dup}; vazias={empty}; lidas={total}; Score={score}; IPI={'SIM' if ipi_col else 'NAO'}; Engine=openpyxl_read_only"))
        gc.collect()
    except Exception as e:
        try:
            if _wb_local: _wb_local.close()
        except Exception:
            pass
        logs.append((file_path.name, sheet_name, "ERRO_LEITURA", str(e)))
        logger.exception("Aba %s/%s", file_path.name, sheet_name)



# =============================================================================
# LEITOR DE INVENTARIO READ_ONLY - preserva regra por mes e almoxarifado
# =============================================================================

def competencia_from_sheet_name(name):
    s = clean_str(name).upper().strip()
    m = re.search(r"(0[1-9]|1[0-2])[^0-9]?(20\d{2})", s)
    if m:
        return f"{m.group(1)}/{m.group(2)}"
    m = re.search(r"(20\d{2})[^0-9]?(0[1-9]|1[0-2])", s)
    if m:
        return f"{m.group(2)}/{m.group(1)}"
    meses = {"JAN":"01","FEV":"02","MAR":"03","ABR":"04","MAI":"05","JUN":"06","JUL":"07","AGO":"08","SET":"09","OUT":"10","NOV":"11","DEZ":"12"}
    for nome, mm in meses.items():
        if nome in s:
            y = re.search(r"20\d{2}", s)
            if y:
                return f"{mm}/{y.group(0)}"
    return "N/D"

def resolve_inventory_cnpj(file_path, config):
    fname = norm_text(Path(file_path).name)
    for token, cnpj in (config.get("cnpj_inventario_por_token_arquivo", {}) or {}).items():
        if norm_text(token) in fname:
            return normalize_cnpj(cnpj), f"token={token}"
    direct = re.search(r"\d{14}", clean_str(Path(file_path).name))
    if direct:
        return normalize_cnpj(direct.group(0)), "cnpj_no_nome_arquivo"
    padrao = normalize_cnpj(config.get("cnpj_inventario_padrao"))
    return padrao, "cnpj_inventario_padrao"

def validate_inventory_quantity_values(sample_vals, config):
    vals = [to_decimal(v) for v in sample_vals if clean_str(v) != ""]
    vals = [v for v in vals if isinstance(v, Decimal)]
    if not vals:
        return False, "Amostra de quantidade física vazia ou inválida."
    if all(v == ZERO for v in vals):
        return True, "Amostra de quantidade física zerada; inventário processado para rastreabilidade."
    return True, "Amostra de quantidade física validada."

def load_inventory_pa(config, logs):
    if not config.get("gerar_aba_inventario_pa", True):
        return {}, []
    allowed_almox = ({normalize_almox(a) for a in config.get("almoxarifados_inventario_pa", ["01", "50"]) if normalize_almox(a)}
                     if config.get("filtrar_almoxarifado_inventario_pa", True) else set())
    meses_validos = set(_month_range(parse_config_date(config.get("periodo_base_inicio")), parse_config_date(config.get("periodo_base_fim"))))
    agg = {}
    detail_logs = []
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        detail_logs.append(("-", "INVENTARIO", "ERRO_LEITURA", f"openpyxl indisponivel: {exc}"))
        logs.extend(detail_logs)
        return agg, detail_logs
    for file_path in discover_inventory_files(config):
        cnpj_inv, cnpj_origem = resolve_inventory_cnpj(file_path, config)
        detail_logs.append((file_path.name, "-", "INVENTARIO_CNPJ", f"CNPJ={cnpj_inv or 'NAO_IDENTIFICADO'}; {cnpj_origem}"))
        try:
            wb_ro = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            detail_logs.append((file_path.name, "-", "ERRO_LEITURA", str(e))); continue
        for sh in wb_ro.sheetnames:
            comp = competencia_from_sheet_name(sh)
            if comp == "N/D" or (meses_validos and comp not in meses_validos):
                continue
            try:
                ws = wb_ro[sh]
                preview_rows = [list(row) for row in ws.iter_rows(min_row=1, max_row=HEADER_SCAN, values_only=True)]
                if not preview_rows:
                    continue
                preview = pd.DataFrame(preview_rows)
                hi = find_inventory_header(preview)
                if hi["score"] < 20:
                    detail_logs.append((file_path.name, sh, "NAO_UTILIZADO", f"Score={hi['score']}")); continue
                headers = make_unique_headers(preview.iloc[hi["idx"]].tolist())
                mapping = {k: pick_inventory_column(headers, k) for k in INVENTORY_CANDIDATES}
                qtde_idx = mapping.get("qtde_inventario")
                if qtde_idx is None or not is_physical_quantity_header(headers[qtde_idx]):
                    detail_logs.append((file_path.name, sh, "NAO_UTILIZADO", "Coluna de QTDE fisica nao localizada.")); continue
                # amostra de quantidade
                sample_vals = []
                for ridx, row in enumerate(ws.iter_rows(min_row=hi["idx"]+2, max_row=hi["idx"]+62, values_only=True), start=hi["idx"]+2):
                    if qtde_idx < len(row): sample_vals.append(row[qtde_idx])
                ok_qtde, msg_qtde = validate_inventory_quantity_values(sample_vals, config)
                if not ok_qtde:
                    detail_logs.append((file_path.name, sh, "NAO_UTILIZADO", msg_qtde)); continue
                imp = skipped = 0
                for ridx, row in enumerate(ws.iter_rows(min_row=hi["idx"]+2, values_only=True), start=hi["idx"]+2):
                    cod = row[mapping["codigo_item"]] if mapping.get("codigo_item") is not None and mapping["codigo_item"] < len(row) else ""
                    if not clean_str(cod):
                        continue
                    almox_val = row[mapping["cod_almox"]] if mapping.get("cod_almox") is not None and mapping["cod_almox"] < len(row) else ""
                    if mapping.get("cod_almox") is not None and allowed_almox and normalize_almox(almox_val) not in allowed_almox:
                        skipped += 1; continue
                    qtde = to_decimal(row[qtde_idx] if qtde_idx < len(row) else "")
                    if qtde is None:
                        continue
                    desc = row[mapping["descricao"]] if mapping.get("descricao") is not None and mapping["descricao"] < len(row) else ""
                    unid = row[mapping["unid_comercial"]] if mapping.get("unid_comercial") is not None and mapping["unid_comercial"] < len(row) else ""
                    ncm = row[mapping["ncm"]] if mapping.get("ncm") is not None and mapping["ncm"] < len(row) else ""
                    ok = _add_inventory_record(agg, cod, comp, desc, unid, ncm, almox_val, qtde, f"{file_path.name}|{sh}|{ridx}", cnpj_inv, config)
                    if ok: imp += 1
                detail_logs.append((file_path.name, sh, "INVENTARIO_PROCESSADO", f"Linhas={imp}; ignoradas_almox={skipped}; competencia={comp}; {msg_qtde}"))
            except Exception as e:
                detail_logs.append((file_path.name, sh, "ERRO_LEITURA", str(e)))
        try:
            wb_ro.close()
        except Exception:
            pass
        gc.collect()
    logs.extend(detail_logs)
    return agg, detail_logs



# apply_cnpj_emitente_from_movimento_input foi removida: era dead code (corpo vazio).
# O CNPJ emitente/destinatário é lido diretamente pelas colunas mapeadas em build_record.


def _parse_periodo_operacao_input(valor, tipo_limite):
    """Converte entrada do operador para data de período.

    Aceita DD/MM/AAAA, AAAA-MM-DD e MM/AAAA. Quando informado MM/AAAA,
    data inicial vira primeiro dia do mês e data final vira último dia do mês.
    """
    t = clean_str(valor)
    if not t:
        return None
    d = only_digits(t)
    # MM/AAAA ou MMAAAA
    if len(d) == 6:
        mes = int(d[:2]); ano = int(d[2:6])
        if 1 <= mes <= 12 and 1900 <= ano <= 2100:
            dia = 1 if tipo_limite == "inicio" else calendar.monthrange(ano, mes)[1]
            return datetime(ano, mes, dia)
        # AAAAMM
        ano = int(d[:4]); mes = int(d[4:6])
        if 1 <= mes <= 12 and 1900 <= ano <= 2100:
            dia = 1 if tipo_limite == "inicio" else calendar.monthrange(ano, mes)[1]
            return datetime(ano, mes, dia)
    return parse_date(t)


def _format_prompt_default(dt):
    return f"{dt.day:02d}/{dt.month:02d}/{dt.year:04d}" if isinstance(dt, datetime) else ""


def _perguntar_data_periodo(rotulo, padrao_dt, tipo_limite):
    while True:
        padrao_txt = _format_prompt_default(padrao_dt)
        entrada = input(f"{rotulo} [{padrao_txt}]: ").strip()
        if not entrada:
            return padrao_dt
        dt = _parse_periodo_operacao_input(entrada, tipo_limite)
        if dt:
            return datetime(dt.year, dt.month, dt.day)
        print("Data inválida. Informe DD/MM/AAAA ou MM/AAAA. Ex.: 01/01/2025, 31/12/2025 ou 03/2025.")


def _aplicar_periodo_execucao(cfg, inicio, fim, origem):
    if not inicio or not fim:
        raise ConfigError("Periodo de processamento nao informado ou invalido.")
    if inicio > fim:
        raise ConfigError("Periodo invalido: a data inicial nao pode ser maior que a data final.")
    inicio_txt = f"{inicio.day:02d}/{inicio.month:02d}/{inicio.year:04d}"
    fim_txt = f"{fim.day:02d}/{fim.month:02d}/{fim.year:04d}"
    cfg["data_inicial_calculo"] = inicio_txt
    cfg["data_final_calculo"] = fim_txt
    cfg["periodo_base_inicio"] = inicio_txt
    cfg["periodo_base_fim"] = fim_txt
    cfg["periodo_definido_interativamente"] = origem
    cfg["periodo_origem"] = origem
    print(f"Periodo selecionado: {inicio_txt} ate {fim_txt}")
    return cfg


def solicitar_periodo_operacao(config):
    """Define o periodo da execucao.

    O periodo continua obrigatorio quando perguntar_periodo_ao_iniciar = Sim.
    No fluxo normal, o BAT pergunta ao operador e entrega as variaveis
    RIR70_DATA_INICIAL/RIR70_DATA_FINAL ao motor. O input direto fica apenas
    como fallback para execucao manual do Python.
    """
    cfg = dict(config or {})
    if not _is_sim(cfg.get("perguntar_periodo_ao_iniciar", "Sim")):
        cfg["periodo_definido_interativamente"] = "Nao - desativado no config"
        return cfg

    inicio_padrao = parse_config_date(cfg.get("data_inicial_calculo") or cfg.get("periodo_base_inicio")) or datetime(datetime.now().year, 1, 1)
    fim_padrao = parse_config_date(cfg.get("data_final_calculo") or cfg.get("periodo_base_fim")) or datetime(datetime.now().year, 12, 31)
    env_inicio = clean_str(os.environ.get("RIR70_DATA_INICIAL"))
    env_fim = clean_str(os.environ.get("RIR70_DATA_FINAL"))
    if env_inicio or env_fim:
        inicio = _parse_periodo_operacao_input(env_inicio or _format_prompt_default(inicio_padrao), "inicio")
        fim = _parse_periodo_operacao_input(env_fim or _format_prompt_default(fim_padrao), "fim")
        return _aplicar_periodo_execucao(cfg, inicio, fim, "Sim - informado no BAT")

    try:
        interativo = bool(sys.stdin and sys.stdin.isatty())
    except Exception:
        interativo = False
    if not interativo:
        raise ConfigError("Periodo nao informado. Execute pelo EXECUTAR_ARBITRAMENTO.bat ou informe RIR70_DATA_INICIAL e RIR70_DATA_FINAL.")

    print("PERIODO DE PROCESSAMENTO")
    print("Informe o periodo desejado para esta execucao.")
    print("Formato aceito: DD/MM/AAAA. Tambem e aceito MM/AAAA; nesse caso o motor usa o mes completo.")
    print("Pressione ENTER para manter o padrao exibido.")
    while True:
        inicio = _perguntar_data_periodo("Data inicial", inicio_padrao, "inicio")
        fim = _perguntar_data_periodo("Data final", fim_padrao, "fim")
        if inicio and fim and inicio <= fim:
            return _aplicar_periodo_execucao(cfg, inicio, fim, "Sim - informado no terminal")
        print("Periodo invalido: a data inicial nao pode ser maior que a data final. Informe novamente.")


def main():
    print(f"Sistema RIR70 | Motor de Arbitramento de Custo | versao {SCRIPT_VER}")
    config = ensure_config()
    config = solicitar_periodo_operacao(config)
    validar_governanca_operacional(config)
    ensure_runtime_folders(config)
    if config.get("validar_campos_obrigatorios_matriz", True):
        _validate_matrix_for_mode(config)
    _novo_output_path(config)
    logging.basicConfig(filename=str(LOG_PATH), level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s", encoding="utf-8")
    inicio = parse_config_date(config.get("periodo_base_inicio")); fim = parse_config_date(config.get("periodo_base_fim"))
    if inicio and fim and inicio > fim:
        raise ValueError("periodo_base_inicio > periodo_base_fim no config.")
    if not config.get("CFOP_VENDA_CALCULO"):
        raise ValueError("CFOP_VENDA_CALCULO vazio no config.")
    logger.info("Início. Config=%s", safe_json_dumps(config))
    records, logs, duplicates = load_inputs(config)
    print(f"Linhas importadas: {len(records)}")
    records = enrich_records(records, config)
    n_part = sum(1 for r in records if r.get("_participa"))
    print(f"Participam do cálculo: {n_part} / {len(records)}")
    write_workbook(records, logs, config)
    write_metadata(records, config)
    # Arquivo de controle para o BAT abrir o Excel final automaticamente.
    (OUTPUT_DIR / "ultimo_arquivo_gerado.txt").write_text(str(OUTPUT_PATH), encoding="utf-8")
    print(f"Concluído. Arquivo: {OUTPUT_PATH}")
    logging.shutdown()
    # Abre o Excel automaticamente se configurado.
    if config.get("abrir_excel_ao_final", True) and os.environ.get("RIR70_EXECUTADO_PELO_BAT") != "1":
        try:
            import subprocess, platform
            if platform.system() == "Windows":
                os.startfile(str(OUTPUT_PATH))
            elif platform.system() == "Darwin":
                subprocess.Popen(["open", str(OUTPUT_PATH)])
            else:
                subprocess.Popen(["xdg-open", str(OUTPUT_PATH)])
        except Exception as exc:
            print(f"Aviso: não foi possível abrir o Excel automaticamente: {exc}")
    # Fecha o terminal/CMD automaticamente se configurado.
    if config.get("fechar_cmd_ao_final", False):
        try:
            import subprocess, platform
            if platform.system() == "Windows":
                subprocess.Popen("exit", shell=True)
        except Exception:
            pass


if __name__=="__main__":
    try: main()
    except Exception as exc:
        logger.exception("ERRO"); print(f"ERRO: {exc}\nLog: {LOG_PATH}"); sys.exit(1)

